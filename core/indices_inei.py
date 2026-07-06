# SPDX-License-Identifier: GPL-3.0-or-later
# Copyright (C) 2026 Marco Sumari / Sumari SAC
# This file is part of IngePresupuestos — https://ingepresupuestos.com
# Licensed under the GNU GPL v3.0 or later. See the LICENSE file.
"""core.indices_inei — gestión del histórico de Índices Unificados de Precios (INEI).

Tablas:
    - ``indices_inei``         : catálogo de los 80 códigos
    - ``indices_inei_areas``   : 6 áreas geográficas estándar
    - ``indices_inei_valores`` : serie histórica (codigo · año · mes · área · valor)

Soporta importación masiva desde Excel publicado por el INEI cada mes y
exportación / importación de la serie completa en JSON para sincronización
entre instalaciones.

Espejo conceptual de la funcionalidad "Importación de Índices de Precios
INEI 2026" de Delphin Express.
"""
from __future__ import annotations

import json
from pathlib import Path

from core.database import get_db


# ── Catálogo de 80 índices unificados de precios INEI ────────────────────────
CATALOGO_INEI: list[tuple[str, str]] = [
    ("01", "Aceite"),
    ("02", "Acero de construcción liso"),
    ("03", "Acero de construcción corrugado"),
    ("04", "Agregado fino"),
    ("05", "Agregado grueso"),
    ("06", "Alambre y cable de cobre desnudo"),
    ("07", "Alambre y cable tipo TW y THW"),
    ("08", "Alambre y cable tipo WP"),
    ("09", "Alcantarilla metálica"),
    ("10", "Aparato sanitario con grifería"),
    ("11", "Artefacto de alumbrado exterior"),
    ("12", "Artefacto de alumbrado interior"),
    ("13", "Asfalto"),
    ("14", "Baldosa acústica"),
    ("15", "Baldosa asfáltica"),
    ("16", "Baldosa vinílica"),
    ("17", "Bloque y ladrillo"),
    ("18", "Cable telefónico"),
    ("19", "Cable NYY-N2XY"),
    ("20", "Cemento asfáltico"),
    ("21", "Cemento Portland tipo I"),
    ("22", "Cemento Portland tipo II"),
    ("23", "Cemento Portland tipo V"),
    ("24", "Cerámica esmaltada y sin esmaltar"),
    ("26", "Cerrajería nacional"),
    ("27", "Detonante"),
    ("28", "Dinamita"),
    ("29", "Dólar"),
    ("30", "Dólar más inflación USA / General ponderado"),
    ("31", "Ducto de concreto"),
    ("32", "Flete terrestre"),
    ("33", "Flete aéreo"),
    ("34", "Gasolina"),
    ("37", "Herramienta manual"),
    ("38", "Hormigón"),
    ("39", "Índice general de precios al consumidor (IPC)"),
    ("40", "Loseta"),
    ("41", "Madera en tiras para piso"),
    ("42", "Madera importada para encofrado y carpintería"),
    ("43", "Madera nacional para encofrado y carpintería"),
    ("44", "Madera terciada para encofrado y carpintería"),
    ("45", "Madera terciada para encofrado"),
    ("46", "Malla de acero"),
    ("47", "Mano de obra (incluido leyes sociales)"),
    ("48", "Maquinaria y equipo nacional"),
    ("49", "Maquinaria y equipo importado"),
    ("50", "Marco y tapa de hierro fundido"),
    ("51", "Perfil de acero liviano"),
    ("52", "Perfil de aluminio"),
    ("53", "Petróleo diesel"),
    ("54", "Pintura látex"),
    ("55", "Pintura temple"),
    ("56", "Plancha de Aero LAC"),
    ("57", "Plancha de Aero LAF"),
    ("59", "Plancha de fibro-cemento"),
    ("60", "Plancha de poliuretano"),
    ("61", "Plancha galvanizada"),
    ("62", "Poste de concreto"),
    ("64", "Terrazo"),
    ("65", "Tubería de acero negro y/o galvanizado"),
    ("66", "Tubería de PVC para agua potable y alcantarillado"),
    ("68", "Tubería de cobre"),
    ("69", "Tubería de concreto simple"),
    ("70", "Tubería de concreto reforzado"),
    ("71", "Tubería de fierro fundido"),
    ("72", "Tubería de PVC para agua"),
    ("73", "Ducto telefónico de PVC"),
    ("74", "Tubería de PVC para electricidad (SAP)"),
    ("77", "Válvula de bronce nacional"),
    ("78", "Válvula de fierro fundido nacional"),
    ("79", "Vidrio incoloro nacional"),
    ("80", "Concreto premezclado"),
]


# ── 6 Áreas geográficas estándar INEI (IUP) ──────────────────────────────────
AREAS_INEI: list[tuple[str, str]] = [
    ("01", "Lima Metropolitana y Callao"),
    ("02", "Norte (Tumbes/Piura/Lambayeque/La Libertad/Áncash)"),
    ("03", "Centro (Lima Provincias/Junín/Pasco/Huánuco)"),
    ("04", "Sur Medio y Selva (Ica/Ayacucho/Huancavelica/Amazonas/San Martín/Loreto)"),
    ("05", "Sur (Arequipa/Moquegua/Tacna/Apurímac/Cusco/Madre de Dios/Puno)"),
    ("06", "Nacional (promedio ponderado)"),
]


# ── Seed (idempotente) ───────────────────────────────────────────────────────
def asegurar_seed(conn=None) -> None:
    """Asegura que el catálogo y las áreas estén poblados. Idempotente."""
    own = conn is None
    if own:
        conn = get_db()
    try:
        for codigo, nombre in CATALOGO_INEI:
            conn.execute(
                "INSERT OR IGNORE INTO indices_inei (codigo, nombre, activo) "
                "VALUES (?, ?, 1)",
                (codigo, nombre)
            )
        for i, (codigo, nombre) in enumerate(AREAS_INEI):
            conn.execute(
                "INSERT OR IGNORE INTO indices_inei_areas "
                "(codigo, nombre, orden) VALUES (?, ?, ?)",
                (codigo, nombre, i)
            )
        conn.commit()
    finally:
        if own:
            conn.close()


# ── Listados ─────────────────────────────────────────────────────────────────
def listar_indices(conn=None) -> list[dict]:
    """Devuelve los 80 códigos con el último valor cargado (de cualquier área)."""
    own = conn is None
    if own:
        conn = get_db()
    asegurar_seed(conn)
    rows = conn.execute(
        """SELECT i.codigo, i.nombre, i.activo,
                  (SELECT COUNT(DISTINCT anio || '-' || mes || '-' || area)
                   FROM indices_inei_valores v WHERE v.codigo = i.codigo)
                  AS n_valores,
                  (SELECT anio || '-' || PRINTF('%02d', mes)
                   FROM indices_inei_valores v WHERE v.codigo = i.codigo
                   ORDER BY anio DESC, mes DESC LIMIT 1)
                  AS ultimo_periodo,
                  (SELECT valor FROM indices_inei_valores v
                   WHERE v.codigo = i.codigo
                   ORDER BY anio DESC, mes DESC LIMIT 1)
                  AS ultimo_valor
           FROM indices_inei i ORDER BY i.codigo"""
    ).fetchall()
    if own:
        conn.close()
    return [dict(r) for r in rows]


def listar_areas(conn=None) -> list[dict]:
    """Devuelve la lista de áreas geográficas."""
    own = conn is None
    if own:
        conn = get_db()
    asegurar_seed(conn)
    rows = conn.execute(
        "SELECT codigo, nombre, orden FROM indices_inei_areas ORDER BY orden"
    ).fetchall()
    if own:
        conn.close()
    return [dict(r) for r in rows]


def obtener_valor(codigo: str, anio: int, mes: int,
                  area: str = '01') -> float | None:
    """Devuelve el valor del índice para (codigo, anio, mes, area) o None."""
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT valor FROM indices_inei_valores "
            "WHERE codigo=? AND anio=? AND mes=? AND area=?",
            (codigo, anio, mes, area)
        ).fetchone()
        return row['valor'] if row else None
    finally:
        conn.close()


def obtener_matriz(codigo: str, area: str = '01') -> dict[int, dict[int, float]]:
    """Devuelve un dict {anio: {mes: valor}} con toda la serie del índice."""
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT anio, mes, valor FROM indices_inei_valores "
            "WHERE codigo=? AND area=? ORDER BY anio, mes",
            (codigo, area)
        ).fetchall()
    finally:
        conn.close()
    out: dict[int, dict[int, float]] = {}
    for r in rows:
        out.setdefault(r['anio'], {})[r['mes']] = r['valor']
    return out


# ── Persistencia ─────────────────────────────────────────────────────────────
def guardar_valor(codigo: str, anio: int, mes: int, valor: float,
                  area: str = '01') -> None:
    """Inserta o reemplaza un valor."""
    conn = get_db()
    try:
        conn.execute(
            "INSERT OR REPLACE INTO indices_inei_valores "
            "(codigo, anio, mes, area, valor) VALUES (?,?,?,?,?)",
            (codigo, int(anio), int(mes), area, float(valor))
        )
        conn.commit()
    finally:
        conn.close()


def guardar_valores(rows: list[dict]) -> tuple[int, int]:
    """Batch upsert. ``rows`` es lista de dicts con codigo/anio/mes/area/valor.
    Retorna (n_insertados_o_actualizados, n_ignorados_por_error)."""
    conn = get_db()
    ok = 0
    err = 0
    try:
        for r in rows:
            try:
                codigo = str(r.get('codigo') or '').strip().zfill(2)[:2]
                anio = int(r.get('anio') or 0)
                mes = int(r.get('mes') or 0)
                area = str(r.get('area') or '01')
                valor = float(r.get('valor') or 0)
                if not codigo or anio < 1900 or not (1 <= mes <= 12) or valor <= 0:
                    err += 1
                    continue
                conn.execute(
                    "INSERT OR REPLACE INTO indices_inei_valores "
                    "(codigo, anio, mes, area, valor) VALUES (?,?,?,?,?)",
                    (codigo, anio, mes, area, valor)
                )
                ok += 1
            except Exception:
                err += 1
        conn.commit()
    finally:
        conn.close()
    return ok, err


def eliminar_valor(codigo: str, anio: int, mes: int, area: str = '01') -> None:
    conn = get_db()
    try:
        conn.execute(
            "DELETE FROM indices_inei_valores "
            "WHERE codigo=? AND anio=? AND mes=? AND area=?",
            (codigo, anio, mes, area)
        )
        conn.commit()
    finally:
        conn.close()


# ── Importación / Exportación ────────────────────────────────────────────────
MESES_MAP = {
    'enero': 1, 'ene': 1, 'jan': 1, 'january': 1,
    'febrero': 2, 'feb': 2, 'february': 2,
    'marzo': 3, 'mar': 3, 'march': 3,
    'abril': 4, 'abr': 4, 'apr': 4, 'april': 4,
    'mayo': 5, 'may': 5,
    'junio': 6, 'jun': 6, 'june': 6,
    'julio': 7, 'jul': 7, 'july': 7,
    'agosto': 8, 'ago': 8, 'aug': 8, 'august': 8,
    'septiembre': 9, 'setiembre': 9, 'sep': 9, 'sept': 9, 'september': 9,
    'octubre': 10, 'oct': 10, 'october': 10,
    'noviembre': 11, 'nov': 11, 'november': 11,
    'diciembre': 12, 'dic': 12, 'dec': 12, 'december': 12,
}


def _parse_mes(texto) -> int | None:
    """Resuelve un encabezado de columna a número de mes (1-12).

    Acepta números (1, 01, 1.0), nombres ('Ene', 'Enero', 'JAN') con o sin
    tildes y acentos. Retorna None si no se puede resolver.
    """
    if texto is None:
        return None
    s = str(texto).strip().lower()
    if not s:
        return None
    s = (s.replace('á', 'a').replace('é', 'e').replace('í', 'i')
           .replace('ó', 'o').replace('ú', 'u'))
    if s in MESES_MAP:
        return MESES_MAP[s]
    try:
        v = int(float(s))
        if 1 <= v <= 12:
            return v
    except Exception:
        pass
    return None


def importar_excel_inei(filepath: str, area: str = '01',
                        anio_override: int | None = None) -> dict:
    """Importa valores desde un Excel publicado por INEI.

    Detecta automáticamente la orientación (índices en filas o columnas) y
    parsea encabezados de mes en español/inglés. Asume todos los valores son
    del mismo año (se infiere de un header tipo "Enero 2026" o se usa
    ``anio_override``).

    Retorna dict::

        {'ok': bool, 'msg': str, 'rows': [...], 'ignorados': int,
         'anio_detectado': int|None, 'codigos_encontrados': set}
    """
    try:
        import openpyxl
    except ImportError:
        return {'ok': False, 'msg': "openpyxl no instalado (pip install openpyxl)",
                'rows': [], 'ignorados': 0}

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return {'ok': False, 'msg': f"No se pudo abrir: {e}",
                'rows': [], 'ignorados': 0}

    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return {'ok': False, 'msg': "El archivo está vacío.",
                'rows': [], 'ignorados': 0}

    # Detectar año: buscar un texto tipo "ENERO 2026" o "2026"
    anio_detectado = anio_override
    if not anio_detectado:
        import re
        for row in data[:20]:
            for cell in row:
                if cell:
                    m = re.search(r'\b(20\d{2})\b', str(cell))
                    if m:
                        anio_detectado = int(m.group(1))
                        break
            if anio_detectado:
                break

    # Buscar fila de encabezado: la primera con al menos 6 valores que sean
    # meses parseables (o números 1..12)
    header_row_idx = -1
    mes_cols: dict[int, int] = {}   # idx_col → mes (1..12)
    codigo_col = -1
    for i, row in enumerate(data[:30]):
        mes_temp: dict[int, int] = {}
        codigo_temp = -1
        for j, val in enumerate(row):
            mes = _parse_mes(val)
            if mes:
                mes_temp[j] = mes
            elif val and 'codigo' in str(val).lower():
                codigo_temp = j
            elif val and ('indice' in str(val).lower() or 'iu' in str(val).lower()):
                if codigo_temp < 0:
                    codigo_temp = j
        if len(mes_temp) >= 6:
            header_row_idx = i
            mes_cols = mes_temp
            codigo_col = codigo_temp if codigo_temp >= 0 else 0
            break

    if header_row_idx < 0 or not mes_cols:
        return {
            'ok': False,
            'msg': ("No se encontró una tabla con encabezados de mes "
                    "(Ene/Feb/Mar… o 1/2/3…)."),
            'rows': [], 'ignorados': 0,
        }

    rows_out: list[dict] = []
    ignorados = 0
    codigos_encontrados: set[str] = set()
    for row in data[header_row_idx + 1:]:
        if not row:
            continue
        # Código en col codigo_col, soportar valores tipo '01', 1, '1.0', '01 - Aceite'
        cod_raw = row[codigo_col] if codigo_col < len(row) else None
        if cod_raw is None:
            continue
        cod_str = str(cod_raw).strip()
        # Extraer dígitos iniciales
        import re
        m = re.match(r'^(\d{1,2})', cod_str)
        if not m:
            continue
        codigo = m.group(1).zfill(2)
        if not (1 <= int(codigo) <= 99):
            continue

        for col_idx, mes in mes_cols.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or val == '':
                continue
            try:
                f = float(val)
                if f <= 0:
                    ignorados += 1
                    continue
            except Exception:
                ignorados += 1
                continue
            if not anio_detectado:
                ignorados += 1
                continue
            rows_out.append({
                'codigo': codigo, 'anio': anio_detectado,
                'mes': mes, 'area': area, 'valor': f,
            })
            codigos_encontrados.add(codigo)

    return {
        'ok': True,
        'msg': f"OK — {len(rows_out)} valores listos para importar.",
        'rows': rows_out,
        'ignorados': ignorados,
        'anio_detectado': anio_detectado,
        'codigos_encontrados': codigos_encontrados,
    }


def exportar_json(filepath: str, area: str | None = None) -> int:
    """Exporta toda la serie a JSON. Si ``area`` no es None, filtra por ella.
    Retorna el número de valores exportados."""
    conn = get_db()
    try:
        q = ("SELECT codigo, anio, mes, area, valor "
             "FROM indices_inei_valores")
        p = []
        if area:
            q += " WHERE area=?"; p.append(area)
        q += " ORDER BY codigo, anio, mes, area"
        rows = [dict(r) for r in conn.execute(q, p).fetchall()]
    finally:
        conn.close()
    payload = {
        'version': 1,
        'tipo': 'indices_inei',
        'area_filtro': area,
        'valores': rows,
    }
    Path(filepath).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding='utf-8'
    )
    return len(rows)


def importar_json(filepath: str) -> dict:
    """Importa valores desde JSON exportado por ``exportar_json``."""
    try:
        payload = json.loads(Path(filepath).read_text(encoding='utf-8'))
    except Exception as e:
        return {'ok': False, 'msg': f"Archivo inválido: {e}", 'n_ok': 0, 'n_err': 0}
    if payload.get('tipo') != 'indices_inei':
        return {'ok': False, 'msg': "El archivo no es de tipo indices_inei.",
                'n_ok': 0, 'n_err': 0}
    rows = payload.get('valores') or []
    ok, err = guardar_valores(rows)
    return {'ok': True, 'msg': f"{ok} valores importados, {err} ignorados",
            'n_ok': ok, 'n_err': err}


# ─── Descarga por URL ────────────────────────────────────────────────────────
def descargar_desde_url(url: str, area: str = '01',
                        anio_override: int | None = None) -> dict:
    """Descarga un Excel desde una URL pública y lo parsea con
    ``importar_excel_inei``. Útil cuando el usuario tiene el link del INEI.

    Acepta solo http/https. Timeout de 30s. Tamaño máximo: 20 MB.
    Retorna el mismo dict que ``importar_excel_inei`` más:
        - 'url': URL fuente
        - 'tamano_kb': tamaño del archivo descargado
    """
    import tempfile
    import urllib.request
    import urllib.error

    if not url:
        return {'ok': False, 'msg': "URL vacía.", 'rows': [], 'ignorados': 0}
    if not (url.startswith("http://") or url.startswith("https://")):
        return {'ok': False, 'msg': "La URL debe empezar con http:// o https://",
                'rows': [], 'ignorados': 0}

    # Headers que muchos servidores requieren para no bloquearnos
    req = urllib.request.Request(url, headers={
        'User-Agent': ('Mozilla/5.0 (X11; Linux x86_64) ingePresupuestos/1.0'),
        'Accept': '*/*',
    })

    tmp_path: str | None = None
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            ct = resp.headers.get('Content-Type', '').lower()
            cl = resp.headers.get('Content-Length')
            if cl and int(cl) > 20 * 1024 * 1024:
                return {'ok': False,
                        'msg': f"Archivo demasiado grande ({int(cl)//1024} KB).",
                        'rows': [], 'ignorados': 0}

            # Si la URL apunta a HTML, lo más probable es que sea una página y
            # no el Excel directo
            if 'html' in ct and not url.lower().endswith(('.xlsx', '.xls')):
                return {
                    'ok': False,
                    'msg': ("La URL devuelve HTML, no un Excel. Abre la "
                            "página en el navegador, haz clic derecho sobre "
                            "el enlace del Excel y elige «Copiar dirección "
                            "del enlace», luego pégala aquí."),
                    'rows': [], 'ignorados': 0,
                }

            data = resp.read()

        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.write(data)
        tmp.close()
        tmp_path = tmp.name

        res = importar_excel_inei(tmp_path, area=area,
                                  anio_override=anio_override)
        res['url'] = url
        res['tamano_kb'] = round(len(data) / 1024, 1)
        return res

    except urllib.error.HTTPError as e:
        return {'ok': False, 'msg': f"Error HTTP {e.code} al descargar.",
                'rows': [], 'ignorados': 0}
    except urllib.error.URLError as e:
        return {'ok': False, 'msg': f"Error de conexión: {e.reason}",
                'rows': [], 'ignorados': 0}
    except Exception as e:
        return {'ok': False, 'msg': f"Error inesperado: {e}",
                'rows': [], 'ignorados': 0}
    finally:
        if tmp_path:
            try:
                import os
                os.unlink(tmp_path)
            except Exception:
                pass


# ─── Pegar desde portapapeles ────────────────────────────────────────────────
def importar_desde_texto(texto: str, area: str = '01',
                          anio_override: int | None = None) -> dict:
    """Parsea contenido tabular (CSV, TSV o tabla pegada desde Excel/web/PDF).

    Soporta los mismos formatos que ``importar_excel_inei`` pero recibe el
    contenido como texto crudo. Detecta separador automáticamente
    (tab, coma, punto y coma, múltiples espacios).
    """
    import csv
    import io
    import re
    import tempfile

    if not texto or not texto.strip():
        return {'ok': False, 'msg': "Nada para pegar.",
                'rows': [], 'ignorados': 0}

    # Detectar separador por mayoría en las primeras 5 líneas
    primeras = "\n".join(texto.splitlines()[:5])
    counts = {
        '\t': primeras.count('\t'),
        ';':  primeras.count(';'),
        ',':  primeras.count(','),
    }
    sep = max(counts, key=counts.get)
    if counts[sep] == 0:
        # Sin separador estándar, tratar de detectar espacios múltiples
        sep_re = re.compile(r' {2,}|\t')
        rows = [sep_re.split(line.strip()) for line in texto.splitlines()
                if line.strip()]
    else:
        rows = list(csv.reader(io.StringIO(texto), delimiter=sep))

    if not rows:
        return {'ok': False, 'msg': "No se pudieron parsear filas.",
                'rows': [], 'ignorados': 0}

    # Escribimos un Excel temporal en memoria para reutilizar el parser robusto
    try:
        import openpyxl
    except ImportError:
        return {'ok': False, 'msg': "openpyxl no instalado.",
                'rows': [], 'ignorados': 0}

    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append([c.strip() if isinstance(c, str) else c for c in r])

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    try:
        wb.save(tmp.name)
        res = importar_excel_inei(tmp.name, area=area,
                                  anio_override=anio_override)
        res['filas_pegadas'] = len(rows)
        return res
    finally:
        try:
            import os
            os.unlink(tmp.name)
        except Exception:
            pass


# ─── Auto-detección del último archivo INEI ──────────────────────────────────
INEI_BASE = "https://www.inei.gob.pe/media/MenuRecursivo/indices_tematicos"
INEI_PATTERN = "06_indices_unificados_de_precios_de_la_construccion_{mes}{ano}.xlsx"
INEI_PATTERN_NUEVA = "07_indices_unificados_de_precios_de_la_construccion_1.xlsx"

_MESES_CORTOS_INEI = ['ene', 'feb', 'mar', 'abr', 'may', 'jun',
                      'jul', 'ago', 'set', 'oct', 'nov', 'dic']


def buscar_ultimo_excel_inei() -> dict:
    """Busca por HEAD el último Excel publicado por el INEI siguiendo el
    patrón estable de URLs ``06_..._{mes_corto}{ano_2dig}.xlsx``.

    Estrategia: empezar desde el mes actual y retroceder hasta encontrar uno
    que responda 200. INEI publica datos del mes anterior alrededor del día 15.

    Retorna::

        {'ok': bool, 'url': str|None, 'msg': str, 'mes_detectado': str|None,
         'anio_detectado': int|None}
    """
    import urllib.request
    from datetime import date

    hoy = date.today()
    # Probar desde el mes actual hasta 18 meses atrás
    candidatos: list[tuple[int, int]] = []
    y, m = hoy.year, hoy.month
    for _ in range(18):
        candidatos.append((y, m))
        m -= 1
        if m < 1:
            m = 12
            y -= 1

    headers = {
        'User-Agent': ('Mozilla/5.0 (X11; Linux x86_64) '
                       'ingePresupuestos/1.0'),
    }
    for y, m in candidatos:
        mes_str = _MESES_CORTOS_INEI[m - 1]
        ano_str = str(y % 100).zfill(2)
        fname = INEI_PATTERN.format(mes=mes_str, ano=ano_str)
        url = f"{INEI_BASE}/{fname}"
        try:
            req = urllib.request.Request(url, method='HEAD', headers=headers)
            with urllib.request.urlopen(req, timeout=6) as resp:
                if resp.status == 200:
                    return {
                        'ok': True,
                        'url': url,
                        'msg': f"Último archivo INEI: {mes_str.title()} {y}",
                        'mes_detectado': mes_str,
                        'anio_detectado': y,
                    }
        except Exception:
            continue

    return {
        'ok': False,
        'url': None,
        'msg': ("No se encontró ningún archivo INEI con el patrón conocido en "
                "los últimos 18 meses. Tu conexión puede estar bloqueada o el "
                "INEI cambió el formato de URL."),
        'mes_detectado': None,
        'anio_detectado': None,
    }


def descargar_ultimo_inei(area: str = '01') -> dict:
    """Conveniencia: busca el último Excel disponible en el INEI y lo importa.

    Retorna el dict de ``descargar_desde_url`` enriquecido con ``url`` y
    ``mes/anio_detectado`` si la búsqueda fue exitosa.
    """
    busq = buscar_ultimo_excel_inei()
    if not busq['ok']:
        return {'ok': False, 'msg': busq['msg'], 'rows': [], 'ignorados': 0,
                'url': None}
    res = descargar_desde_url(busq['url'], area=area,
                              anio_override=busq['anio_detectado'])
    res['mes_detectado'] = busq['mes_detectado']
    res['anio_detectado_url'] = busq['anio_detectado']
    return res
