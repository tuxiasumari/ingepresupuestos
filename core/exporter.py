# SPDX-License-Identifier: GPL-3.0-or-later
# Copyright (C) 2026 Marco Sumari / Sumari SAC
# This file is part of IngePresupuestos — https://ingepresupuestos.com
# Licensed under the GNU GPL v3.0 or later. See the LICENSE file.
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.worksheet.page import PageMargins
import os
import io
from core.database import get_db, calcular_totales, get_acu_items, get_decimales_metrado

# Fuente Inter (variable) empaquetada con la app — un solo .ttf que
# contiene todos los pesos y variantes (italic en su propio archivo).
# Path original (Ubuntu en `static/fonts/ubuntu/`) era código muerto
# — esos archivos nunca existieron en el repo. Migrado a Inter para
# coherencia con el resto de reportes (PDF/Word/Excel).
_FONTS_DIR = os.path.join(os.path.dirname(__file__), '..', 'resources', 'fonts')
_FONT_R  = os.path.abspath(os.path.join(_FONTS_DIR, 'Inter.ttf'))
_FONT_B  = _FONT_R          # variable: el mismo file maneja Bold
_FONT_RI = os.path.abspath(os.path.join(_FONTS_DIR, 'Inter-Italic.ttf'))
_FONT_BI = _FONT_RI         # variable: el mismo file maneja Bold Italic


# ─── Paleta del sistema de diseño (espejo del PDF) ───────────────────────────

C_ORANGE      = 'F37329'
C_ORANGE_DARK = 'C0621A'
C_ORANGE_SOFT = 'FEF5EB'
C_SLATE_900   = '1F2A38'
C_SLATE_700   = '2E3C52'
C_SLATE_500   = '485A6C'
C_SLATE_100   = 'E2E8F0'


# ─── Helper: altura de fila basada en longitud de texto ──────────────────────

def _alto_fila(texto: str, chars_por_linea: int) -> float:
    """Altura mínima en pt para que un texto envuelto sea visible
    completamente. ~15pt por línea con Inter 11."""
    n = max(1, -(-len(texto or '') // chars_por_linea))
    return max(18.0, n * 15.0)


def _card_border_factory(n_cols: int):
    """Devuelve un closure `_card_border(col, top=, bottom=)` que aplica
    bordes laterales slate-700 cuando col es el extremo izquierdo o derecho
    de la card (col=1 o col=n_cols). Espejo del patrón de ACUs."""
    side_card = Side(style='medium', color=C_SLATE_700)
    def _border(col, *, top=None, bottom=None):
        sides = {}
        if top is not None:    sides['top']    = top
        if bottom is not None: sides['bottom'] = bottom
        if col == 1:           sides['left']   = side_card
        if col == n_cols:      sides['right']  = side_card
        return Border(**sides)
    return _border, side_card


# ─── Bordes Excel ─────────────────────────────────────────────────────────────

def thin_border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

def _borde_titulo():
    """Borde con línea gruesa arriba y abajo — para títulos nivel 1."""
    m = Side(style='medium'); t = Side(style='thin')
    return Border(left=t, right=t, top=m, bottom=m)

def _borde_seccion():
    """Borde con línea media abajo — para cabeceras de columna y totales."""
    t = Side(style='thin'); m = Side(style='medium')
    return Border(left=t, right=t, top=t, bottom=m)

def _borde_total():
    """Borde con línea media arriba — para filas de total/resumen."""
    t = Side(style='thin'); m = Side(style='medium')
    return Border(left=t, right=t, top=m, bottom=t)


# ─── Firmantes (legacy, tabla pie_presupuesto) ────────────────────────────────

def _get_pie(conn, proyecto_id):
    rows = conn.execute(
        "SELECT * FROM pie_presupuesto WHERE proyecto_id=?", (proyecto_id,)
    ).fetchall()
    return {r['rol']: dict(r) for r in rows}


def _escribir_pie(ws, r, pie, n_cols=None):
    roles = [
        ('elaborado',   'ELABORADO POR'),
        ('revisado',    'REVISADO POR'),
        ('aprobado',    'APROBADO POR'),
        ('visto_bueno', 'VISTO BUENO'),
    ]
    firmantes = [(t, pie.get(rol, {})) for rol, t in roles
                 if pie.get(rol, {}).get('nombre') or pie.get(rol, {}).get('cargo')]
    if not firmantes:
        return
    thin = Side(style='thin')
    r += 2
    # Distribuir firmantes dentro del ancho útil (n_cols) si se especifica
    if n_cols and len(firmantes):
        cols_per = max(1, n_cols // len(firmantes))
    else:
        cols_per = 3
    for i, (titulo, f) in enumerate(firmantes):
        c = 1 + i * cols_per; ce = c + cols_per - 1
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=ce)
        cell = ws.cell(r, c, titulo)
        cell.font = Font(name='Inter',bold=True, size=11)
        cell.alignment = Alignment(horizontal='center')
    r += 1
    for i, (_, f) in enumerate(firmantes):
        c = 1 + i * cols_per; ce = c + cols_per - 1
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=ce)
        ws.cell(r, c).border = Border(bottom=thin)
    r += 1
    for i, (_, f) in enumerate(firmantes):
        c = 1 + i * cols_per; ce = c + cols_per - 1
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=ce)
        cell = ws.cell(r, c, f.get('nombre', ''))
        cell.font = Font(name='Inter',bold=True, size=11)
        cell.alignment = Alignment(horizontal='center')
    r += 1
    for i, (_, f) in enumerate(firmantes):
        c = 1 + i * cols_per; ce = c + cols_per - 1
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=ce)
        cell = ws.cell(r, c, f.get('cargo', ''))
        cell.font = Font(name='Inter', size=11, italic=True)
        cell.alignment = Alignment(horizontal='center')
    r += 1
    for i, (_, f) in enumerate(firmantes):
        if f.get('cip'):
            c = 1 + i * cols_per; ce = c + cols_per - 1
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=ce)
            cell = ws.cell(r, c, f"CIP Nº {f['cip']}")
            cell.font = Font(name='Inter', size=11)
            cell.alignment = Alignment(horizontal='center')


# ─── Monto en letras (español) ────────────────────────────────────────────────

_UNIDADES = ['', 'UN', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE', 'OCHO', 'NUEVE',
             'DIEZ', 'ONCE', 'DOCE', 'TRECE', 'CATORCE', 'QUINCE', 'DIECISÉIS', 'DIECISIETE',
             'DIECIOCHO', 'DIECINUEVE']
_DECENAS  = ['', 'DIEZ', 'VEINTE', 'TREINTA', 'CUARENTA', 'CINCUENTA',
             'SESENTA', 'SETENTA', 'OCHENTA', 'NOVENTA']
_CENTENAS = ['', 'CIENTO', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS', 'QUINIENTOS',
             'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS', 'NOVECIENTOS']


def _n2l(n):
    if n == 0:
        return 'CERO'
    if n < 0:
        return 'MENOS ' + _n2l(-n)
    p = []
    if n >= 1_000_000:
        m = n // 1_000_000; n %= 1_000_000
        p.append('UN MILLÓN' if m == 1 else _n2l(m) + ' MILLONES')
    if n >= 1_000:
        k = n // 1_000; n %= 1_000
        p.append('MIL' if k == 1 else _n2l(k) + ' MIL')
    if n >= 100:
        c = n // 100; n %= 100
        p.append('CIEN' if (c == 1 and n == 0) else _CENTENAS[c])
    if n >= 20:
        d, u = n // 10, n % 10
        p.append(_DECENAS[d] if u == 0 else _DECENAS[d] + ' Y ' + _UNIDADES[u])
    elif n > 0:
        p.append(_UNIDADES[n])
    return ' '.join(p)


def _monto_letras(monto, moneda='SOLES'):
    try:
        monto   = round(float(monto), 2)
        entero  = int(monto)
        cts     = round((monto - entero) * 100)
        return f"{_n2l(entero)} CON {cts:02d}/100 {moneda}"
    except Exception:
        return ''


# ─── Cálculo dinámico de rubros del pie ───────────────────────────────────────

def _calcular_rubros_pie(conn, proyecto_id, cd):
    # None = el proyecto nunca ha tenido pie configurado → usar fallback de porcentajes
    # []   = pie configurado pero sin rubros activos → mostrar solo Costo Directo
    tiene_pie = conn.execute(
        "SELECT 1 FROM pie_rubros WHERE proyecto_id=? LIMIT 1", (proyecto_id,)
    ).fetchone()
    if not tiene_pie:
        return None, cd

    rubros = conn.execute(
        "SELECT * FROM pie_rubros WHERE proyecto_id=? AND activo=1 ORDER BY orden",
        (proyecto_id,)
    ).fetchall()
    if not rubros:
        return [], cd

    result   = []
    acum     = cd
    last_sub = cd

    for r in rubros:
        tipo   = r['tipo']
        nombre = r['nombre']
        pct    = r['pct'] or 0
        codigo = r['codigo']

        mp = r['mostrar_pct'] if r['mostrar_pct'] is not None else 1
        if tipo == 'subtotal':
            last_sub = acum
            result.append({'tipo': tipo, 'nombre': nombre, 'valor': acum, 'codigo': codigo, 'pct': pct, 'mostrar_pct': mp})
        elif tipo == 'pct_sub':
            val   = last_sub * pct / 100
            acum += val
            result.append({'tipo': tipo, 'nombre': nombre, 'valor': val, 'codigo': codigo, 'pct': pct, 'mostrar_pct': mp})
        else:
            has_items = False
            if tipo == 'rubro':
                # Monto manual del rubro (si existe) tiene prioridad — igual que
                # el Resumen en pantalla (`_filas_resumen`).
                manual = conn.execute(
                    "SELECT precio FROM gastos_generales"
                    " WHERE proyecto_id=? AND rubro=? AND tipo='manual'",
                    (proyecto_id, codigo)
                ).fetchone()
                if manual is not None:
                    val = manual['precio'] or 0
                    has_items = True
                else:
                    gg = conn.execute(
                        "SELECT * FROM gastos_generales WHERE proyecto_id=? AND rubro=? AND tipo='item'",
                        (proyecto_id, codigo)
                    ).fetchall()
                    if gg:
                        val = sum(
                            (i['cantidad'] or 0)
                            * ((i['pct_participacion'] or 100) / 100)
                            * (i['precio'] or 0)
                            for i in gg
                        )
                        has_items = True
                    else:
                        val = cd * pct / 100
            else:
                val = cd * pct / 100
            pct_real = round(val / cd * 100, 2) if cd else 0
            acum += val
            result.append({'tipo': tipo, 'nombre': nombre, 'valor': val, 'codigo': codigo,
                           'pct': pct, 'mostrar_pct': mp, 'has_items': has_items, 'pct_real': pct_real})

    return result, acum


# ─── EXCEL: Fórmula Polinómica ────────────────────────────────────────────────

def exportar_formula_polinomica(proyecto_id):
    conn     = get_db()
    proyecto = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    monomios = conn.execute(
        "SELECT * FROM formula_monomios WHERE proyecto_id=? ORDER BY orden",
        (proyecto_id,)
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fórmula Polinómica"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    N = 5  # columnas: Símbolo | Descripción | Índice INEI | Coeficiente k | % Participación
    t = Side(style='thin')
    m = Side(style='medium')
    b_hdr  = Border(left=t, right=t, top=m, bottom=m)
    b_data = Border(left=t, right=t, top=t, bottom=t)
    b_tot  = Border(left=t, right=t, top=m, bottom=t)
    al_c   = Alignment(horizontal='center', vertical='center')
    al_r   = Alignment(horizontal='right',  vertical='center')
    al_l   = Alignment(horizontal='left',   vertical='center')

    r = 1
    # ── Encabezado del proyecto ──────────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    ws.cell(r, 1, 'FÓRMULA POLINÓMICA').font      = Font(name='Inter', bold=True, italic=True, size=16)
    ws.cell(r, 1).alignment = al_c
    r += 1

    info = [('Proyecto', proyecto['nombre'] or ''),
            ('Cliente',  proyecto['cliente'] or ''),
            ('Ubicación',proyecto['ubicacion'] or '')]
    for label, val in info:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=N)
        ws.cell(r, 1, label).font = Font(name='Inter', bold=True, italic=True, size=11)
        ws.cell(r, 2, val).font   = Font(name='Inter', bold=(label == 'Proyecto'), size=11)
        r += 1

    # separador
    for c in range(1, N + 1):
        ws.cell(r, c).border = Border(top=m, bottom=t)
    r += 1

    # ── Expresión de la fórmula ──────────────────────────────────────────────
    if monomios:
        partes = []
        for mo in monomios:
            k   = float(mo['coeficiente'] or 0)
            sim = mo['simbolo'] or '?'
            partes.append(f"{k:.4f}·({sim}r/{sim}o)")
        expresion = "K = " + " + ".join(partes)
    else:
        expresion = "K = (sin monomios)"

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    ws.cell(r, 1, expresion).font      = Font(name='Courier New', bold=True, size=12)
    ws.cell(r, 1).alignment            = al_l
    ws.cell(r, 1).border               = b_data
    r += 1

    ws.row_dimensions[r - 1].height = 18
    r += 1  # espacio

    # ── Cabecera tabla ───────────────────────────────────────────────────────
    hdrs = ['Símbolo', 'Descripción', 'Índice INEI', 'Coeficiente k', '% Participación']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(r, c, h)
        cell.font      = Font(name='Inter', bold=True, italic=True, size=11)
        cell.alignment = al_c
        cell.border    = b_hdr
    r += 1

    # ── Filas de monomios ────────────────────────────────────────────────────
    for mo in monomios:
        k   = float(mo['coeficiente'] or 0)
        pct = k * 100
        vals = [mo['simbolo'] or '', mo['descripcion'] or '',
                mo['indice_inei'] or '', k, f"{pct:.2f}%"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font      = Font(name='Inter', size=11)
            cell.border    = b_data
            cell.alignment = al_r if c in (4, 5) else (al_c if c in (1, 3) else al_l)
            if c == 4:
                cell.number_format = '0.0000'
    r += 1

    # ── Fila total ───────────────────────────────────────────────────────────
    suma_k   = sum(float(mo['coeficiente'] or 0) for mo in monomios)
    suma_pct = suma_k * 100
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(r, 1, 'TOTAL').font      = Font(name='Inter', bold=True, size=11)
    ws.cell(r, 1).alignment          = al_r
    ws.cell(r, 1).border             = b_tot
    ws.cell(r, 2).border             = b_tot
    ws.cell(r, 3).border             = b_tot
    ws.cell(r, 4, suma_k).font       = Font(name='Inter', bold=True, size=11)
    ws.cell(r, 4).number_format      = '0.0000'
    ws.cell(r, 4).alignment          = al_r
    ws.cell(r, 4).border             = b_tot
    ws.cell(r, 5, f"{suma_pct:.2f}%").font = Font(name='Inter', bold=True, size=11)
    ws.cell(r, 5).alignment          = al_r
    ws.cell(r, 5).border             = b_tot

    # ── Anchos de columna ────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Helper: configuración de impresión ──────────────────────────────────────

def _setup_impresion(ws, n_filas_encabezado=8, n_cols=None, *, proyecto=None):
    """Ajusta la hoja para imprimir en A4 ajustado al ancho de la tabla.

    Si `n_cols` se especifica, fija el área de impresión a las primeras
    `n_cols` columnas y centra horizontalmente el contenido en la página.
    Si `proyecto` se especifica, agrega el pie tripartito espejo del PDF
    (Cliente | Fecha del reporte | Página X de N).
    """
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1   # 1 página de ancho (no escala en alto)
    ws.page_setup.fitToHeight  = 0   # sin límite de páginas en alto
    ws.page_setup.paperSize    = ws.PAPERSIZE_A4
    ws.page_setup.orientation  = 'portrait'
    # Márgenes compactos — el header tripartito ya ocupa ~30pt; un top margin
    # generoso (default 0.75") deja un hueco visible antes de la fila 1 que
    # Marco prefiere reducir al mínimo para que el reporte arranque arriba.
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.35, bottom=0.5,
        header=0.15, footer=0.2
    )
    # Repetir filas de encabezado (proyecto/título) en cada página impresa
    ws.print_title_rows = f'1:{n_filas_encabezado}'
    # Área de impresión + centrado horizontal (si se proporciona el ancho)
    if n_cols:
        last_col = get_column_letter(n_cols)
        last_row = ws.max_row
        ws.print_area = f'A1:{last_col}{last_row}'
        ws.print_options.horizontalCentered = True

    # Pie tripartito espejo del PDF — `&P` página actual, `&N` total páginas
    # (openpyxl los traduce a códigos nativos que Excel/Calc reemplazan al
    # imprimir o exportar). `&K666` setea color slate-500, `&7` size 7pt.
    if proyecto is not None:
        from datetime import datetime as _dt_p
        cliente = (proyecto['cliente'] or '').strip() if 'cliente' in proyecto.keys() else ''
        fecha   = _dt_p.now().strftime('%d/%m/%Y')
        color   = '485A6C'  # slate-500
        # El tamaño (&7) DEBE ir antes del color (&K). Si va al final y el
        # texto empieza con dígito (la fecha), Excel concatena `&7`+dígitos
        # como un solo tamaño (&701 → 701pt, texto gigante). Con `&K` último
        # lee exactamente 6 hex y el dígito siguiente ya cuenta como texto.
        prefix  = f'&7&K{color}'
        ws.oddFooter.left.text   = f"{prefix}Cliente: {cliente}" if cliente else f"{prefix} "
        ws.oddFooter.center.text = f"{prefix}{fecha}"
        ws.oddFooter.right.text  = f"{prefix}Página &P de &N"


# ─── Helper: encabezado Excel estilo PDF (3 zonas) ─────────────────────────

def _xlsx_header_pdf_style(ws, proyecto, titulo, n_cols, *, cols_partition=None):
    """Encabezado al estilo del `_draw_header` del PDF — 3 zonas:
    IZQ "ingePresupuestos" + subtítulo, CENTRO título + nombre del proyecto,
    DER "Costo al: …" + modalidad. Devuelve la fila siguiente.

    Estructura:
      Fila 1: empresa | título reporte    | Costo al
      Fila 2: subtítulo | nombre proyecto | modalidad
      Fila 3: separador slate-300

    `cols_partition=(L_END, C_END)`: opcional, especifica las columnas finales
    de las zonas izquierda y centro. Sin esto se usa el reparto por terceros
    (`third = n_cols // 3`). Útil para reportes con anchos de col muy desiguales
    como Presupuesto (`(2, 5)` → L=A-B, C=C-E, R=F-G).
    """
    if cols_partition is not None:
        if len(cols_partition) == 3:
            # Override completo: (L_END, C_END, R_START) — permite "huecos"
            # entre center y right (cols intermedias sin contenido en header).
            L_END, C_END, R_START = cols_partition
        else:
            L_END, C_END = cols_partition
            R_START = C_END + 1
    else:
        third = max(2, n_cols // 3)
        L_END = third
        C_END = min(n_cols, 2 * third)
        R_START = C_END + 1
    C_START = L_END + 1
    R_END   = n_cols

    # Empresa + subtítulo desde la configuración (espejo del PDF, que también
    # lee de `rep_empresa_nombre` / `rep_empresa_subtitulo`).
    from core.pdf_reports import get_formato
    from utils.theme import accent_reportes
    _fmt = get_formato()
    _empresa  = _fmt.get('rep_empresa_nombre')    or 'IngePresupuestos'
    _subtitulo = _fmt.get('rep_empresa_subtitulo') or 'Sistema de Presupuestos de Obra Pública'
    # Color del header: respeta `rep_color_marca_dk` custom, fallback a `od`
    # de `accent_reportes()` (slate-800 en sobrio, naranja-700 en marca).
    _o_acc, _od_acc, _ = accent_reportes()
    _hdr_dk = (_fmt.get('rep_color_marca_dk') or _od_acc).lstrip('#').upper()

    # ── Fila 1: empresa | título | costo al ───────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=L_END)
    c_emp = ws.cell(1, 1, _empresa)
    c_emp.font      = Font(name='Inter', bold=True, size=12, color=_hdr_dk)
    c_emp.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(start_row=1, start_column=C_START, end_row=1, end_column=C_END)
    c_tit = ws.cell(1, C_START, titulo)
    c_tit.font      = Font(name='Inter', bold=True, size=11, color='1F2A38')
    c_tit.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)

    ws.merge_cells(start_row=1, start_column=R_START, end_row=1, end_column=R_END)
    from core.pdf_reports import _clean_costo_al
    c_costo = ws.cell(1, R_START, f"Costo al: {_clean_costo_al(proyecto['costo_al'])}")
    c_costo.font      = Font(name='Inter', bold=True, size=9, color='2E3C52')
    c_costo.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[1].height = 18

    # ── Fila 2: subtítulo | nombre proyecto (wrap) | modalidad ────────────
    # Subtítulo y modalidad alineados ARRIBA — cuando el nombre del proyecto
    # wrappea a 2-3 líneas la fila crece, y con vertical='center' los textos
    # cortos quedan flotando al medio. Top los pega contra empresa/costo de
    # la fila 1, mismo look que el PDF.
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=L_END)
    c_sub = ws.cell(2, 1, _subtitulo)
    c_sub.font      = Font(name='Inter', size=8, color='667885')
    c_sub.alignment = Alignment(horizontal='left', vertical='top')

    # Nombre del proyecto en el centro — espejo del PDF (justo debajo del
    # título "Presupuesto"). Solo dejamos `wrap_text=True` y openpyxl /
    # LibreOffice se encargan del wrap natural. Calculamos n_lineas SOLO
    # para fijar la altura de la fila (sin esto, el viewer recorta).
    nombre = (proyecto['nombre'] or '').strip()
    centro_w = sum(
        (ws.column_dimensions[get_column_letter(c)].width or 8)
        for c in range(C_START, C_END + 1)
    )
    # Inter 7pt italic UPPERCASE: ~1.15 chars por col-unit en LibreOffice.
    chars_por_linea = max(20, int(centro_w * 1.15))
    n_lineas = min(3, max(1, -(-len(nombre) // chars_por_linea)))

    ws.merge_cells(start_row=2, start_column=C_START, end_row=2, end_column=C_END)
    c_proy = ws.cell(2, C_START, nombre)
    c_proy.font      = Font(name='Inter', italic=True, size=7, color='485A6C')
    # vertical='top' — alinea con subtítulo y modalidad (row 1+ continuación),
    # da look más limpio cuando el nombre wrappea a 2-3 líneas.
    c_proy.alignment = Alignment(horizontal='center', vertical='top',
                                  wrap_text=True)

    ws.merge_cells(start_row=2, start_column=R_START, end_row=2, end_column=R_END)
    c_mod = ws.cell(2, R_START, (proyecto['modalidad'] or '').strip())
    c_mod.font      = Font(name='Inter', italic=True, size=8, color='667885')
    c_mod.alignment = Alignment(horizontal='right', vertical='top')

    # Altura fila 2 = max(subtitulo+modalidad mínimos, n_lineas del proyecto)
    ws.row_dimensions[2].height = max(14, n_lineas * 11.0 + 6)

    # ── Fila 3: separador sutil (border-bottom slate-300) ─────────────────
    thin_slate = Side(style='thin', color='CBD5E1')
    for c in range(1, n_cols + 1):
        ws.cell(3, c).border = Border(bottom=thin_slate)
    ws.row_dimensions[3].height = 2

    return 4


# ─── Helper: encabezado Excel (legacy, estilo PowerCost) ────────────────────

def _xlsx_encabezado(ws, proyecto, titulo, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws['A1'] = titulo
    ws['A1'].font      = Font(name='Inter', bold=False, italic=False, size=21)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    info = [
        ('Proyecto', proyecto['nombre']          or ''),
        ('Sub Pto',  proyecto['sub_presupuesto'] or proyecto['nombre'] or ''),
        ('Cliente',  proyecto['cliente']         or ''),
        ('Ubicación', proyecto['ubicacion']      or ''),
    ]
    r = 2
    # Las primeras 3 filas de info ocupan hasta n_cols; la última fila deja
    # las 2 columnas finales libres para "Costo a:"
    for i, (label, val) in enumerate(info):
        # Etiqueta: regular (sin negrita ni cursiva), con wrap para que
        # "Sub Presupuesto" se ajuste en dos líneas si col A es estrecha
        lbl_cell = ws.cell(r, 1, label)
        lbl_cell.font = Font(name='Inter', bold=False, italic=False, size=11)
        lbl_cell.alignment = Alignment(horizontal='left', vertical='center',
                                        wrap_text=True)
        end_col = (n_cols - 2) if i == len(info) - 1 else n_cols
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=end_col)
        val_cell = ws.cell(r, 2, val)
        val_cell.font = Font(name='Inter',
            bold=(label == 'Sub Pto'), size=11)
        # Wrap + justify para que nombres largos se ajusten dentro del ancho
        val_cell.alignment = Alignment(horizontal='justify', vertical='center',
                                       wrap_text=True)
        # Altura proporcional al texto del VALOR o del LABEL (lo que sea mayor)
        ancho_aprox = sum(
            (ws.column_dimensions[get_column_letter(c)].width or 8)
            for c in range(2, end_col + 1)
        )
        chars_por_linea = max(40, int(ancho_aprox * 1.1))
        n_lineas_val = max(1, -(-len(val or '') // chars_por_linea))
        # Label en col A (width ~10): "Sub Presupuesto" (15 chars) wrappea a 2 líneas
        ancho_a = ws.column_dimensions['A'].width or 10
        chars_lbl = max(6, int(ancho_a * 1.1))
        n_lineas_lbl = max(1, -(-len(label) // chars_lbl))
        n_lineas = max(n_lineas_val, n_lineas_lbl)
        ws.row_dimensions[r].height = max(18.0, n_lineas * 15.0)
        r += 1
    ub_row = r - 1
    # "Costo a:" en las 2 últimas columnas de la fila Ubicación, fusionadas
    ws.merge_cells(start_row=ub_row, start_column=n_cols - 1, end_row=ub_row, end_column=n_cols)
    costo_txt = f"Costo a :  {proyecto['costo_al'] or ''}"
    ws.cell(ub_row, n_cols - 1, costo_txt).font = Font(name='Inter', bold=True, italic=True, size=11)
    ws.cell(ub_row, n_cols - 1).alignment = Alignment(horizontal='right')

    # Línea separadora doble debajo del encabezado
    for c in range(1, n_cols + 1):
        ws.cell(r, c).border = Border(
            top=Side(style='medium'), bottom=Side(style='thin'))
    return r + 1


# ─── EXCEL: Presupuesto ───────────────────────────────────────────────────────

def exportar_presupuesto(proyecto_id):
    """Export Excel del reporte de Presupuesto — espejo visual del PDF.

    Reglas de paridad con el PDF (`core/pdf_reports._html_presupuesto`):
    - Sin gridlines nativos de Excel (fondo blanco limpio).
    - Cabecera 3 zonas idéntica al `_draw_header` del PDF.
    - Tabla de partidas: SIN bordes en filas de partida, zebra muy sutil.
    - Cabecera de columnas: solo border-bottom naranja medium (nada más).
    - N1 uppercase + subrayado tipográfico, SIN bordes top/bottom (el PDF
      tampoco los muestra: QTextDocument ignora el CSS `tr.titulo1 td border`).
    - N2 rojo, N3 azul, N4 morado, N5+ marrón italic — SIN bordes (solo color).
    - Spacer rows antes de cada título (no si es primer hijo de su padre).
    - Pie SIN bordes (PDF tampoco los muestra). Diferenciación por peso/tamaño/color:
      'gran' bold 12pt slate-900 · 'sub' bold 11pt slate-900 · 'else' regular 10pt slate-700.
    - "Son: ..." italic slate-700 derecha.
    """
    from openpyxl.styles import PatternFill

    conn     = get_db()
    proyecto = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    pie      = _get_pie(conn, proyecto_id)
    items, totales = calcular_totales(proyecto_id)
    conn.close()

    # ── Paleta (mirror PDF) ─────────────────────────────────────────────────
    C_ORANGE      = 'F37329'
    C_ORANGE_DARK = 'C0621A'
    C_SLATE_900   = '1F2A38'
    C_SLATE_700   = '2E3C52'
    C_SLATE_500   = '64748B'
    C_SLATE_300   = 'CBD5E1'
    C_ALT_BG      = 'FBFBFC'
    # Colores de fuente por nivel = espejo del programa (NIVEL_ESTILO en
    # proyecto_view.py): N1 rojo, N2 arándano, N3 morado, N4 rosa.
    C_TITULO1     = 'B71C1C'   # rojo (capítulos principales)
    C_TITULO2     = '0D52BF'   # arándano (sub-capítulos)
    C_TITULO3     = '6A1B9A'   # morado
    C_TITULO4     = 'AD1457'   # rosa
    C_TITULO5     = '92400E'   # marrón (sin equivalente en programa)

    side_orange_md = Side(style='medium', color=C_ORANGE)
    side_slate_md  = Side(style='medium', color=C_SLATE_700)
    side_slate_th  = Side(style='thin',   color=C_SLATE_700)
    side_slate_300 = Side(style='thin',   color=C_SLATE_300)
    side_thick     = Side(style='thick',  color=C_SLATE_900)
    side_med_900   = Side(style='medium', color=C_SLATE_900)
    alt_fill = PatternFill(start_color=C_ALT_BG, end_color=C_ALT_BG, fill_type='solid')

    # Alturas de spacer por nivel (espejo del `_spacer_h` del PDF)
    SPACER_H = {1: 7, 2: 6, 3: 5, 4: 4, 5: 4}

    # ── Workbook setup ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    # Sin gridlines NI en pantalla NI en impresión — clave para fondo blanco
    # limpio como el PDF. `showGridLines` solo afecta vista, hay que setear
    # también `print_options.gridLines = False` para que la impresión coincida.
    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines    = False
    # `gridLinesSet="1"` = "el usuario seteó gridLines explícitamente" →
    # LibreOffice respeta el `gridLines="0"`. Con "0" Calc usa su default
    # (que en LibreOffice puede ser ON aunque Excel diga OFF).
    ws.print_options.gridLinesSet = True
    # Defensa adicional: `printGridLines` en sheet properties
    ws.sheet_properties.outlinePr = None
    # Algunos viewers respetan estos del pageSetup
    ws.page_setup.blackAndWhite = False

    # 7 cols físicas (Descripción se SPLIT en B+C para balancear el tripartite
    # del header — sin esto el center queda muy angosto y el título
    # "Presupuesto" no alinea con el nombre del proyecto. En cada fila de
    # datos B+C se mergean → el usuario sigue viendo 1 columna Descripción).
    # Tripartite resultante: Left=A+B=28 · Center=C+D=39 · Right=E+F+G=42.
    N = 7  # Ítem · Desc1 · Desc2 · Und · Metrado · Precio Unit. · Parcial
    ws.column_dimensions['A'].width = 10   # Ítem
    ws.column_dimensions['B'].width = 18   # Descripción parte 1
    ws.column_dimensions['C'].width = 43   # Descripción parte 2 (ancha para
                                           # que el centro del header respire)
    ws.column_dimensions['D'].width = 7    # Und
    ws.column_dimensions['E'].width = 12   # Metrado
    ws.column_dimensions['F'].width = 14   # Precio Unit.
    ws.column_dimensions['G'].width = 16   # Parcial

    # Helper para mergear Descripción (cols 2+3) en una fila dada.
    def _merge_descripcion(row_idx: int):
        ws.merge_cells(start_row=row_idx, start_column=2,
                          end_row=row_idx, end_column=3)

    # ── Encabezado 3 zonas (espejo del PDF) ─────────────────────────────────
    # cols_partition=(2, 5) → Left=A-B (28) · Center=C-E (51) · Right=F-G (30).
    # Sin el override, la heurística por terceros (L=A-B, C=C-D, R=E-G) deja
    # el centro muy angosto (19 col-units) y el nombre del proyecto en el
    # centro no respira.
    r = _xlsx_header_pdf_style(ws, proyecto, 'Presupuesto', N,
                                  cols_partition=(2, 5))

    # ── Título h2 "Presupuesto" — usa `od` puro de accent_reportes (espejo
    # del CSS `h2 { color: {od} }` del PDF). En modo sobrio = slate-900 (negro);
    # en modo marca = naranja-700. NO respeta `rep_color_marca_dk` custom
    # porque el CSS del PDF tampoco lo hace.
    from utils.theme import accent_reportes as _accent_h2
    _h2_color = _accent_h2()[1].lstrip('#').upper()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    c_h2 = ws.cell(r, 1, 'Presupuesto')
    c_h2.font      = Font(name='Inter', bold=True, size=14, color=_h2_color)
    c_h2.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 22
    r += 1

    # ── Cabecera de columnas — bg accent_soft (espejo del PDF + card ACU).
    # Descripción merge B+C; otras columnas físicas: A · (B+C) · D · E · F · G
    from utils.theme import accent_reportes as _accent_hdr_p
    _, _, _os_hdr_p = _accent_hdr_p()
    C_HDR_BG = _os_hdr_p.lstrip('#').upper()
    hdr_fill = PatternFill(start_color=C_HDR_BG, end_color=C_HDR_BG,
                              fill_type='solid')
    hdrs_fisicos = [
        (1, 'Ítem',         'left'),
        (2, 'Descripción',  'left'),    # merge con col 3
        (4, 'Und',          'center'),
        (5, 'Cantidad',     'right'),
        (6, 'Precio',       'right'),
        (7, 'Parcial',      'right'),
    ]
    # PRE-aplicar fill a TODAS las cols (incluye col 3 que será MergedCell).
    for col_ in range(1, N + 1):
        ws.cell(r, col_).fill = hdr_fill
    for col_idx, h, al in hdrs_fisicos:
        cell = ws.cell(r, col_idx, h)
        cell.font      = Font(name='Inter', bold=True, size=10, color=C_SLATE_900)
        cell.alignment = Alignment(horizontal=al, vertical='center')
    _merge_descripcion(r)
    filas_titulo_impresion = r
    r += 1

    # ── Loop de items ───────────────────────────────────────────────────────
    fmt_money = '[$-0409]#,##0.00'
    _dm = get_decimales_metrado()
    fmt_met = ('[$-0409]#,##0.' + '0' * _dm) if _dm else '[$-0409]#,##0'

    def _font_titulo(nivel: int) -> Font:
        if nivel == 1:
            return Font(name='Inter', bold=True, size=11,
                        color=C_TITULO1, underline='single')
        if nivel == 2:
            return Font(name='Inter', bold=True, size=11, color=C_TITULO2)
        if nivel == 3:
            return Font(name='Inter', bold=True, size=11, color=C_TITULO3)
        if nivel == 4:
            return Font(name='Inter', bold=True, size=11, color=C_TITULO4)
        return Font(name='Inter', bold=True, italic=True, size=11, color=C_TITULO5)

    hoja_idx = 0
    prev_titulo = False
    prev_nivel  = 0
    # Contador propio de partidas hoja para zebra continuo (los títulos no
    # cuentan, así la alternancia visual se mantiene al cruzar un título).
    partida_idx = 0
    # Sangría jerárquica de la Descripción (espejo del PDF/programa): nivel de
    # indent = profundidad por puntos del ítem, relativa al más superficial.
    # `nivel` está topeado en 4 → se usa el código del ítem.
    _dots = [(it['partida'].get('item') or '').count('.') for it in items]
    _min_dots = min(_dots) if _dots else 0

    for idx, entry in enumerate(items):
        p     = entry['partida']
        total = entry['total']
        es_titulo = bool(p.get('es_titulo'))
        nivel = int(p.get('nivel') or 1)
        desc = p['descripcion'] or ''
        _depth = max(0, (p.get('item') or '').count('.') - _min_dots)

        # Spacer row — solo si el anterior NO es padre directo
        if es_titulo and idx > 0:
            if not (prev_titulo and prev_nivel < nivel):
                ws.row_dimensions[r].height = SPACER_H.get(nivel, 4)
                r += 1

        if es_titulo and nivel == 1:
            # N1: uppercase + underline + slate-900 bold — SIN bordes.
            # Descripción merge cols 2-6 (B-F: cubre Desc1+Desc2+Und+Met+Precio),
            # monto en col 7 (G).
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            f = _font_titulo(1)
            ws.cell(r, 1, p['item']).font = f
            ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
            c_d = ws.cell(r, 2, desc.upper()); c_d.font = f
            c_d.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=_depth)
            c_v = ws.cell(r, 7, total); c_v.font = f
            c_v.number_format = fmt_money
            c_v.alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[r].height = max(22, _alto_fila(desc, 75))

        elif es_titulo and nivel == 2:
            # N2 rojo — solo color, SIN bordes (igual que el PDF visible)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            f = _font_titulo(2)
            ws.cell(r, 1, p['item']).font = f
            ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
            c_d = ws.cell(r, 2, desc); c_d.font = f
            c_d.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=_depth)
            c_v = ws.cell(r, 7, total); c_v.font = f
            c_v.number_format = fmt_money
            c_v.alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[r].height = max(20, _alto_fila(desc, 65))

        elif es_titulo:
            # N3, N4, N5+ — solo color de texto, SIN bordes
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            f = _font_titulo(nivel)
            ws.cell(r, 1, p['item']).font = f
            ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
            c_d = ws.cell(r, 2, desc); c_d.font = f
            c_d.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=_depth)
            c_v = ws.cell(r, 7, total); c_v.font = f
            c_v.number_format = fmt_money
            c_v.alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[r].height = max(18, _alto_fila(desc, 60))

        else:
            # Partida hoja — texto regular, SIN bordes, zebra `#FBFCFD`.
            # Descripción ocupa col 2+3 mergeados (split del header → user
            # ve 1 sola columna). Resto: Und=4, Met=5, Precio=6, Parcial=7.
            fh = Font(name='Inter', size=10, color=C_SLATE_900)
            zebra = (partida_idx % 2 == 1)
            zebra_fill = (PatternFill(start_color='FBFCFD',
                                          end_color='FBFCFD',
                                          fill_type='solid')
                          if zebra else None)
            partida_idx += 1
            celdas_fila = []
            ws.cell(r, 1, p['item']).font = fh
            ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='top')
            celdas_fila.append(ws.cell(r, 1))
            c_desc = ws.cell(r, 2, desc); c_desc.font = fh
            c_desc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=_depth)
            celdas_fila.append(c_desc)
            # PRE-estilar col 3 antes del merge (luego es MergedCell read-only)
            ws.cell(r, 3).font = fh
            celdas_fila.append(ws.cell(r, 3))
            ws.cell(r, 4, p['unidad']).font = fh
            ws.cell(r, 4).alignment = Alignment(horizontal='center', vertical='top')
            celdas_fila.append(ws.cell(r, 4))
            c_met = ws.cell(r, 5, p['metrado']); c_met.font = fh
            c_met.number_format = fmt_met
            c_met.alignment = Alignment(horizontal='right', vertical='top')
            celdas_fila.append(c_met)
            c_pu = ws.cell(r, 6, p['precio_unitario']); c_pu.font = fh
            c_pu.number_format = fmt_money
            c_pu.alignment = Alignment(horizontal='right', vertical='top')
            celdas_fila.append(c_pu)
            c_to = ws.cell(r, 7, total); c_to.font = fh
            c_to.number_format = fmt_money
            c_to.alignment = Alignment(horizontal='right', vertical='top')
            celdas_fila.append(c_to)
            if zebra_fill:
                for c_ in celdas_fila:
                    c_.fill = zebra_fill
            # Mergear Descripción (cols 2+3) DESPUÉS de estilar — para que
            # el fill se aplique antes a ambos cells.
            _merge_descripcion(r)
            # chars_por_linea conservador (42) para LibreOffice/PlanMaker.
            ws.row_dimensions[r].height = max(18, _alto_fila(desc, 42))

        r += 1
        prev_titulo = es_titulo
        prev_nivel  = nivel

    # ── Pie de presupuesto ──────────────────────────────────────────────────
    # Espacio entre tabla y pie (espejo del `margin-top:28pt` del PDF)
    r += 2

    from core.pdf_reports import _build_pie_rows
    pie_rows = _build_pie_rows(proyecto_id, totales.get('cd', 0) or 0)

    def _fila_pie(label, val, cls):
        nonlocal r
        # SIN bordes en el pie (espejo del PDF visible: QTextDocument ignora
        # `tr.gran td border` y similares, el PDF sale sin líneas en el pie).
        # Diferenciación visual solo por peso/tamaño/color.
        if cls == 'gran':
            bold, size, color = True, 12, C_SLATE_900
        elif cls == 'sub':
            bold, size, color = True, 11, C_SLATE_900
        else:
            bold, size, color = False, 10, C_SLATE_700
        # Label cols 2-6 right (Desc1+Desc2+Und+Met+Precio), valor en col 7.
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c_l = ws.cell(r, 2, label)
        c_l.font = Font(name='Inter', bold=bold, size=size, color=color)
        c_l.alignment = Alignment(horizontal='right', vertical='center')
        c_v = ws.cell(r, 7, val)
        c_v.font = Font(name='Inter', bold=bold, size=size, color=color)
        c_v.number_format = fmt_money
        c_v.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[r].height = 20 if cls in ('gran', 'sub') else 16
        r += 1

    for label, val, cls in pie_rows:
        _fila_pie(label.upper(), val, cls)

    # ── "Son: ..." — italic slate-700, alineado a la derecha ────────────────
    monto_total = pie_rows[-1][1] if pie_rows else (totales.get('total', 0) or 0)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    for c in range(1, N + 1):
        ws.cell(r, c).border = Border()
    c_son = ws.cell(r, 1, f"Son: {_monto_letras(monto_total)}.")
    c_son.font = Font(name='Inter', italic=True, size=10, color=C_SLATE_700)
    c_son.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 22

    # ── Firmantes + impresión ───────────────────────────────────────────────
    _escribir_pie(ws, r + 2, pie, n_cols=N)
    _setup_impresion(ws, n_filas_encabezado=filas_titulo_impresion, n_cols=N,
                       proyecto=proyecto)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── EXCEL: ACUs ──────────────────────────────────────────────────────────────

def exportar_acus(proyecto_id):
    """Export Excel del reporte ACU — espejo visual del PDF.

    Espejo de `core/pdf_reports._html_acus`:
    - Header 3 zonas + nombre proyecto full-width (vía `_xlsx_header_pdf_style`).
    - h2 "Análisis de Costos Unitarios" izq, color `od` de `accent_reportes`.
    - Por cada partida: head con `{item}  {descripción}` (slate-900 bold) y
      meta line "Unidad · Rendimiento · Costo Unit." (slate-500 italic).
    - Tabla 7 cols: Tipo · Recurso · Und · Cuadrilla · Cantidad · Precio · Parcial.
    - Filas con tipo en color (MO ámbar · MAT verde · EQ slate · SC morado).
    - Subtotales por tipo: solo color/peso, SIN bordes.
    - COSTO UNITARIO TOTAL: bold slate-900, SIN bordes.
    - Cabecera de columnas: border-bottom slate-300 thin (no naranja).
    - Sin gridlines nativos.
    """
    from openpyxl.styles import PatternFill

    conn     = get_db()
    proyecto = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    partidas = conn.execute(
        "SELECT * FROM partidas WHERE proyecto_id=? AND es_titulo=0 ORDER BY item",
        (proyecto_id,)
    ).fetchall()

    # ── Paleta (mirror PDF) ─────────────────────────────────────────────────
    C_SLATE_900 = '1F2A38'
    C_SLATE_700 = '2E3C52'
    C_SLATE_500 = '64748B'
    C_SLATE_300 = 'CBD5E1'
    # Colores por tipo de recurso (mirror utils/theme.py)
    TIPO_FG = {'MO': 'B7791F', 'MAT': '15803D', 'EQ': '475569', 'SC': '6B21A8'}
    TIPO_LABEL = {'MO': 'Mano de obra', 'MAT': 'Materiales',
                  'EQ': 'Equipos', 'SC': 'Sub-contratos'}

    side_slate_300 = Side(style='thin', color=C_SLATE_300)

    # ── Workbook setup ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACUs"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines = False
    ws.print_options.gridLinesSet = True

    # 8 cols físicas (Recurso se SPLIT en B+C — espejo del patrón Presupuesto:
    # da más espacio al center del tripartite del header. B+C se mergean en
    # cada fila → el usuario sigue viendo 1 columna Recurso).
    N = 8  # Tipo · Rec1 · Rec2 · Und · Cuadrilla · Cantidad · Precio · Parcial
    ws.column_dimensions['A'].width = 8    # Tipo
    ws.column_dimensions['B'].width = 20   # Recurso parte 1
    ws.column_dimensions['C'].width = 24   # Recurso parte 2 (ancha para center)
    ws.column_dimensions['D'].width = 7    # Und
    ws.column_dimensions['E'].width = 11   # Cuadrilla
    ws.column_dimensions['F'].width = 12   # Cantidad
    ws.column_dimensions['G'].width = 12   # Precio
    ws.column_dimensions['H'].width = 14   # Parcial

    # Helper para mergear Recurso (cols 2+3) en una fila dada.
    def _merge_recurso(row_idx: int):
        ws.merge_cells(start_row=row_idx, start_column=2,
                          end_row=row_idx, end_column=3)

    # ── Encabezado 3 zonas + nombre proyecto en row 2 centro ────────────────
    # cols_partition=(2, 6) → Left=A+B (28) · Center=C+D+E+F (54) · Right=G+H (26).
    r = _xlsx_header_pdf_style(ws, proyecto, 'Análisis de Costos Unitarios', N,
                                  cols_partition=(2, 6))

    # ── Título h2 — color `od` de accent_reportes (slate-900 en sobrio) ─────
    from utils.theme import accent_reportes
    _h2_color = accent_reportes()[1].lstrip('#').upper()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    c_h2 = ws.cell(r, 1, 'Análisis de Costos Unitarios')
    c_h2.font      = Font(name='Inter', bold=True, size=14, color=_h2_color)
    c_h2.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 22
    r += 1

    # Símbolo de moneda + decimales
    try:
        from utils.formatting import _moneda_simbolo
        sym = _moneda_simbolo(proyecto['moneda'] or 'Soles')
    except Exception:
        sym = 'S/'

    fmt_money = '[$-0409]#,##0.00'
    fmt_4 = '#,##0.0000'

    for partida in partidas:
        # Usar `get_acu_items` (la misma fuente que usa el PDF) para que
        # los `parcial` de items con unidad `%MO`/`%MAT` se calculen como
        # porcentaje del subtotal del tipo base, NO como `cantidad * precio`.
        # Retorna (items, totales_tipo_aux) — solo necesitamos los items.
        items_acu, _ = get_acu_items(conn, partida['id'])
        if not items_acu:
            continue

        # ── Header de la partida ────────────────────────────────────────────
        # Bg soft del accent (espejo del PDF `acu-head { background: {os_} }`).
        # Cubre AMBAS filas: la del item + descripción Y la de meta (Unidad ·
        # Rendimiento · Costo Unit.) — visualmente un "card header" suave.
        from utils.theme import accent_reportes as _accent_card
        _, _, _os_card = _accent_card()
        C_CARD_BG = _os_card.lstrip('#').upper()
        card_fill = PatternFill(start_color=C_CARD_BG, end_color=C_CARD_BG,
                                  fill_type='solid')

        # Línea 1: ítem + descripción (slate-900 bold)
        # PRE-aplicar fill en todas las cols ANTES de mergear.
        for col_ in range(1, N + 1):
            ws.cell(r, col_).fill = card_fill
        head_txt = f"{partida['item'] or ''}   {partida['descripcion'] or ''}"
        c_head = ws.cell(r, 1, head_txt)
        c_head.font      = Font(name='Inter', bold=True, size=11, color=C_SLATE_900)
        c_head.alignment = Alignment(horizontal='left', vertical='center',
                                      wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        ws.row_dimensions[r].height = max(18, _alto_fila(head_txt, 95))
        r += 1

        # Línea 2: meta (Unidad · Rendimiento · Costo Unit.) — espejo del PDF:
        # regular slate-500, con los VALORES en bold (ej. "UND", "1.00 UND/día").
        # Uso CellRichText para mezclar regular+bold en una sola celda.
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        rendimiento = partida['rendimiento'] or 0
        unidad = partida['unidad'] or '—'
        cu = partida['precio_unitario'] or 0
        rend_txt = (f"{rendimiento:.2f} {unidad}/día"
                    if rendimiento else "—")
        f_reg  = InlineFont(rFont='Inter', sz=9, color=C_SLATE_500)
        f_bold = InlineFont(rFont='Inter', sz=9, color=C_SLATE_500, b=True)
        meta_rich = CellRichText(
            TextBlock(f_reg,  'Unidad: '),
            TextBlock(f_bold, unidad),
            TextBlock(f_reg,  '   ·   Rendimiento: '),
            TextBlock(f_bold, rend_txt),
            TextBlock(f_reg,  '   ·   Costo Unit.: '),
            TextBlock(f_bold, f'{sym} {cu:,.2f}'),
        )
        # PRE-aplicar mismo fill que la fila del item (acu-head card) en todas
        # las cols antes de mergear — sin esto el bg solo queda en col 1 y se
        # ve un cuadro desigual.
        for col_ in range(1, N + 1):
            ws.cell(r, col_).fill = card_fill
        c_meta = ws.cell(r, 1)
        c_meta.value = meta_rich
        c_meta.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        ws.row_dimensions[r].height = 14
        r += 1

        # ── Cabecera de columnas (border-bottom slate-300 thin) ─────────────
        # Cols físicas: A · (B+C merged) · D · E · F · G · H
        hdrs_fisicos = [
            (1, 'Tipo',      'left'),
            (2, 'Recurso',   'left'),    # merge con col 3
            (4, 'Und',       'center'),
            (5, 'Cuadrilla', 'right'),
            (6, 'Cantidad',  'right'),
            (7, 'Precio',    'right'),
            (8, 'Parcial',   'right'),
        ]
        for col_idx, h, al in hdrs_fisicos:
            cell = ws.cell(r, col_idx, h)
            cell.font      = Font(name='Inter', bold=True, size=10, color=C_SLATE_500)
            cell.alignment = Alignment(horizontal=al, vertical='center')
            cell.border    = Border(bottom=side_slate_300)
        _merge_recurso(r)
        ws.row_dimensions[r].height = 16
        r += 1

        # ── Filas por tipo (MO/MAT/EQ/SC) — TODOS los datos primero, luego
        # subtotales agrupados al final (espejo del PDF). Marco prefiere
        # que los montos de MO/MAT/EQ aparezcan juntos antes del COSTO
        # UNITARIO TOTAL, no interleaved entre cada sección.
        totales_tipo = {'MO': 0.0, 'MAT': 0.0, 'EQ': 0.0, 'SC': 0.0}
        # Contador propio para zebra continuo en TODAS las filas de datos del
        # ACU (no se resetea entre tipos MO/MAT/EQ/SC).
        recurso_idx = 0
        for tipo in ('MO', 'MAT', 'EQ', 'SC'):
            tipo_items = [x for x in items_acu if x['tipo'] == tipo]
            if not tipo_items:
                continue
            tipo_color = TIPO_FG[tipo]

            for it in tipo_items:
                # `parcial`, `precio` y `cantidad` vienen ya calculados por
                # `get_acu_items` (incluye el caso `%MO`/`%MAT` donde el
                # parcial = cantidad/100 * subtotal_base y precio = base).
                precio  = it['precio']   or 0
                cant    = it['cantidad'] or 0
                parcial = it['parcial']  or 0
                totales_tipo[tipo] += parcial

                # Zebra `#FBFCFD` cada fila impar — espejo del `tr.alt` del
                # PDF que usa `background:#FBFBFC` cuando j%2==1.
                zebra = (recurso_idx % 2 == 1)
                zebra_fill = (PatternFill(start_color='FBFCFD',
                                              end_color='FBFCFD',
                                              fill_type='solid')
                              if zebra else None)
                recurso_idx += 1
                # Col 1: Tipo (texto bold del color del tipo)
                c_t = ws.cell(r, 1, tipo)
                c_t.font = Font(name='Inter', bold=True, size=10, color=tipo_color)
                c_t.alignment = Alignment(horizontal='left', vertical='top')
                # Col 2: Recurso (descripción) — luego mergear con col 3
                c_d = ws.cell(r, 2, it['descripcion'] or '')
                c_d.font = Font(name='Inter', size=10, color=C_SLATE_900)
                c_d.alignment = Alignment(horizontal='left', vertical='top',
                                           wrap_text=True)
                # PRE-estilar col 3 para que el fill aplique antes de mergear
                ws.cell(r, 3).font = Font(name='Inter', size=10, color=C_SLATE_900)
                # Col 4-7: Und, Cuadrilla, Cantidad, Precio
                num_font = Font(name='Inter', size=10, color=C_SLATE_900)
                c_u = ws.cell(r, 4, it['unidad'] or ''); c_u.font = num_font
                c_u.alignment = Alignment(horizontal='center', vertical='top')
                c_cu = ws.cell(r, 5, it['cuadrilla'] or 0); c_cu.font = num_font
                c_cu.number_format = fmt_4
                c_cu.alignment = Alignment(horizontal='right', vertical='top')
                c_ca = ws.cell(r, 6, cant); c_ca.font = num_font
                c_ca.number_format = fmt_4
                c_ca.alignment = Alignment(horizontal='right', vertical='top')
                c_p = ws.cell(r, 7, precio); c_p.font = num_font
                c_p.number_format = fmt_money
                c_p.alignment = Alignment(horizontal='right', vertical='top')
                # Col 8: Parcial
                c_pa = ws.cell(r, 8, parcial); c_pa.font = num_font
                c_pa.number_format = fmt_money
                c_pa.alignment = Alignment(horizontal='right', vertical='top')
                # Aplicar zebra fill a TODAS las cols ANTES de mergear (post-
                # merge col 3 sería MergedCell read-only).
                if zebra_fill:
                    for c_ in range(1, N + 1):
                        ws.cell(r, c_).fill = zebra_fill
                _merge_recurso(r)
                ws.row_dimensions[r].height = max(16, _alto_fila(
                    it['descripcion'] or '', 38))
                r += 1

        # ── Subtotales agrupados al final (espejo del PDF) ──────────────────
        # Después de TODAS las filas de datos, listar Mano de obra, Materiales,
        # Equipos, Sub-contratos en bloque antes del COSTO UNITARIO TOTAL.
        for tipo in ('MO', 'MAT', 'EQ', 'SC'):
            if totales_tipo[tipo] <= 0:
                continue
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            c_sl = ws.cell(r, 1, TIPO_LABEL[tipo])
            c_sl.font = Font(name='Inter', bold=True, size=10, color=C_SLATE_700)
            c_sl.alignment = Alignment(horizontal='right', vertical='center')
            c_sv = ws.cell(r, 8, totales_tipo[tipo])
            c_sv.font = Font(name='Inter', bold=True, size=10, color=C_SLATE_900)
            c_sv.number_format = fmt_money
            c_sv.alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[r].height = 16
            r += 1

        # ── COSTO UNITARIO TOTAL — bold slate-900, SIN bordes (Marco prefiere
        # el look sobrio: diferenciación solo por peso/tamaño/color).
        cu_total = sum(totales_tipo.values())
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c_tl = ws.cell(r, 1, 'COSTO UNITARIO TOTAL')
        c_tl.font = Font(name='Inter', bold=True, size=11, color=C_SLATE_900)
        c_tl.alignment = Alignment(horizontal='right', vertical='center')
        c_tv = ws.cell(r, 8, cu_total)
        c_tv.font = Font(name='Inter', bold=True, size=11, color=C_SLATE_900)
        c_tv.number_format = fmt_money
        c_tv.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[r].height = 20
        r += 1

        # Espacio entre ACUs (mirror PDF: párrafo de separación)
        r += 2

    conn.close()
    # Solo repetir las 3 filas del tripartito en cada página — el h2 y la
    # cabecera de columnas se repiten POR partida dentro del cuerpo.
    _setup_impresion(ws, n_filas_encabezado=3, n_cols=N, proyecto=proyecto)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── EXCEL: Listado de insumos ────────────────────────────────────────────────

def exportar_insumos(proyecto_id):
    """Export Excel del reporte Insumos — espejo visual del PDF `_html_insumos`.

    - Header 3 zonas + nombre proyecto full-width.
    - h2 "Relación de Insumos" izq, color `od` de `accent_reportes`.
    - Tabla única 6 cols: Código · Descripción · Und · Cantidad · Precio · Parcial.
    - Por tipo (MO/MAT/EQ/SC): barra de sección con fondo suave del tipo +
      filas + subtotal del tipo en color, sin bordes.
    - Resumen final: tabla "Resumen por Tipo de Insumo" con subtotales y
      TOTAL COSTO DIRECTO en bold slate-900.
    """
    from openpyxl.styles import PatternFill
    from core.database import get_insumos_proyecto as _get_ins

    conn     = get_db()
    proyecto = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    # Usar `get_insumos_proyecto` (la misma fuente del PDF y la pestaña):
    # incluye overhead `%MO`/`%MAT` y aplica distribución proporcional al
    # parcial de cada partida → sum(insumos) ≈ CD exactamente.
    insumos = _get_ins(conn, proyecto_id)
    conn.close()
    _orden_tipo = {'MO': 0, 'MAT': 1, 'EQ': 2, 'SC': 3}
    insumos = sorted(insumos, key=lambda r: (_orden_tipo.get(r['tipo'], 99),
                                              r['descripcion'] or ''))

    # ── Paleta ──────────────────────────────────────────────────────────────
    C_SLATE_900 = '1F2A38'
    C_SLATE_500 = '64748B'
    # Colores por tipo (mirror utils/theme.py)
    TIPO_FG   = {'MO': 'B7791F', 'MAT': '15803D', 'EQ': '475569', 'SC': '6B21A8'}
    TIPO_SOFT = {'MO': 'FEF3DD', 'MAT': 'E3F4EA', 'EQ': 'ECEEF1', 'SC': 'F1E7F9'}
    TIPO_LABEL = {'MO': 'Mano de Obra', 'MAT': 'Materiales',
                  'EQ': 'Equipos y Herramientas',
                  'SC': 'Sub-contratos / Servicios'}

    # ── Workbook setup ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Insumos"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines = False
    ws.print_options.gridLinesSet = True

    # 7 cols físicas (Descripción se SPLIT en B+C — patrón espejo de
    # Presupuesto/ACU/Metrados: balancea el tripartite del header. B+C se
    # mergean en cada fila → user ve 1 sola columna Descripción).
    N = 7  # Código · Desc1 · Desc2 · Und · Cantidad · Precio · Parcial
    ws.column_dimensions['A'].width = 12   # Código
    ws.column_dimensions['B'].width = 18   # Descripción parte 1
    ws.column_dimensions['C'].width = 24   # Descripción parte 2
    ws.column_dimensions['D'].width = 7    # Und
    ws.column_dimensions['E'].width = 12   # Cantidad
    ws.column_dimensions['F'].width = 12   # Precio
    ws.column_dimensions['G'].width = 14   # Parcial

    # Helper para mergear Descripción (cols 2+3) en una fila dada.
    def _merge_descripcion(row_idx: int):
        ws.merge_cells(start_row=row_idx, start_column=2,
                          end_row=row_idx, end_column=3)

    # ── Encabezado 3 zonas + nombre proyecto en row 2 centro ────────────────
    # cols_partition=(2, 5) → Left=A+B (30) · Center=C+D+E (43) · Right=F+G (26).
    r = _xlsx_header_pdf_style(ws, proyecto, 'Relación de Insumos', N,
                                  cols_partition=(2, 5))

    # ── Título h2 ───────────────────────────────────────────────────────────
    from utils.theme import accent_reportes
    _h2_color = accent_reportes()[1].lstrip('#').upper()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    c_h2 = ws.cell(r, 1, 'Relación de Insumos')
    c_h2.font      = Font(name='Inter', bold=True, size=14, color=_h2_color)
    c_h2.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 22
    r += 1

    fmt_money = '[$-0409]#,##0.00'
    fmt_4     = '#,##0.0000'
    side_slate_300 = Side(style='thin', color='CBD5E1')

    # Símbolo moneda
    try:
        from utils.formatting import _moneda_simbolo
        sym = _moneda_simbolo(proyecto['moneda'] or 'Soles')
    except Exception:
        sym = 'S/'

    # ── Cabecera de columnas — bg accent_soft (espejo de Presupuesto).
    # Cols físicas: A · (B+C merged) · D · E · F · G
    from utils.theme import accent_reportes as _accent_hdr_i
    _, _, _os_hdr_i = _accent_hdr_i()
    C_HDR_BG_I = _os_hdr_i.lstrip('#').upper()
    hdr_fill_i = PatternFill(start_color=C_HDR_BG_I, end_color=C_HDR_BG_I,
                                fill_type='solid')
    hdrs_fisicos = [
        (1, 'Código',      'left'),
        (2, 'Descripción', 'left'),    # merge con col 3
        (4, 'Und',         'center'),
        (5, 'Cantidad',    'right'),
        (6, 'Precio',      'right'),
        (7, 'Parcial',     'right'),
    ]
    # PRE-aplicar fill a TODAS las cols (incluye col 3 que será MergedCell).
    for col_ in range(1, N + 1):
        ws.cell(r, col_).fill = hdr_fill_i
    for col_idx, h, al in hdrs_fisicos:
        cell = ws.cell(r, col_idx, h)
        cell.font      = Font(name='Inter', bold=True, size=10, color=C_SLATE_500)
        cell.alignment = Alignment(horizontal=al, vertical='center')
        cell.border    = Border(bottom=side_slate_300)
    _merge_descripcion(r)
    ws.row_dimensions[r].height = 16
    r += 1

    # ── Loop por tipo: barra de sección + items + subtotal ─────────────────
    por_tipo_lista = {}
    subtotal_tipo  = {'MO': 0.0, 'MAT': 0.0, 'EQ': 0.0, 'SC': 0.0}
    for ins in insumos:
        por_tipo_lista.setdefault(ins['tipo'], []).append(ins)

    primer_tipo = True
    for tipo in ('MO', 'MAT', 'EQ', 'SC'):
        lst = por_tipo_lista.get(tipo)
        if not lst:
            continue
        if not primer_tipo:
            r += 1  # espaciador
        primer_tipo = False

        tipo_color = TIPO_FG[tipo]
        soft_fill = PatternFill(start_color=TIPO_SOFT[tipo],
                                 end_color=TIPO_SOFT[tipo], fill_type='solid')

        # ── Barra de sección (fondo suave del tipo + texto del color) ───────
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        c_bar = ws.cell(r, 1, f"{tipo}    {TIPO_LABEL[tipo].upper()}")
        c_bar.font = Font(name='Inter', bold=True, size=11, color=tipo_color)
        c_bar.alignment = Alignment(horizontal='left', vertical='center',
                                     indent=1)
        for c in range(1, N + 1):
            ws.cell(r, c).fill = soft_fill
        ws.row_dimensions[r].height = 22
        r += 1

        # ── Items del tipo ──────────────────────────────────────────────────
        # Cols físicas: A=Cód · (B+C merged)=Desc · D=Und · E=Cant · F=Pre · G=Parc
        sub = 0.0
        num_font = Font(name='Inter', size=10, color=C_SLATE_900)
        for ins in lst:
            c_cod = ws.cell(r, 1, ins['codigo']); c_cod.font = num_font
            c_cod.alignment = Alignment(horizontal='left', vertical='top')
            c_dsc = ws.cell(r, 2, ins['descripcion'] or ''); c_dsc.font = num_font
            c_dsc.alignment = Alignment(horizontal='left', vertical='top',
                                         wrap_text=True)
            # PRE-estilar col 3 antes del merge (post-merge sería MergedCell)
            ws.cell(r, 3).font = num_font
            c_un = ws.cell(r, 4, ins['unidad'] or ''); c_un.font = num_font
            c_un.alignment = Alignment(horizontal='center', vertical='top')
            c_ct = ws.cell(r, 5, ins['cantidad_total'] or 0); c_ct.font = num_font
            c_ct.number_format = fmt_4
            c_ct.alignment = Alignment(horizontal='right', vertical='top')
            c_pr = ws.cell(r, 6, ins['precio'] or 0); c_pr.font = num_font
            c_pr.number_format = fmt_money
            c_pr.alignment = Alignment(horizontal='right', vertical='top')
            parcial = ins.get('parcial_total') or 0
            sub += parcial
            c_pa = ws.cell(r, 7, parcial); c_pa.font = num_font
            c_pa.number_format = fmt_money
            c_pa.alignment = Alignment(horizontal='right', vertical='top')
            _merge_descripcion(r)
            ws.row_dimensions[r].height = max(16, _alto_fila(
                ins['descripcion'] or '', 38))
            r += 1

        # ── Subtotal del tipo (label cols 1-6 right + valor en col 7) ───────
        subtotal_tipo[tipo] = sub
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c_sl = ws.cell(r, 1, f"Subtotal {TIPO_LABEL[tipo]}")
        c_sl.font = Font(name='Inter', bold=True, size=10, color=tipo_color)
        c_sl.alignment = Alignment(horizontal='right', vertical='center')
        c_sv = ws.cell(r, 7, sub)
        c_sv.font = Font(name='Inter', bold=True, size=10, color=tipo_color)
        c_sv.number_format = fmt_money
        c_sv.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[r].height = 18
        r += 1

    # ── Resumen por Tipo de Insumo (al pie) ─────────────────────────────────
    total_general = sum(subtotal_tipo.values())
    r += 3  # separación equivalente al margin-top:28pt del PDF

    # h3 "Resumen por Tipo de Insumo"
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    c_h3 = ws.cell(r, 1, 'Resumen por Tipo de Insumo')
    c_h3.font = Font(name='Inter', bold=True, size=11, color='2E3C52')
    c_h3.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 18
    r += 1

    # Cabecera resumen — FULL-WIDTH: Tipo (cols 1-5 merged) · Subtotal (col 6) · % CD (col 7).
    # bg accent_soft espejo del PDF para diferenciarse del cuerpo.
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    for col_ in range(1, N + 1):
        ws.cell(r, col_).fill = hdr_fill_i
    c_h_t = ws.cell(r, 1, 'Tipo')
    c_h_t.font = Font(name='Inter', bold=True, size=10, color=C_SLATE_900)
    c_h_t.alignment = Alignment(horizontal='left', vertical='center')
    c_h_s = ws.cell(r, 6, 'Subtotal')
    c_h_s.font = Font(name='Inter', bold=True, size=10, color=C_SLATE_900)
    c_h_s.alignment = Alignment(horizontal='right', vertical='center')
    c_h_p = ws.cell(r, 7, '% CD')
    c_h_p.font = Font(name='Inter', bold=True, size=10, color=C_SLATE_900)
    c_h_p.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[r].height = 16
    r += 1

    for tipo, label in [('MO', 'Mano de Obra'), ('MAT', 'Materiales'),
                        ('EQ', 'Equipos / Herramientas'),
                        ('SC', 'Sub-contratos / Servicios')]:
        v = subtotal_tipo.get(tipo, 0)
        pct = (v / total_general * 100) if total_general else 0
        # Tipo + label en cols 1-5 merged (full-width)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c_t = ws.cell(r, 1, f"{tipo}   {label}")
        c_t.font = Font(name='Inter', size=10, color=C_SLATE_900)
        c_t.alignment = Alignment(horizontal='left', vertical='center')
        # Subtotal en col 6
        c_s = ws.cell(r, 6, v)
        c_s.font = Font(name='Inter', bold=True, size=10, color=C_SLATE_900)
        c_s.number_format = fmt_money
        c_s.alignment = Alignment(horizontal='right', vertical='center')
        # % en col 7
        c_p = ws.cell(r, 7, f"{pct:.1f}%")
        c_p.font = Font(name='Inter', size=10, color=C_SLATE_500)
        c_p.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[r].height = 16
        r += 1

    # ── TOTAL COSTO DIRECTO — bold slate-900, SIN bordes (full-width) ───────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c_tl = ws.cell(r, 1, 'TOTAL COSTO DIRECTO')
    c_tl.font = Font(name='Inter', bold=True, size=11, color=C_SLATE_900)
    c_tl.alignment = Alignment(horizontal='left', vertical='center')
    c_tv = ws.cell(r, 6, total_general)
    c_tv.font = Font(name='Inter', bold=True, size=11, color=C_SLATE_900)
    c_tv.number_format = fmt_money
    c_tv.alignment = Alignment(horizontal='right', vertical='center')
    c_tp = ws.cell(r, 7, "100.0%")
    c_tp.font = Font(name='Inter', bold=True, size=11, color=C_SLATE_900)
    c_tp.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[r].height = 20
    r += 1

    _setup_impresion(ws, n_filas_encabezado=4, n_cols=N, proyecto=proyecto)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── EXCEL: Metrados ──────────────────────────────────────────────────────────

def _hoja_metrados(wb, proyecto_id):
    """Hoja "Metrados" del Excel — espejo visual del PDF `_html_metrados`.

    - Header 3 zonas + nombre proyecto full-width.
    - h2 "Hoja de Metrados" izq, color `od` de `accent_reportes` (slate-900 en sobrio).
    - Solo partidas HOJA con detalle (no títulos N1/N2 — el PDF tampoco los muestra).
    - Por partida: head bold + meta line italic + tabla 8 cols sin bordes excesivos.
    - Total de partida: bold slate-900, sin bordes.
    """
    conn     = get_db()
    proyecto = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    todas    = conn.execute(
        "SELECT * FROM partidas WHERE proyecto_id=? ORDER BY item", (proyecto_id,)
    ).fetchall()
    # Excluir partidas con acero (exclusivas de la hoja de acero)
    acero_ids = {r['partida_id'] for r in conn.execute(
        "SELECT DISTINCT partida_id FROM acero_detalle WHERE partida_id IN "
        "(SELECT id FROM partidas WHERE proyecto_id=?)", (proyecto_id,)
    ).fetchall()}
    # Solo partidas HOJA (no títulos) que no tengan acero
    partidas = [p for p in todas
                if not p['es_titulo'] and p['id'] not in acero_ids]

    ws = wb.create_sheet("Metrados")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines = False
    ws.print_options.gridLinesSet = True

    N = 8  # Descripción · N° estr. · N° elem. · Área · Largo · Ancho · Alto · Parcial
    ws.column_dimensions['A'].width = 32   # Descripción
    ws.column_dimensions['B'].width = 9    # N° estr.
    ws.column_dimensions['C'].width = 9    # N° elem.
    ws.column_dimensions['D'].width = 10   # Área
    ws.column_dimensions['E'].width = 10   # Largo
    ws.column_dimensions['F'].width = 10   # Ancho
    ws.column_dimensions['G'].width = 10   # Alto
    ws.column_dimensions['H'].width = 14   # Parcial

    # ── Encabezado 3 zonas + nombre proyecto en row 2 centro ────────────────
    # cols_partition=(1, 6) → Left=A (32) · Center=B-F (48) · Right=G-H (24).
    # Col A "Descripción" es ancha → cabe holgado "ingePresupuestos" y el
    # subtítulo. El centro (B-F) toma 5 cols numéricas estrechas que juntas
    # dan espacio holgado para el nombre del proyecto.
    r = _xlsx_header_pdf_style(ws, proyecto, 'Hoja de Metrados', N,
                                  cols_partition=(1, 6))

    # ── Título h2 — color `od` de accent_reportes (slate-900 en sobrio) ─────
    from utils.theme import accent_reportes
    C_SLATE_900 = '1F2A38'
    C_SLATE_500 = '64748B'
    C_SLATE_300 = 'CBD5E1'
    _h2_color = accent_reportes()[1].lstrip('#').upper()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    c_h2 = ws.cell(r, 1, 'Hoja de Metrados')
    c_h2.font      = Font(name='Inter', bold=True, size=14, color=_h2_color)
    c_h2.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 22
    r += 1

    side_slate_300 = Side(style='thin', color=C_SLATE_300)
    _dm = get_decimales_metrado()
    fmt_num = ('[$-0409]#,##0.' + '0' * _dm) if _dm else '[$-0409]#,##0'

    # Bg suave del accent (espejo del PDF `acu-head` y ACU del Excel) para
    # el "card header" de cada partida (líneas item+descripción y meta).
    from openpyxl.styles import PatternFill
    from utils.theme import accent_reportes as _accent_card_m
    _, _, _os_card_m = _accent_card_m()
    C_CARD_BG = _os_card_m.lstrip('#').upper()
    card_fill = PatternFill(start_color=C_CARD_BG, end_color=C_CARD_BG,
                              fill_type='solid')

    for partida in partidas:
        detalles = conn.execute(
            "SELECT * FROM metrados_detalle WHERE partida_id=? ORDER BY orden",
            (partida['id'],)
        ).fetchall()
        # Fallback: metrado directo (sin planilla) → fila sintética con
        # el valor en "N° elem" (mismo criterio que el PDF)
        if not detalles and (partida['metrado'] or 0):
            detalles = [{
                'descripcion':   'Metrado directo',
                'n_estructuras': None,
                'n_elementos':   partida['metrado'] or 0,
                'area':          None, 'largo': None,
                'ancho':         None, 'alto':  None,
                'parcial':       partida['metrado'] or 0,
            }]
        # Saltar si la partida no tiene detalle ni metrado directo (como el PDF)
        if not detalles:
            continue

        # ── Header de la partida (card con bg accent_soft) ──────────────────
        # Línea 1: ítem + descripción (slate-900 bold) — PRE-aplicar fill
        # en TODAS las cols antes de mergear (col 2..N quedan MergedCell).
        for col_ in range(1, N + 1):
            ws.cell(r, col_).fill = card_fill
        head_txt = f"{partida['item'] or ''}   {partida['descripcion'] or ''}"
        c_head = ws.cell(r, 1, head_txt)
        c_head.font      = Font(name='Inter', bold=True, size=11, color=C_SLATE_900)
        c_head.alignment = Alignment(horizontal='left', vertical='center',
                                      wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        ws.row_dimensions[r].height = max(18, _alto_fila(head_txt, 90))
        r += 1

        # Línea 2: meta (Unidad · Metrado total) — espejo del ACU: regular
        # slate-500 con valores en BOLD (CellRichText), NO italic.
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        unidad = partida['unidad'] or '—'
        metrado_total = partida['metrado'] or 0
        f_reg  = InlineFont(rFont='Inter', sz=9, color=C_SLATE_500)
        f_bold = InlineFont(rFont='Inter', sz=9, color=C_SLATE_500, b=True)
        meta_rich = CellRichText(
            TextBlock(f_reg,  'Unidad: '),
            TextBlock(f_bold, unidad),
            TextBlock(f_reg,  '   ·   Metrado total: '),
            TextBlock(f_bold, f'{metrado_total:,.{_dm}f}'),
        )
        for col_ in range(1, N + 1):
            ws.cell(r, col_).fill = card_fill
        c_meta = ws.cell(r, 1)
        c_meta.value = meta_rich
        c_meta.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        ws.row_dimensions[r].height = 14
        r += 1

        # ── Cabecera de columnas (border-bottom slate-300 thin) ─────────────
        hdrs = [('Descripción', 'left'), ('N° estr.', 'right'),
                ('N° elem.', 'right'), ('Área', 'right'),
                ('Largo', 'right'), ('Ancho', 'right'),
                ('Alto', 'right'), ('Parcial', 'right')]
        for c, (h, al) in enumerate(hdrs, 1):
            cell = ws.cell(r, c, h)
            cell.font      = Font(name='Inter', bold=True, size=10, color=C_SLATE_500)
            cell.alignment = Alignment(horizontal=al, vertical='center')
            cell.border    = Border(bottom=side_slate_300)
        ws.row_dimensions[r].height = 16
        r += 1

        # ── Filas de detalle (zebra sutil `#FBFCFD` en filas impares) ───────
        total_partida = 0.0
        for det_idx, det in enumerate(detalles):
            det = dict(det)
            zebra = (det_idx % 2 == 1)
            zebra_fill = (PatternFill(start_color='FBFCFD',
                                          end_color='FBFCFD',
                                          fill_type='solid')
                          if zebra else None)
            num_font = Font(name='Inter', size=10, color=C_SLATE_900)
            c_d = ws.cell(r, 1, det.get('descripcion') or '')
            c_d.font = num_font
            c_d.alignment = Alignment(horizontal='left', vertical='top',
                                       wrap_text=True)
            for col, key in [(2, 'n_estructuras'), (3, 'n_elementos'),
                             (4, 'area'), (5, 'largo'),
                             (6, 'ancho'), (7, 'alto')]:
                v = det.get(key)
                cn = ws.cell(r, col)
                if v:
                    cn.value = v
                    cn.font = num_font
                    cn.number_format = fmt_num
                    cn.alignment = Alignment(horizontal='right', vertical='top')
                else:
                    cn.font = num_font
            parcial = det.get('parcial') or 0
            total_partida += parcial
            c_p = ws.cell(r, 8, parcial); c_p.font = num_font
            c_p.number_format = fmt_num
            c_p.alignment = Alignment(horizontal='right', vertical='top')
            if zebra_fill:
                for c_ in range(1, N + 1):
                    ws.cell(r, c_).fill = zebra_fill
            ws.row_dimensions[r].height = max(16, _alto_fila(
                det.get('descripcion') or '', 30))
            r += 1

        # ── Total de partida — bold slate-900, SIN bordes ───────────────────
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c_tl = ws.cell(r, 1, f"Total  ({unidad})")
        c_tl.font = Font(name='Inter', bold=True, size=11, color=C_SLATE_900)
        c_tl.alignment = Alignment(horizontal='right', vertical='center')
        c_tv = ws.cell(r, 8, total_partida)
        c_tv.font = Font(name='Inter', bold=True, size=11, color=C_SLATE_900)
        c_tv.number_format = fmt_num
        c_tv.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[r].height = 20
        r += 1

        # Espacio entre partidas
        r += 2

    conn.close()
    # Solo repetir las 3 filas del tripartito en cada página — el h2 y la
    # cabecera de columnas se repiten POR partida dentro del cuerpo.
    _setup_impresion(ws, n_filas_encabezado=3, n_cols=N, proyecto=proyecto)


# ─── EXCEL: Metrados de Acero ────────────────────────────────────────────────

def _hoja_acero_metrados(wb, proyecto_id):
    """Hoja "Acero" del Excel — espejo visual del PDF (rama de aceros en `_html_metrados`).

    - Header 3 zonas + nombre proyecto full-width.
    - h2 "Hoja de Metrados de Acero" izq, color `od` de `accent_reportes`.
    - Solo partidas HOJA con detalle de acero (igual que el PDF).
    - Por partida: head bold + meta line italic + tabla 7 cols sin bordes excesivos.
    - Total kg de partida: bold slate-900, sin bordes.
    """
    conn     = get_db()
    proyecto = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    todas    = conn.execute(
        "SELECT * FROM partidas WHERE proyecto_id=? ORDER BY item", (proyecto_id,)
    ).fetchall()
    acero_ids = {r['partida_id'] for r in conn.execute(
        "SELECT DISTINCT partida_id FROM acero_detalle WHERE partida_id IN "
        "(SELECT id FROM partidas WHERE proyecto_id=?)", (proyecto_id,)
    ).fetchall()}
    if not acero_ids:
        conn.close()
        return  # no hay datos de acero, no crear hoja
    # Solo partidas HOJA con acero (sin títulos N1/N2 — espejo del PDF)
    partidas = [p for p in todas if not p['es_titulo'] and p['id'] in acero_ids]

    ws = wb.create_sheet("Acero")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines = False
    ws.print_options.gridLinesSet = True

    N = 7  # Descripción · Ø · Veces · N° elem. · Longitud · kg/m · Parcial
    ws.column_dimensions['A'].width = 28  # Descripción
    ws.column_dimensions['B'].width = 8   # Ø
    ws.column_dimensions['C'].width = 9   # Veces
    ws.column_dimensions['D'].width = 9   # N° elem.
    ws.column_dimensions['E'].width = 11  # Longitud
    ws.column_dimensions['F'].width = 10  # kg/m
    ws.column_dimensions['G'].width = 13  # Parcial

    # ── Encabezado 3 zonas + nombre proyecto en row 2 centro ────────────────
    # cols_partition=(1, 5) → Left=A (28) · Center=B-E (37) · Right=F-G (23).
    # Col A "Descripción" ya es ancha (28) → suficiente para "ingePresupuestos".
    r = _xlsx_header_pdf_style(ws, proyecto, 'Hoja de Metrados de Acero', N,
                                  cols_partition=(1, 5))

    # ── Título h2 ───────────────────────────────────────────────────────────
    from utils.theme import accent_reportes
    _h2_color = accent_reportes()[1].lstrip('#').upper()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    c_h2 = ws.cell(r, 1, 'Hoja de Metrados de Acero')
    c_h2.font      = Font(name='Inter', bold=True, size=14, color=_h2_color)
    c_h2.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 22
    r += 1

    side_slate_300 = Side(style='thin', color='CBD5E1')
    fmt_num = '#,##0.00'
    fmt_4   = '#,##0.0000'

    for partida in partidas:
        detalles = conn.execute(
            "SELECT * FROM acero_detalle WHERE partida_id=? ORDER BY orden",
            (partida['id'],)
        ).fetchall()
        if not detalles:
            continue

        # Total kg de la partida (suma de parcial)
        detalles_tot = sum(d['parcial'] or 0 for d in detalles)

        # ── Header de la partida ────────────────────────────────────────────
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        head_txt = f"{partida['item'] or ''}   {partida['descripcion'] or ''}"
        c_head = ws.cell(r, 1, head_txt)
        c_head.font      = Font(name='Inter', bold=True, size=11, color='1F2A38')
        c_head.alignment = Alignment(horizontal='left', vertical='center',
                                      wrap_text=True)
        ws.row_dimensions[r].height = max(18, _alto_fila(head_txt, 75))
        r += 1

        # Línea 2: meta (Unidad · Metrado total) en italic slate-500
        unidad = partida['unidad'] or 'kg'
        meta = (f"Unidad: {unidad}   ·   "
                f"Metrado total: {detalles_tot:,.3f} kg")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        c_meta = ws.cell(r, 1, meta)
        c_meta.font = Font(name='Inter', italic=True, size=9, color='64748B')
        c_meta.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[r].height = 14
        r += 1

        # ── Cabecera de columnas (border-bottom slate-300 thin) ─────────────
        hdrs = [('Descripción', 'left'), ('Ø', 'center'),
                ('Veces', 'right'), ('N° elem.', 'right'),
                ('Longitud', 'right'), ('kg/m', 'right'),
                ('Parcial', 'right')]
        for c, (h, al) in enumerate(hdrs, 1):
            cell = ws.cell(r, c, h)
            cell.font      = Font(name='Inter', bold=True, size=10, color='64748B')
            cell.alignment = Alignment(horizontal=al, vertical='center')
            cell.border    = Border(bottom=side_slate_300)
        ws.row_dimensions[r].height = 16
        r += 1

        # ── Filas de detalle ────────────────────────────────────────────────
        for det in detalles:
            det = dict(det)
            num_font = Font(name='Inter', size=10, color='1F2A38')
            c_d = ws.cell(r, 1, det.get('descripcion') or '')
            c_d.font = num_font
            c_d.alignment = Alignment(horizontal='left', vertical='top',
                                       wrap_text=True)
            c_dia = ws.cell(r, 2, det.get('diametro') or ''); c_dia.font = num_font
            c_dia.alignment = Alignment(horizontal='center', vertical='top')
            # Veces (n_veces), N° elem. (n_elementos)
            for col, key in [(3, 'n_veces'), (4, 'n_elementos')]:
                v = det.get(key)
                if v is not None:
                    cn = ws.cell(r, col, v); cn.font = num_font
                    cn.number_format = fmt_num
                    cn.alignment = Alignment(horizontal='right', vertical='top')
            # Longitud
            if det.get('longitud') is not None:
                cn = ws.cell(r, 5, det['longitud']); cn.font = num_font
                cn.number_format = fmt_4
                cn.alignment = Alignment(horizontal='right', vertical='top')
            # kg/m
            if det.get('kg_ml'):
                cn = ws.cell(r, 6, det['kg_ml']); cn.font = num_font
                cn.number_format = fmt_4
                cn.alignment = Alignment(horizontal='right', vertical='top')
            # Parcial
            parcial = det.get('parcial') or 0
            c_p = ws.cell(r, 7, parcial); c_p.font = num_font
            c_p.number_format = fmt_4
            c_p.alignment = Alignment(horizontal='right', vertical='top')
            ws.row_dimensions[r].height = max(16, _alto_fila(
                det.get('descripcion') or '', 28))
            r += 1

        # ── Total kg — bold slate-900, SIN bordes ───────────────────────────
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c_tl = ws.cell(r, 1, 'Total acero (kg)')
        c_tl.font = Font(name='Inter', bold=True, size=11, color='1F2A38')
        c_tl.alignment = Alignment(horizontal='right', vertical='center')
        c_tv = ws.cell(r, 7, detalles_tot)
        c_tv.font = Font(name='Inter', bold=True, size=11, color='1F2A38')
        c_tv.number_format = fmt_4
        c_tv.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[r].height = 20
        r += 1

        # Espacio entre partidas
        r += 2

    conn.close()
    # Solo repetir las 3 filas del tripartito en cada página — espejo del
    # comportamiento de la Hoja de Metrados (sister sheet).
    _setup_impresion(ws, n_filas_encabezado=3, n_cols=N, proyecto=proyecto)


# ─── EXCEL: Especificaciones ──────────────────────────────────────────────────

def _hoja_especificaciones(wb, proyecto_id):
    conn     = get_db()
    proyecto = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    partidas = conn.execute(
        """SELECT * FROM partidas WHERE proyecto_id=? AND es_titulo=0
           AND especificaciones IS NOT NULL AND especificaciones != ''
           ORDER BY item""", (proyecto_id,)
    ).fetchall()
    conn.close()

    ws = wb.create_sheet("Especificaciones")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    ws.merge_cells('A1:D1')
    ws['A1'] = 'ESPECIFICACIONES TÉCNICAS'
    ws['A1'].font      = Font(name='Inter', bold=True, italic=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.cell(2, 1, 'Proyecto').font = Font(name='Inter',bold=True, italic=True, size=11)
    ws.merge_cells('B2:D2')
    ws.cell(2, 2, proyecto['nombre']).font = Font(name='Inter', size=11)

    r = 4
    if not partidas:
        ws.merge_cells('A4:D4')
        ws.cell(4, 1, 'Sin especificaciones técnicas generadas.').font = Font(name='Inter',italic=True)
    else:
        for partida in partidas:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            unidad_str = f" ({partida['unidad']})" if partida['unidad'] else ''
            ws.cell(r, 1, f"{partida['item']}  {partida['descripcion']}{unidad_str}")
            ws.cell(r, 1).font      = Font(name='Inter',bold=True, size=11)
            ws.cell(r, 1).alignment = Alignment(wrap_text=True)
            r += 1
            spec_text = partida['especificaciones'] or ''
            for parrafo in [p.strip() for p in spec_text.split('\n') if p.strip()]:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                cell = ws.cell(r, 1, parrafo)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border    = thin_border()
                ws.row_dimensions[r].height = 15 * max(1, len(parrafo) // 120 + 1)
                r += 1
            r += 1

    ws.column_dimensions['A'].width = 120
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    _setup_impresion(ws)


# ─── EXCEL: Gastos Generales ──────────────────────────────────────────────────

def _hoja_gastos_generales(wb, proyecto_id):
    conn      = get_db()
    proyecto  = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    _, totales = calcular_totales(proyecto_id)
    cd        = totales['cd']
    rubros_pie, total_final = _calcular_rubros_pie(conn, proyecto_id, cd)

    ws = wb.create_sheet("Pie de Presupuesto")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Espejo del PDF: header de columnas sombreado (gris suave accent_soft),
    # títulos de rubro en rojo, grupos subrayados, items con sangría (tab).
    from openpyxl.styles import PatternFill
    from utils.theme import accent_reportes as _accent_gg
    _, _, _hdr_soft = _accent_gg()
    _hdr_soft = _hdr_soft.lstrip('#').upper()
    _hdr_fill = PatternFill(start_color=_hdr_soft, end_color=_hdr_soft,
                            fill_type='solid')
    # Subtotales/total: solo línea superior (espejo del PDF `border-top` slate),
    # NO cuadro. COSTO DIRECTO y filas de resumen van SIN borde, como el PDF.
    # El header de columnas solo lleva sombreado (el PDF no dibuja línea abajo).
    _b_top = Border(top=Side(style='thin', color='2E3C52'))

    # Columnas: Descripción(1-2 merge) · Unidad(3) · % Participación(4) ·
    #           Cantidad(5) · Precio(6) · Total(7).
    N = 7
    # Encabezado tripartito estilo PDF (igual que Presupuesto): empresa | título
    # +proyecto | costo+modalidad. cols_partition (2,5): L=A-B, C=C-E, R=F-G.
    r = _xlsx_header_pdf_style(ws, proyecto, 'Desagregado del Pie de Presupuesto',
                              N, cols_partition=(2, 5))
    # El título es largo y la zona central (cols C-E) es angosta → a size 11
    # se parte en 2 líneas y se encima con el nombre del proyecto. Lo bajamos a
    # size 9 para que entre en una sola línea (C_START=3 con partition (2,5)).
    ws.cell(1, 3).font = Font(name='Inter', bold=True, size=10, color='1F2A38')
    filas_enc = r - 1   # filas del encabezado a repetir al imprimir

    # Costo Directo
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N - 1)
    ws.cell(r, 1, 'COSTO DIRECTO').font = Font(name='Inter',bold=True, size=11)
    ws.cell(r, 1).alignment = Alignment(horizontal='right')
    ws.cell(r, N, cd).number_format = '[$-0409]#,##0.00'
    ws.cell(r, N).font = Font(name='Inter',bold=True, size=11)
    # COSTO DIRECTO sin borde (espejo del PDF)
    r += 1

    if not rubros_pie:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N - 1)
        ws.cell(r, 1, f"GASTOS GENERALES ({proyecto['gf_pct']}% del CD)").font = Font(name='Inter',bold=True, size=11)
        ws.cell(r, 1).alignment = Alignment(horizontal='right')
        ws.cell(r, N, totales['gf']).number_format = '[$-0409]#,##0.00'
        ws.cell(r, N).font = Font(name='Inter',bold=True, size=11)
        for c in range(1, N + 1):
            ws.cell(r, c).border = _b_top
        conn.close()
        return

    for rub in rubros_pie:
        if rub['tipo'] in ('subtotal', 'pct_cd', 'pct_sub'):
            is_bold = rub['tipo'] == 'subtotal' or rub['nombre'].upper().startswith('TOTAL')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N - 1)
            ws.cell(r, 1, rub['nombre'].upper()).font = Font(name='Inter',bold=is_bold, size=11)
            ws.cell(r, 1).alignment = Alignment(horizontal='right')
            ws.cell(r, N, rub['valor']).number_format = '[$-0409]#,##0.00'
            ws.cell(r, N).font = Font(name='Inter',bold=is_bold, size=11)
            if is_bold:                       # subtotal → solo línea superior
                for c in range(1, N + 1):
                    ws.cell(r, c).border = _b_top
            r += 1
            continue

        # Tipo 'rubro': encabezado de sección — título en rojo (espejo del PDF)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        ws.cell(r, 1, rub['nombre'].upper()).font = Font(name='Inter', bold=True,
                                                         size=11, color='B71C1C')
        ws.cell(r, 1).alignment = Alignment(horizontal='left', indent=1)
        # Sin borde en el título de rubro (espejo del PDF: solo texto rojo).
        r += 1

        gg_items = conn.execute(
            "SELECT * FROM gastos_generales WHERE proyecto_id=? AND rubro=? ORDER BY orden",
            (proyecto_id, rub['codigo'])
        ).fetchall()

        if gg_items:
            hdrs_gg = ['Descripción', '', 'Unidad', '% Particip.',
                       'Cantidad', 'Precio', 'Total']
            for c, h in enumerate(hdrs_gg, 1):
                cell = ws.cell(r, c, h)
                cell.font      = Font(name='Inter',bold=True, italic=True, size=11)
                cell.alignment = Alignment(horizontal='center')
                cell.fill      = _hdr_fill   # solo sombreado (espejo del PDF, sin línea)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            r += 1

            grupo_activo = False
            for item in gg_items:
                if item['tipo'] not in ('grupo', 'item'):
                    continue   # filas 'manual' no se listan en el detalle
                if item['tipo'] == 'grupo':
                    grupo_activo = True
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
                    # Grupo subrayado (espejo del PDF) — sin borde/cuadrícula
                    ws.cell(r, 1, item['descripcion']).font = Font(
                        name='Inter', bold=True, size=11, underline='single')
                    r += 1
                else:
                    pctp = item['pct_participacion'] if item['pct_participacion'] is not None else 100
                    cant = item['cantidad'] or 0
                    pr_  = item['precio']   or 0
                    total = cant * (pctp / 100) * pr_

                    ws.cell(r, 1, item['descripcion']).font = Font(name='Inter', size=11)
                    # Tab del item: sangría 1 nivel si está bajo un grupo
                    ws.cell(r, 1).alignment = Alignment(
                        horizontal='left', indent=(1 if grupo_activo else 0))
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                    ws.cell(r, 3, item['unidad'] or '').font = Font(name='Inter', size=11)
                    ws.cell(r, 3).alignment = Alignment(horizontal='center')
                    ws.cell(r, 4, pctp).number_format = '#,##0.00'
                    if cant: ws.cell(r, 5, cant).number_format = '#,##0.00'
                    ws.cell(r, 6, pr_).number_format  = '#,##0.00'
                    ws.cell(r, 7, total).number_format = '[$-0409]#,##0.00'
                    # Sin borde/cuadrícula en los items (espejo del PDF).
                    r += 1

        # Subtotal del rubro
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N - 1)
        ws.cell(r, 1, rub['nombre'].upper()).font = Font(name='Inter',bold=True, size=11)
        ws.cell(r, 1).alignment = Alignment(horizontal='right')
        ws.cell(r, N, rub['valor']).number_format = '[$-0409]#,##0.00'
        ws.cell(r, N).font = Font(name='Inter',bold=True, size=11)
        for c in range(1, N + 1):     # subtotal del rubro → solo línea superior
            ws.cell(r, c).border = _b_top
        r += 1

    # Total final → solo línea superior (sin cuadro), espejo del PDF
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N - 1)
    ws.cell(r, 1, 'COSTO TOTAL DE OBRA').font = Font(name='Inter',bold=True, size=11)
    ws.cell(r, 1).alignment = Alignment(horizontal='right')
    ws.cell(r, N, total_final).number_format = '[$-0409]#,##0.00'
    ws.cell(r, N).font = Font(name='Inter',bold=True, size=11)
    for c in range(1, N + 1):
        ws.cell(r, c).border = _b_top

    conn.close()
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    # Pie tripartito + área/centrado igual que Presupuesto (espejo del PDF).
    _setup_impresion(ws, n_filas_encabezado=filas_enc, n_cols=N, proyecto=proyecto)


def exportar_valorizacion(val_id):
    """Export Excel de una valorización (Base·Anterior·Actual·Acumulado·Saldo por
    partida, subtotales de título y fila TOTAL). Devuelve un buffer BytesIO."""
    from openpyxl.styles import PatternFill
    import core.valorizacion as _val
    from core.database import get_decimales_ppto
    val = _val.get_valorizacion(val_id)
    if not val:
        raise ValueError("Valorización no encontrada")
    conn = get_db()
    proyecto = conn.execute("SELECT * FROM proyectos WHERE id=?",
                            (val['proyecto_id'],)).fetchone()
    conn.close()
    filas, resumen = _val.get_valorizacion_detalle(val_id)
    dm = get_decimales_metrado(); dp = get_decimales_ppto()
    C_SLATE_700 = '2E3C52'; C_HEAD = '273445'; C_TOTAL = 'F0F1F2'
    C_TIT = {1: 'B71C1C', 2: '0D52BF', 3: '6A1B9A', 4: 'AD1457'}

    def _m(v):
        return round(v, dm) if v is not None else None

    def _d(v):
        return round(v, dp) if v is not None else None

    from openpyxl.styles import Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Valorización"
    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines = False; ws.print_options.gridLinesSet = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    N = 18
    GHBG = 'F1F5F9'; BANANA = 'FFF9CC'; ACCENT = 'F37329'
    r = _xlsx_header_pdf_style(ws, proyecto, f"Valorización N° {val['numero']}", N)

    def _fch(iso):
        p = str(iso or '').split('-')
        return f"{p[2]}/{p[1]}/{p[0]}" if (len(p) == 3 and len(p[0]) == 4) else str(iso or '')

    sub = []
    if val.get('periodo_desde') or val.get('periodo_hasta'):
        sub.append(f"Período: {_fch(val['periodo_desde'])} al {_fch(val['periodo_hasta'])}")
    sub.append(f"Avance físico acumulado: {resumen.get('pct_fisico', 0):.2f}%")
    scell = ws.cell(r, 1, "   ·   ".join(sub))
    scell.font = Font(name='Inter', size=10, color=C_SLATE_700, italic=True)
    scell.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    r += 2

    # ── Cabecera de 2 niveles ────────────────────────────────────────────────
    # Columnas: 1 Ítem · 2 Descripción · 3 Und · grupos de 3 (Metrado/Valor./%).
    groups = [('Contractual (Base)', ['Metrado', 'P.U.', 'Parcial'], False),
              ('Anterior', ['Metrado', 'Valorizado', '%'], False),
              ('Actual', ['Metrado', 'Valorizado', '%'], True),
              ('Acumulado', ['Metrado', 'Valorizado', '%'], False),
              ('Saldo', ['Metrado', 'Valorizado', '%'], False)]
    r_grp, r_sub = r, r + 1
    bot = Border(bottom=Side(style='medium', color=ACCENT))
    for c, h in ((1, 'Ítem'), (2, 'Descripción'), (3, 'Und')):
        ws.merge_cells(start_row=r_grp, start_column=c, end_row=r_sub, end_column=c)
        cell = ws.cell(r_grp, c, h)
        cell.font = Font(name='Inter', size=9, bold=True, color='1F2A38')
        cell.fill = PatternFill('solid', fgColor=GHBG)
        cell.alignment = Alignment(horizontal='center' if c == 3 else 'left',
                                   vertical='center', wrap_text=True)
        ws.cell(r_sub, c).border = bot
        cell.border = Border(bottom=Side(style='medium', color=ACCENT))
    col = 4
    grp_end_cols = {3}   # cols con borde derecho (fin de grupo) → separador
    for glabel, subs, banana in groups:
        span = len(subs)
        fill = BANANA if banana else GHBG
        ws.merge_cells(start_row=r_grp, start_column=col,
                       end_row=r_grp, end_column=col + span - 1)
        g = ws.cell(r_grp, col, glabel)
        g.font = Font(name='Inter', size=9, bold=True, color='1F2A38')
        g.fill = PatternFill('solid', fgColor=fill)
        g.alignment = Alignment(horizontal='center', vertical='center')
        for i, sname in enumerate(subs):
            s = ws.cell(r_sub, col + i, sname)
            s.font = Font(name='Inter', size=8, bold=True, color='2E3C52')
            s.fill = PatternFill('solid', fgColor=fill)
            s.alignment = Alignment(horizontal='right', vertical='center',
                                    wrap_text=True)
            s.border = bot
        grp_end_cols.add(col + span - 1)
        col += span
    r += 2
    data_start = r

    # Bordes reutilizables para las filas.
    grid = Side(style='thin', color='E2E8F0')
    sep = Side(style='thin', color='94A3B8')

    # Metadatos por columna (1-based).
    METR = {4, 7, 10, 13, 16}; PU = {5}; VAL = {6, 8, 11, 14, 17}
    PCT = {9, 12, 15, 18}; ACT = {10, 11, 12}

    def _fila(f, tot=False):
        nonlocal r
        es_tit = bool(f.get('es_titulo')) and not tot
        base = f.get('base_val') or 0

        def _pv(v):
            return ((v or 0) / base * 100) if base else 0

        blank = (es_tit or tot)
        vals = [
            f.get('item') or '', f.get('descripcion') or '',
            '' if blank else (f.get('unidad') or ''),
            None if blank else _m(f.get('base_metr')),
            None if blank else _d(f.get('precio_unitario')), _d(f.get('base_val')),
            None if blank else _m(f.get('ant_metr')), _d(f.get('ant_val')),
            _pv(f.get('ant_val')),
            None if blank else _m(f.get('act_metr')), _d(f.get('act_val')),
            _pv(f.get('act_val')),
            None if blank else _m(f.get('acu_metr')), _d(f.get('acu_val')),
            (f.get('pct') or 0),
            None if blank else _m(f.get('sal_metr')), _d(f.get('sal_val')),
            _pv(f.get('sal_val'))]
        # Color de fuente: en títulos, SOLO Ítem/Descripción llevan color de
        # nivel; los valores van en negro.
        niv_col = C_TIT.get(min(max(f.get('nivel', 1), 1), 4), C_SLATE_700)
        base_fill = (PatternFill('solid', fgColor=C_TOTAL) if tot else None)
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c)
            if c in PCT:
                cell.value = v
                cell.number_format = '#,##0.00"%"'
            else:
                cell.value = v
                if v is not None and (c in METR or c in PU or c in VAL):
                    nd = dp if (c in PU or c in VAL) else dm
                    cell.number_format = f'#,##0.{"0"*nd}'
            if es_tit:
                fcol = niv_col if c in (1, 2) else '1F2A38'
            elif tot:
                fcol = '1F2A38'
            else:
                fcol = C_SLATE_700
            cell.font = Font(name='Inter', size=9, bold=(es_tit or tot), color=fcol)
            cell.alignment = Alignment(horizontal='right' if c >= 4
                                       else 'center' if c == 3 else 'left')
            # Relleno: banana en el grupo «Actual»; total gris.
            if base_fill and c not in ACT:
                cell.fill = base_fill
            elif c in ACT:
                cell.fill = PatternFill('solid', fgColor=BANANA)
            # Bordes: rejilla inferior + separador derecho al fin de grupo.
            cell.border = Border(bottom=grid,
                                 right=(sep if c in grp_end_cols else None))
        r += 1

    for f in filas:
        _fila(f)
    _fila({'item': '', 'descripcion': 'TOTAL', 'base_val': resumen.get('base_val'),
           'ant_val': resumen.get('ant_val'), 'act_val': resumen.get('act_val'),
           'acu_val': resumen.get('acu_val'), 'sal_val': resumen.get('sal_val'),
           'pct': resumen.get('pct_fisico', 0)}, tot=True)

    widths = [10, 38, 6, 9, 10, 13, 9, 13, 8, 9, 13, 8, 9, 13, 8, 9, 13, 8]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    # Inmovilizar las filas del encabezado (quedan fijas al hacer scroll).
    ws.freeze_panes = ws.cell(row=data_start, column=1)
    _setup_impresion(ws, n_cols=N, proyecto=proyecto)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def exportar_metrados(proyecto_id):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _hoja_metrados(wb, proyecto_id)
    _hoja_acero_metrados(wb, proyecto_id)  # solo crea hoja si hay datos de acero
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def exportar_gastos_generales(proyecto_id):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _hoja_gastos_generales(wb, proyecto_id)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── EXCEL: Reporte Completo (6 hojas) ────────────────────────────────────────

def exportar_reporte_completo(proyecto_id):
    conn     = get_db()
    proyecto = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    pie      = _get_pie(conn, proyecto_id)
    items, totales = calcular_totales(proyecto_id)
    rubros_pie, total_final = _calcular_rubros_pie(conn, proyecto_id, totales['cd'])
    conn.close()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 1. Presupuesto ────────────────────────────────────────────────────────
    ws_p = wb.create_sheet("Presupuesto")
    ws_p.page_setup.paperSize = ws_p.PAPERSIZE_A4
    N = 9
    r  = _xlsx_encabezado(ws_p, proyecto, 'PRESUPUESTO', N)

    hdrs_p = ['Ítem', 'Descripción', '', 'Und', 'Metrado', 'Precio', 'Parcial', 'Subtotal', 'Total']
    for c, h in enumerate(hdrs_p, 1):
        cell = ws_p.cell(r, c, h)
        cell.font      = Font(name='Inter',bold=True, italic=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = _borde_seccion()
    ws_p.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    r += 1

    for entry in items:
        p     = entry['partida']
        total = entry['total']
        if p['nivel'] == 1:
            ws_p.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            ws_p.cell(r, 1, p['item']).font = Font(name='Inter',bold=True, size=11)
            ws_p.cell(r, 2, p['descripcion']).font = Font(name='Inter',bold=True, size=11)
            ws_p.cell(r, 9, total).number_format = '[$-0409]#,##0.00'
            ws_p.cell(r, 9).font = Font(name='Inter',bold=True, size=11)
            for c in range(1, N + 1):
                ws_p.cell(r, c).border = _borde_titulo()
        elif p['es_titulo']:
            ws_p.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            ws_p.cell(r, 1, p['item']).font = Font(name='Inter',bold=True, italic=True, size=11)
            ws_p.cell(r, 2, p['descripcion']).font = Font(name='Inter',bold=True, italic=True, size=11)
            ws_p.cell(r, 8, total).number_format = '[$-0409]#,##0.00'
            ws_p.cell(r, 8).font = Font(name='Inter',bold=True, size=11)
            for c in range(1, N + 1):
                ws_p.cell(r, c).border = thin_border()
        else:
            ws_p.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            ws_p.cell(r, 1, p['item']).font = Font(name='Inter', size=11)
            ws_p.cell(r, 2, p['descripcion']).font = Font(name='Inter', size=11)
            ws_p.cell(r, 4, p['unidad']).font = Font(name='Inter', size=11)
            ws_p.cell(r, 5, p['metrado']).number_format = '[$-0409]#,##0.00'
            ws_p.cell(r, 6, p['precio_unitario']).number_format = '[$-0409]#,##0.00'
            ws_p.cell(r, 7, total).number_format = '[$-0409]#,##0.00'
            for c in range(1, N + 1):
                ws_p.cell(r, c).border = thin_border()
        r += 1

    r += 1

    def _fila_p(label, val, bold=True):
        nonlocal r
        ws_p.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws_p.cell(r, 1, label).font = Font(name='Inter',bold=bold, size=11)
        ws_p.cell(r, 1).alignment   = Alignment(horizontal='right')
        ws_p.cell(r, 9, val).number_format = '[$-0409]#,##0.00'
        ws_p.cell(r, 9).font = Font(name='Inter',bold=bold, size=11)
        for c in range(1, N + 1):
            ws_p.cell(r, c).border = _borde_total()
        r += 1

    if rubros_pie is not None:
        _fila_p('COSTO DIRECTO', totales['cd'])
        for rub in rubros_pie:
            is_b = rub['tipo'] == 'subtotal' or rub['nombre'].upper().startswith('TOTAL')
            _fila_p(rub['nombre'].upper(), rub['valor'], bold=is_b)
    else:
        for lb, v in [
            ('COSTO DIRECTO', totales['cd']),
            (f"GASTOS GENERALES ({proyecto['gf_pct']}%)", totales['gf']),
            (f"UTILIDAD ({proyecto['utilidad_pct']}%)", totales['utilidad']),
            ('SUB TOTAL', totales['subtotal']),
            (f"IGV ({proyecto['igv_pct']}%)", totales['igv']),
            ('PRESUPUESTO TOTAL', totales['total']),
        ]:
            _fila_p(lb, v)

    monto_final = totales['total'] if rubros_pie is None else total_final
    ws_p.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=N)
    ws_p.cell(r + 1, 1, f"Son :   {_monto_letras(monto_final)}").font = Font(name='Inter',bold=True, italic=True, size=11)
    _escribir_pie(ws_p, r + 2, pie)
    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'], [10, 42, 6, 7, 10, 12, 12, 12, 14]):
        ws_p.column_dimensions[col].width = w
    _setup_impresion(ws_p)

    # ── 2. ACUs ───────────────────────────────────────────────────────────────
    conn = get_db()
    ws_a = wb.create_sheet("ACUs")
    ws_a.page_setup.paperSize = ws_a.PAPERSIZE_A4
    Na = 10
    ra = _xlsx_encabezado(ws_a, proyecto, 'ANÁLISIS DE COSTOS UNITARIOS', Na)

    partidas_hoja = conn.execute(
        "SELECT * FROM partidas WHERE proyecto_id=? AND es_titulo=0 ORDER BY item",
        (proyecto_id,)
    ).fetchall()

    for partida in partidas_hoja:
        items_acu = conn.execute(
            """SELECT ai.*, r.codigo, r.descripcion as rdesc, r.tipo, r.unidad,
                      COALESCE(ai.precio, r.precio, 0) as precio
               FROM acu_items ai JOIN recursos r ON r.id=ai.recurso_id
               WHERE ai.partida_id=?""", (partida['id'],)
        ).fetchall()
        if not items_acu:
            continue

        ws_a.cell(ra, 1, f"Partida   {partida['item']}").font = Font(name='Inter', bold=True, size=11)
        ws_a.merge_cells(start_row=ra, start_column=2, end_row=ra, end_column=7)
        ws_a.cell(ra, 2, partida['descripcion']).font = Font(name='Inter', bold=True, size=11)
        rend_txt = f"Rend:  {(partida['rendimiento'] or 0):.4f}  {partida['unidad'] or ''}/DÍA"
        ws_a.merge_cells(start_row=ra, start_column=8, end_row=ra, end_column=Na)
        ws_a.cell(ra, 8, rend_txt).font = Font(name='Inter',italic=True, size=11)
        ws_a.cell(ra, 8).alignment = Alignment(horizontal='right')
        for c in range(1, Na + 1):
            ws_a.cell(ra, c).border = _borde_titulo()
        ra += 1

        hdrs_a = ['Código', 'Descripción Insumo', '', 'Unidad', 'Cuadrilla', 'Cantidad', 'Precio', '', 'Parcial', '']
        for c, h in enumerate(hdrs_a, 1):
            cell = ws_a.cell(ra, c, h)
            cell.font      = Font(name='Inter',bold=True, italic=True, size=11)
            cell.alignment = Alignment(horizontal='center')
            cell.border    = _borde_seccion()
        ws_a.merge_cells(start_row=ra, start_column=2, end_row=ra, end_column=3)
        ws_a.merge_cells(start_row=ra, start_column=8, end_row=ra, end_column=10)
        ws_a.cell(ra, 8, 'Parcial').font = Font(name='Inter',bold=True, italic=True, size=11)
        ws_a.cell(ra, 8).alignment = Alignment(horizontal='center')
        ra += 1

        for tipo, label in [('MO', 'Mano de Obra'), ('MAT', 'Materiales'),
                            ('EQ', 'Equipo'), ('SC', 'Sub-contratos')]:
            tipo_items = [x for x in items_acu if x['tipo'] == tipo]
            if not tipo_items:
                continue
            ws_a.merge_cells(start_row=ra, start_column=1, end_row=ra, end_column=Na)
            ws_a.cell(ra, 1, label).font = Font(name='Inter',bold=True, size=11)
            ws_a.cell(ra, 1).border = thin_border()
            ra += 1
            subtotal = 0
            for it in tipo_items:
                precio = it['precio'] or 0; cant = it['cantidad'] or 0
                parcial = precio * cant; subtotal += parcial
                ws_a.cell(ra, 1, it['codigo']).font = Font(name='Inter', size=11)
                ws_a.merge_cells(start_row=ra, start_column=2, end_row=ra, end_column=3)
                ws_a.cell(ra, 2, it['rdesc']).font = Font(name='Inter', size=11)
                ws_a.cell(ra, 4, it['unidad']).font = Font(name='Inter', size=11)
                ws_a.cell(ra, 5, it['cuadrilla']).number_format = '[$-0409]#,##0.0000'
                ws_a.cell(ra, 6, cant).number_format = '[$-0409]#,##0.0000'
                ws_a.cell(ra, 7, precio).number_format = '[$-0409]#,##0.00'
                ws_a.merge_cells(start_row=ra, start_column=8, end_row=ra, end_column=10)
                ws_a.cell(ra, 8, parcial).number_format = '[$-0409]#,##0.00'
                ws_a.cell(ra, 8).alignment = Alignment(horizontal='right')
                for c in range(1, Na + 1):
                    ws_a.cell(ra, c).border = thin_border()
                ra += 1
            ws_a.cell(ra, 10, subtotal).number_format = '[$-0409]#,##0.00'
            ws_a.cell(ra, 10).font = Font(name='Inter',bold=True, size=11)
            for c in range(1, Na + 1):
                ws_a.cell(ra, c).border = _borde_seccion()
            ra += 1

        ws_a.merge_cells(start_row=ra, start_column=1, end_row=ra, end_column=9)
        ws_a.cell(ra, 1, f"Costo Unitario por {partida['unidad'] or 'UND'} :").font = Font(name='Inter',bold=True, size=11)
        ws_a.cell(ra, 1).alignment = Alignment(horizontal='right')
        ws_a.cell(ra, 10, partida['precio_unitario']).number_format = '[$-0409]#,##0.00'
        ws_a.cell(ra, 10).font = Font(name='Inter',bold=True, size=11)
        for c in range(1, Na + 1):
            ws_a.cell(ra, c).border = _borde_total()
        ra += 3

    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'], [14, 36, 6, 8, 10, 12, 12, 6, 6, 14]):
        ws_a.column_dimensions[col].width = w
    _setup_impresion(ws_a)

    # ── 3. Insumos ────────────────────────────────────────────────────────────
    ws_i = wb.create_sheet("Insumos")
    Ni = 8
    ri = _xlsx_encabezado(ws_i, proyecto, 'LISTADO TOTAL DE INSUMOS', Ni)
    insumos = conn.execute(
        """SELECT r.codigo, r.descripcion, r.tipo, r.unidad,
                  AVG(COALESCE(ai.precio, r.precio, 0)) as precio,
                  SUM(ai.cantidad * p.metrado) as cantidad_total,
                  SUM(ai.cantidad * p.metrado * COALESCE(ai.precio, r.precio, 0)) as parcial
           FROM acu_items ai JOIN recursos r ON r.id=ai.recurso_id
           JOIN partidas p ON p.id=ai.partida_id
           WHERE p.proyecto_id=? AND p.es_titulo=0 AND SUBSTR(r.unidad,1,1)!='%'
           GROUP BY r.id
           ORDER BY r.tipo,
                    CASE WHEN r.tipo='MO' THEN mo_rank(r.descripcion) ELSE 0 END,
                    r.descripcion""",
        (proyecto_id,)
    ).fetchall()

    hdrs_i = ['IU  Código', 'Descripción', '', 'Unidad', 'Cantidad', 'Precio', 'Parcial', '']
    for c, h in enumerate(hdrs_i, 1):
        cell = ws_i.cell(ri, c, h)
        cell.font      = Font(name='Inter',bold=True, italic=True, size=11)
        cell.alignment = Alignment(horizontal='center')
        cell.border    = _borde_seccion()
    ws_i.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=3)
    ws_i.merge_cells(start_row=ri, start_column=7, end_row=ri, end_column=8)
    ri += 1

    TIPO_LABEL = {'MAT': 'MATERIALES', 'EQ': 'EQUIPO', 'MO': 'MANO DE OBRA',
                  'SC': 'SUB-CONTRATOS'}
    current_tipo_i = None; subtotales_i = {}
    for ins in insumos:
        tipo = ins['tipo']
        if tipo != current_tipo_i:
            current_tipo_i = tipo
            ws_i.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=8)
            ws_i.cell(ri, 1, TIPO_LABEL.get(tipo, tipo)).font = Font(name='Inter',bold=True, size=11)
            ws_i.cell(ri, 1).border = thin_border()
            ri += 1; subtotales_i[tipo] = 0
        ws_i.cell(ri, 1, ins['codigo']).font = Font(name='Inter', size=11)
        ws_i.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=3)
        ws_i.cell(ri, 2, ins['descripcion']).font = Font(name='Inter', size=11)
        ws_i.cell(ri, 4, ins['unidad']).font = Font(name='Inter', size=11)
        ws_i.cell(ri, 5, ins['cantidad_total']).number_format = '[$-0409]#,##0.00'
        ws_i.cell(ri, 6, ins['precio']).number_format = '[$-0409]#,##0.00'
        parcial_i = ins['parcial'] or 0
        ws_i.merge_cells(start_row=ri, start_column=7, end_row=ri, end_column=8)
        ws_i.cell(ri, 7, parcial_i).number_format = '[$-0409]#,##0.00'
        ws_i.cell(ri, 7).alignment = Alignment(horizontal='right')
        subtotales_i[tipo] = subtotales_i.get(tipo, 0) + parcial_i
        for c in range(1, Ni + 1): ws_i.cell(ri, c).border = thin_border()
        ri += 1

    ri += 1
    for tipo in ['MAT', 'EQ', 'MO', 'SC']:
        if tipo in subtotales_i:
            ws_i.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=6)
            ws_i.cell(ri, 1, f"Sub Total {TIPO_LABEL[tipo]}").font = Font(name='Inter',bold=True, size=11)
            ws_i.cell(ri, 1).alignment = Alignment(horizontal='right')
            ws_i.cell(ri, 7, subtotales_i[tipo]).number_format = '[$-0409]#,##0.00'
            ws_i.cell(ri, 7).font = Font(name='Inter',bold=True, size=11)
            for c in range(1, Ni + 1):
                ws_i.cell(ri, c).border = _borde_total()
            ri += 1

    ri += 1

    def _fila_i(label, val, bold=True):
        nonlocal ri
        ws_i.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=7)
        ws_i.cell(ri, 1, label).font = Font(name='Inter',bold=bold, size=11)
        ws_i.cell(ri, 1).alignment   = Alignment(horizontal='right')
        ws_i.cell(ri, 8, val).number_format = '[$-0409]#,##0.00'
        ws_i.cell(ri, 8).font = Font(name='Inter',bold=bold, size=11)
        for c in range(1, Ni + 1):
            ws_i.cell(ri, c).border = _borde_total()
        ri += 1

    if rubros_pie is not None:
        _fila_i('COSTO DIRECTO', totales['cd'])
        for rub in rubros_pie:
            is_b = rub['tipo'] == 'subtotal' or rub['nombre'].upper().startswith('TOTAL')
            _fila_i(rub['nombre'].upper(), rub['valor'], bold=is_b)
    else:
        for lb, v in [('COSTO DIRECTO', totales['cd']),
                      (f"GG ({proyecto['gf_pct']}%)", totales['gf']),
                      ('TOTAL', totales['total'])]:
            _fila_i(lb, v)

    ws_i.merge_cells(start_row=ri + 1, start_column=1, end_row=ri + 1, end_column=Ni)
    ws_i.cell(ri + 1, 1, f"Son :   {_monto_letras(monto_final)}").font = Font(name='Inter',bold=True, italic=True, size=11)
    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'], [14, 36, 6, 8, 12, 12, 12, 6]):
        ws_i.column_dimensions[col].width = w
    _setup_impresion(ws_i)

    conn.close()

    # ── 4. Metrados ───────────────────────────────────────────────────────────
    _hoja_metrados(wb, proyecto_id)

    # ── 5. Metrados de Acero (solo si hay datos) ──────────────────────────────
    _hoja_acero_metrados(wb, proyecto_id)

    # ── 6. Gastos Generales ───────────────────────────────────────────────────
    _hoja_gastos_generales(wb, proyecto_id)

    # Insumos antes que ACUs (mismo orden que el Centro de Reportes)
    try:
        if 'Insumos' in wb.sheetnames and 'ACUs' in wb.sheetnames:
            wb.move_sheet('Insumos',
                          offset=wb.sheetnames.index('ACUs')
                                 - wb.sheetnames.index('Insumos'))
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── PDF: Canvas con numeración P.X/Y ────────────────────────────────────────

def _make_page_canvas():
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.units import cm

    class PageNumCanvas(rl_canvas.Canvas):
        def __init__(self, *args, **kwargs):
            rl_canvas.Canvas.__init__(self, *args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            n = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self._draw_pnum(n)
                rl_canvas.Canvas.showPage(self)
            rl_canvas.Canvas.save(self)

        def _draw_pnum(self, total):
            self.saveState()
            self.setFont('Inter', 7)
            self.setFillColorRGB(0.4, 0.4, 0.4)
            self.drawString(1.5 * cm, 0.6 * cm, f"P.{self._pageNumber}/{total}")
            self.restoreState()

    return PageNumCanvas


# ─── PDF: Reporte Completo ────────────────────────────────────────────────────

def exportar_pdf(proyecto_id):
    """Genera PDF completo del expediente técnico, estilo PowerCost limpio."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, PageBreak,
                                    HRFlowable, KeepTogether)
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    pdfmetrics.registerFont(TTFont('Inter',            _FONT_R))
    pdfmetrics.registerFont(TTFont('Inter-Bold',       _FONT_B))
    pdfmetrics.registerFont(TTFont('Inter-Italic',     _FONT_RI))
    pdfmetrics.registerFont(TTFont('Inter-BoldItalic', _FONT_BI))
    pdfmetrics.registerFontFamily('Inter',
        normal='Inter', bold='Inter-Bold',
        italic='Inter-Italic', boldItalic='Inter-BoldItalic'
    )

    conn     = get_db()
    proyecto = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    items, totales = calcular_totales(proyecto_id)
    rubros_pie, total_final = _calcular_rubros_pie(conn, proyecto_id, totales['cd'])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.8*cm,  bottomMargin=1.8*cm,
        title=f"Expediente Técnico - {proyecto['nombre']}"
    )

    C_BLACK = colors.black
    C_GREY  = colors.HexColor('#888888')

    st_titulo = ParagraphStyle('titulo',
        fontName='Inter-BoldItalic', fontSize=16,
        alignment=TA_CENTER, spaceAfter=6, spaceBefore=0)
    st_lbl = ParagraphStyle('lbl',
        fontName='Inter-Italic', fontSize=9,
        leading=12, spaceAfter=0)
    st_lbl_bold = ParagraphStyle('lbl_bold',
        fontName='Inter-BoldItalic', fontSize=9,
        leading=12, spaceAfter=0)
    st_val = ParagraphStyle('val',
        fontName='Inter', fontSize=9,
        leading=12, spaceAfter=0)
    st_val_bold = ParagraphStyle('val_bold',
        fontName='Inter-Bold', fontSize=9,
        leading=12, spaceAfter=0)
    st_fecha = ParagraphStyle('fecha',
        fontName='Inter-Bold', fontSize=11,
        alignment=TA_RIGHT, leading=14, spaceAfter=0)
    st_normal = ParagraphStyle('normal',
        fontName='Inter', fontSize=9, leading=12)
    st_bold   = ParagraphStyle('bold',
        fontName='Inter-Bold', fontSize=9, leading=12)
    st_small  = ParagraphStyle('small',
        fontName='Inter', fontSize=9, leading=12)
    st_spec   = ParagraphStyle('spec',
        fontName='Inter', fontSize=11, leading=16.5,
        alignment=TA_JUSTIFY, spaceAfter=4, leftIndent=14)
    st_spec_label = ParagraphStyle('spec_label',
        fontName='Inter', fontSize=11, leading=16.5,
        alignment=TA_LEFT, spaceAfter=2, leftIndent=14)
    st_spec_h = ParagraphStyle('spec_h',
        fontName='Inter-Bold', fontSize=12,
        alignment=TA_LEFT, spaceAfter=4, spaceBefore=14)
    st_letras = ParagraphStyle('letras',
        fontName='Inter-BoldItalic', fontSize=9,
        spaceBefore=4)

    def _fmt(v, dec=2):
        try:    return f"{float(v or 0):,.{dec}f}"
        except: return '0.00'

    def _encabezado(titulo_seccion):
        nombre   = proyecto['nombre']          or ''
        sub_ppto = proyecto['sub_presupuesto'] or proyecto['nombre'] or ''
        cliente  = proyecto['cliente']         or ''
        ubicacion = proyecto['ubicacion']      or ''
        costo_al  = proyecto['costo_al']       or ''

        story = [Paragraph(titulo_seccion, st_titulo)]
        hdr_data = [
            [Paragraph('<i>Proyecto</i>',        st_lbl), Paragraph(nombre,                st_val),      '', ''],
            [Paragraph('<i>Sub Presupuesto</i>', st_lbl), Paragraph(f'<b>{sub_ppto}</b>', st_val_bold), '', ''],
            [Paragraph('<i>Cliente</i>',         st_lbl), Paragraph(cliente,               st_val),      '', ''],
            [Paragraph('<i>Ubicación</i>',       st_lbl), Paragraph(ubicacion,             st_val),
             Paragraph('<i>Costo a :</i>', st_lbl),       Paragraph(f'<b>{costo_al}</b>', st_fecha)],
        ]
        hdr_t = Table(hdr_data, colWidths=[3*cm, 9.5*cm, 2*cm, 3.5*cm])
        hdr_t.setStyle(TableStyle([
            ('FONTSIZE',      (0, 0), (-1, -1), 9),
            ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ('TOPPADDING',    (0, 0), (-1, -1), 1),
            ('LEFTPADDING',   (0, 0), (-1, -1), 0),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 2),
            ('SPAN', (1, 0), (3, 0)),
            ('SPAN', (1, 1), (3, 1)),
            ('SPAN', (1, 2), (3, 2)),
        ]))
        story.append(hdr_t)
        story.append(HRFlowable(width='100%', thickness=2, color=C_BLACK,
                                spaceAfter=4, spaceBefore=2))
        return story

    W = A4[0] - 3*cm  # ~18 cm

    story = []

    # ══════════════════════════════════════════════════════════════════════════
    # 1. PRESUPUESTO
    # ══════════════════════════════════════════════════════════════════════════
    story += _encabezado('Presupuesto')
    story.append(Spacer(1, 0.2*cm))

    col_p = [1.4*cm, 7.0*cm, 1.0*cm, 1.7*cm, 1.8*cm, 2.0*cm, 1.6*cm, 1.5*cm]
    cab_p = [['Ítem', 'Descripción', 'Und', 'Metrado', 'Precio', 'Parcial', 'Subtotal', 'Total']]

    ppto_data = list(cab_p)
    for entry in items:
        p     = entry['partida']
        total = entry['total']
        if p['nivel'] == 1:
            ppto_data.append([p['item'], p['descripcion'], '', '', '', '', '', _fmt(total)])
        elif p['es_titulo']:
            ppto_data.append([p['item'], p['descripcion'], '', '', '', '', _fmt(total), ''])
        else:
            ppto_data.append([
                p['item'], p['descripcion'], p['unidad'] or '',
                _fmt(p['metrado']), _fmt(p['precio_unitario']),
                _fmt(total), '', ''
            ])

    t_p = Table(ppto_data, colWidths=col_p, repeatRows=1)
    ts_p = TableStyle([
        # Cabecera
        ('FONTNAME',      (0, 0), (-1, 0), 'Inter-BoldItalic'),
        ('LINEBELOW',     (0, 0), (-1, 0), 1, C_BLACK),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('GRID',          (0, 0), (-1, -1), 0.3, C_GREY),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',         (3, 0), (-1, -1), 'RIGHT'),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING',   (0, 0), (1, -1), 3),
    ])
    for i, entry in enumerate(items, start=1):
        p = entry['partida']
        if p['nivel'] == 1:
            ts_p.add('FONTNAME',   (0, i), (-1, i), 'Inter-Bold')
            ts_p.add('LINEABOVE',  (0, i), (-1, i), 1,   C_BLACK)
            ts_p.add('LINEBELOW',  (0, i), (-1, i), 0.5, C_BLACK)
        elif p['es_titulo']:
            ts_p.add('FONTNAME',   (0, i), (-1, i), 'Inter-BoldItalic')

    t_p.setStyle(ts_p)
    story.append(t_p)

    story.append(Spacer(1, 0.1*cm))
    col_t2 = [sum(col_p[:-1]), col_p[-1]]

    if rubros_pie is not None:
        tot_rows = [['COSTO DIRECTO', _fmt(totales['cd'])]]
        for rub in rubros_pie:
            tot_rows.append([rub['nombre'].upper(), _fmt(rub['valor'])])
    else:
        tot_rows = [
            ['COSTO DIRECTO',                             _fmt(totales['cd'])],
            [f"GASTOS GENERALES ({proyecto['gf_pct']}%)", _fmt(totales['gf'])],
            [f"UTILIDAD ({proyecto['utilidad_pct']}%)",   _fmt(totales['utilidad'])],
            ['SUB TOTAL',                                 _fmt(totales['subtotal'])],
            [f"IGV ({proyecto['igv_pct']}%)",             _fmt(totales['igv'])],
            ['PRESUPUESTO TOTAL',                         _fmt(totales['total'])],
        ]

    t_tot = Table(tot_rows, colWidths=col_t2)
    ts_tot = TableStyle([
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('FONTNAME',      (0, 0), (0, -1), 'Inter-Bold'),
        ('ALIGN',         (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN',         (1, 0), (1, -1), 'RIGHT'),
        ('GRID',          (0, 0), (-1, -1), 0.3, C_GREY),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LINEABOVE',     (0, 0), (-1, 0), 1, C_BLACK),
        ('LINEBELOW',     (0, -1), (-1, -1), 1, C_BLACK),
    ])
    for i, row in enumerate(tot_rows):
        label = row[0]
        if 'TOTAL' in label or 'SUB' in label:
            ts_tot.add('FONTNAME',  (0, i), (-1, i), 'Inter-Bold')
            ts_tot.add('LINEABOVE', (0, i), (-1, i), 0.5, C_BLACK)
    t_tot.setStyle(ts_tot)
    story.append(t_tot)

    monto_final = totales['total'] if rubros_pie is None else total_final
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(f"<b>Son :</b>   {_monto_letras(monto_final)}", st_letras))

    # ══════════════════════════════════════════════════════════════════════════
    # 2. ANÁLISIS DE COSTOS UNITARIOS
    # ══════════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story += _encabezado('Análisis de Costos Unitarios')
    story.append(Spacer(1, 0.2*cm))

    partidas_hoja = conn.execute(
        "SELECT * FROM partidas WHERE proyecto_id=? AND es_titulo=0 ORDER BY item",
        (proyecto_id,)
    ).fetchall()

    col_acu = [2.0*cm, 6.5*cm, 1.0*cm, 1.8*cm, 1.8*cm, 2.0*cm, 2.9*cm]

    for partida in partidas_hoja:
        items_acu = conn.execute(
            """SELECT ai.*, r.codigo, r.descripcion as rdesc, r.tipo, r.unidad,
                      COALESCE(ai.precio, r.precio, 0) as precio
               FROM acu_items ai JOIN recursos r ON r.id=ai.recurso_id
               WHERE ai.partida_id=?""", (partida['id'],)
        ).fetchall()
        if not items_acu:
            continue

        acu_data = []
        rend_txt = f"Rend:  {(partida['rendimiento'] or 0):.4f}  {partida['unidad'] or ''}/DÍA"
        acu_data.append([
            f"Partida   {partida['item']}",
            partida['descripcion'], '', '', '', rend_txt, ''
        ])
        acu_data.append(['Código', 'Descripción Insumo', 'Und', 'Cuadrilla', 'Cantidad', 'Precio', 'Parcial'])

        row_tipos  = []
        row_subtot = []

        for tipo, label in [('MO', 'Mano de Obra'), ('MAT', 'Materiales'),
                            ('EQ', 'Equipo'), ('SC', 'Sub-contratos')]:
            tipo_items = [x for x in items_acu if x['tipo'] == tipo]
            if not tipo_items:
                continue
            row_tipos.append(len(acu_data))
            acu_data.append(['', label, '', '', '', '', ''])
            subtipo = 0
            for it in tipo_items:
                precio  = it['precio']   or 0
                cant    = it['cantidad'] or 0
                parcial = precio * cant
                subtipo += parcial
                acu_data.append([
                    it['codigo'], it['rdesc'], it['unidad'] or '',
                    _fmt(it['cuadrilla'], 4), _fmt(cant, 4),
                    _fmt(precio), _fmt(parcial)
                ])
            row_subtot.append(len(acu_data))
            acu_data.append(['', '', '', '', '', '', _fmt(subtipo)])

        row_cu = len(acu_data)
        cu_txt = f"Costo Unitario por {partida['unidad'] or 'UND'} :"
        acu_data.append(['', '', '', '', '', cu_txt, _fmt(partida['precio_unitario'])])

        t_acu = Table(acu_data, colWidths=col_acu)
        ts_acu = TableStyle([
            ('FONTSIZE',      (0, 0), (-1, -1), 9),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ('ALIGN',         (3, 1), (-1, -1), 'RIGHT'),
            ('LEFTPADDING',   (0, 0), (1, -1), 3),
            # Fila 0: encabezado partida
            ('FONTNAME',      (0, 0), (-1, 0), 'Inter-Bold'),
            ('LINEABOVE',     (0, 0), (-1, 0), 1,   C_BLACK),
            ('LINEBELOW',     (0, 0), (-1, 0), 0.5, C_BLACK),
            ('SPAN',          (0, 0), (4, 0)),
            ('ALIGN',         (5, 0), (5, 0), 'RIGHT'),
            # Fila 1: cabecera columnas
            ('FONTNAME',      (0, 1), (-1, 1), 'Inter-BoldItalic'),
            ('LINEBELOW',     (0, 1), (-1, 1), 1, C_BLACK),
            ('GRID',          (0, 1), (-1, -2), 0.3, C_GREY),
            # Fila CU
            ('FONTNAME',      (0, row_cu), (-1, row_cu), 'Inter-Bold'),
            ('FONTSIZE',      (6, row_cu), (6, row_cu), 10),
            ('LINEABOVE',     (0, row_cu), (-1, row_cu), 1, C_BLACK),
            ('LINEBELOW',     (0, row_cu), (-1, row_cu), 1, C_BLACK),
            ('GRID',          (0, row_cu), (-1, row_cu), 0.3, C_GREY),
            ('BOX',           (0, 0), (-1, -1), 0.5, C_BLACK),
        ])
        for idx in row_tipos:
            ts_acu.add('FONTNAME',  (1, idx), (1, idx), 'Inter-Bold')
            ts_acu.add('LINEABOVE', (0, idx), (-1, idx), 0.5, C_BLACK)
            ts_acu.add('SPAN',      (1, idx), (-1, idx))
        for idx in row_subtot:
            ts_acu.add('FONTNAME',  (6, idx), (6, idx), 'Inter-Bold')
            ts_acu.add('LINEABOVE', (0, idx), (-1, idx), 0.5, C_GREY)

        t_acu.setStyle(ts_acu)
        story.append(KeepTogether([t_acu, Spacer(1, 0.4*cm)]))

    # ══════════════════════════════════════════════════════════════════════════
    # 3. LISTADO TOTAL DE INSUMOS
    # ══════════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story += _encabezado('Listado Total de Insumos')
    story.append(Spacer(1, 0.2*cm))

    insumos = conn.execute(
        """SELECT r.codigo, r.descripcion, r.tipo, r.unidad,
                  AVG(COALESCE(ai.precio, r.precio, 0)) as precio,
                  SUM(ai.cantidad * p.metrado) as cantidad_total,
                  SUM(ai.cantidad * p.metrado * COALESCE(ai.precio, r.precio, 0)) as parcial
           FROM acu_items ai JOIN recursos r ON r.id=ai.recurso_id
           JOIN partidas p ON p.id=ai.partida_id
           WHERE p.proyecto_id=? AND p.es_titulo=0 AND SUBSTR(r.unidad,1,1)!='%'
           GROUP BY r.id
           ORDER BY r.tipo,
                    CASE WHEN r.tipo='MO' THEN mo_rank(r.descripcion) ELSE 0 END,
                    r.descripcion""",
        (proyecto_id,)
    ).fetchall()

    col_ins = [2.3*cm, 8.2*cm, 1.2*cm, 2.0*cm, 2.0*cm, 2.3*cm]
    ins_data = [['IU  Código', 'Descripción', 'Und', 'Cantidad', 'Precio', 'Parcial']]
    TIPO_LABEL = {'MAT': 'MATERIALES', 'EQ': 'EQUIPO', 'MO': 'MANO DE OBRA',
                  'SC': 'SUB-CONTRATOS'}
    current_tipo = None
    tipo_rows    = {}
    subtotales   = {}
    subtot_rows  = {}

    for ins in insumos:
        tipo = ins['tipo']
        if tipo != current_tipo:
            current_tipo = tipo
            tipo_rows[len(ins_data)] = tipo
            ins_data.append(['', TIPO_LABEL.get(tipo, tipo), '', '', '', ''])
            subtotales[tipo] = 0
        ins_data.append([
            ins['codigo'], ins['descripcion'], ins['unidad'] or '',
            _fmt(ins['cantidad_total']), _fmt(ins['precio']),
            _fmt(ins['parcial'] or 0)
        ])
        subtotales[tipo] = subtotales.get(tipo, 0) + (ins['parcial'] or 0)

    for tipo in ['MAT', 'EQ', 'MO', 'SC']:
        if tipo in subtotales:
            subtot_rows[len(ins_data)] = tipo
            ins_data.append(['', f"Sub Total {TIPO_LABEL[tipo]}", '', '', '', _fmt(subtotales[tipo])])

    t_ins = Table(ins_data, colWidths=col_ins, repeatRows=1)
    ts_ins = TableStyle([
        ('FONTNAME',      (0, 0), (-1, 0), 'Inter-BoldItalic'),
        ('LINEBELOW',     (0, 0), (-1, 0), 1, C_BLACK),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('GRID',          (0, 0), (-1, -1), 0.3, C_GREY),
        ('ALIGN',         (3, 0), (-1, -1), 'RIGHT'),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING',   (0, 0), (1, -1), 3),
    ])
    for idx in tipo_rows:
        ts_ins.add('FONTNAME',  (0, idx), (-1, idx), 'Inter-Bold')
        ts_ins.add('LINEABOVE', (0, idx), (-1, idx), 0.5, C_BLACK)
        ts_ins.add('SPAN',      (1, idx), (-1, idx))
    for idx in subtot_rows:
        ts_ins.add('FONTNAME',  (0, idx), (-1, idx), 'Inter-Bold')
        ts_ins.add('LINEABOVE', (0, idx), (-1, idx), 0.5, C_BLACK)
    t_ins.setStyle(ts_ins)
    story.append(t_ins)

    story.append(Spacer(1, 0.1*cm))
    if rubros_pie is not None:
        tot_ins = [['COSTO DIRECTO', _fmt(totales['cd'])]]
        for rub in rubros_pie:
            tot_ins.append([rub['nombre'].upper(), _fmt(rub['valor'])])
    else:
        tot_ins = [
            ['COSTO DIRECTO', _fmt(totales['cd'])],
            [f"GG ({proyecto['gf_pct']}%)", _fmt(totales['gf'])],
            ['TOTAL', _fmt(totales['total'])],
        ]

    col_ti = [sum(col_ins[:-1]), col_ins[-1]]
    t_ti = Table(tot_ins, colWidths=col_ti)
    ts_ti = TableStyle([
        ('FONTNAME',      (0, 0), (0, -1), 'Inter-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('ALIGN',         (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN',         (1, 0), (1, -1), 'RIGHT'),
        ('GRID',          (0, 0), (-1, -1), 0.3, C_GREY),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LINEABOVE',     (0, 0), (-1, 0), 1, C_BLACK),
        ('LINEBELOW',     (0, -1), (-1, -1), 1, C_BLACK),
    ])
    for i, row in enumerate(tot_ins):
        if 'TOTAL' in row[0] or 'SUB' in row[0]:
            ts_ti.add('FONTNAME',  (0, i), (-1, i), 'Inter-Bold')
            ts_ti.add('LINEABOVE', (0, i), (-1, i), 0.5, C_BLACK)
    t_ti.setStyle(ts_ti)
    story.append(t_ti)

    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(f"<b>Son :</b>   {_monto_letras(monto_final)}", st_letras))

    # ══════════════════════════════════════════════════════════════════════════
    # 4. PLANILLA DE SUSTENTO DE METRADOS
    # ══════════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story += _encabezado('Planilla de Sustento de Metrados')
    story.append(Spacer(1, 0.2*cm))

    all_partidas = conn.execute(
        "SELECT * FROM partidas WHERE proyecto_id=? ORDER BY item", (proyecto_id,)
    ).fetchall()

    col_met_det = [7.0*cm, 1.5*cm, 1.5*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2.6*cm]

    for p in all_partidas:
        if p['nivel'] == 1:
            blk = [Paragraph(
                f"<b>{p['item']}  &nbsp;&nbsp;{p['descripcion']}</b>",
                ParagraphStyle('t1', fontName='Inter-Bold', fontSize=9,
                               leftIndent=4, spaceAfter=2, spaceBefore=6,
                               borderPad=2, borderWidth=0,
                               borderColor=colors.black)
            )]
            story.append(KeepTogether(blk))
        elif p['es_titulo']:
            story.append(Paragraph(
                f"<b><i>{p['item']}  &nbsp;&nbsp;{p['descripcion']}</i></b>",
                ParagraphStyle('t2', fontName='Inter-BoldItalic', fontSize=9,
                               leftIndent=4, spaceAfter=2)
            ))
        else:
            detalles = conn.execute(
                "SELECT * FROM metrados_detalle WHERE partida_id=? ORDER BY orden",
                (p['id'],)
            ).fetchall()

            total_txt = f"{(p['metrado'] or 0):.2f}  {p['unidad'] or ''}"
            met_data = []
            met_data.append([
                f"Partida  {p['item']}  —  {p['descripcion']}",
                '', '', '', 'Total :', '', total_txt
            ])
            met_data.append(['Descripción', 'N Estr.', 'N Elem', 'Largo', 'Ancho', 'Alto', 'Parcial'])

            if detalles:
                for det in detalles:
                    met_data.append([
                        det['descripcion'] or '',
                        _fmt(det['n_estructuras']) if det['n_estructuras'] else '',
                        _fmt(det['n_elementos'])   if det['n_elementos']   else '',
                        _fmt(det['largo'])         if det['largo']         else '',
                        _fmt(det['ancho'])         if det['ancho']         else '',
                        _fmt(det['alto'])          if det['alto']          else '',
                        _fmt(det['parcial'] or 0)
                    ])
            else:
                met_data.append([p['descripcion'] or '—', '', '', '', '', '', _fmt(p['metrado'] or 0)])

            t_met = Table(met_data, colWidths=col_met_det)
            ts_met = TableStyle([
                ('FONTSIZE',      (0, 0), (-1, -1), 9),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING',    (0, 0), (-1, -1), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                ('ALIGN',         (1, 0), (-1, -1), 'RIGHT'),
                ('LEFTPADDING',   (0, 0), (0, -1),  3),
                # Fila 0: encabezado partida
                ('FONTNAME',      (0, 0), (-1, 0), 'Inter-Bold'),
                ('LINEABOVE',     (0, 0), (-1, 0), 1,   C_BLACK),
                ('LINEBELOW',     (0, 0), (-1, 0), 0.5, C_BLACK),
                ('SPAN',          (0, 0), (3, 0)),
                # Fila 1: cabecera columnas
                ('FONTNAME',      (0, 1), (-1, 1), 'Inter-BoldItalic'),
                ('LINEBELOW',     (0, 1), (-1, 1), 1, C_BLACK),
                ('GRID',          (0, 1), (-1, -1), 0.3, C_GREY),
                ('BOX',           (0, 0), (-1, -1), 0.5, C_BLACK),
            ])
            t_met.setStyle(ts_met)
            story.append(KeepTogether([t_met, Spacer(1, 0.3*cm)]))

    # ══════════════════════════════════════════════════════════════════════════
    # 5. DESAGREGADO DE GASTOS GENERALES
    # ══════════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story += _encabezado('Desagregado de Gastos Generales')
    story.append(Spacer(1, 0.2*cm))

    cd_row = [['COSTO DIRECTO', _fmt(totales['cd'])]]
    col_gg2 = [sum(col_ins[:-1]), col_ins[-1]]
    t_cd = Table(cd_row, colWidths=col_gg2)
    t_cd.setStyle(TableStyle([
        ('FONTNAME',      (0, 0), (-1, -1), 'Inter-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('GRID',          (0, 0), (-1, -1), 0.3, C_GREY),
        ('LINEABOVE',     (0, 0), (-1, 0), 1, C_BLACK),
        ('LINEBELOW',     (0, 0), (-1, 0), 1, C_BLACK),
        ('ALIGN',         (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN',         (0, 0), (0, -1), 'RIGHT'),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(t_cd)
    story.append(Spacer(1, 0.2*cm))

    # Columnas: Descripción · Unidad · % Participación · Cantidad · Precio · Total
    col_gg = [6.0*cm, 1.4*cm, 2.6*cm, 1.8*cm, 2.0*cm, 2.6*cm]

    if rubros_pie is not None:
        for rub in rubros_pie:
            if rub['tipo'] in ('subtotal', 'pct_cd', 'pct_sub'):
                is_bold = rub['tipo'] == 'subtotal' or 'TOTAL' in rub['nombre'].upper()
                resumen = [[rub['nombre'].upper(), _fmt(rub['valor'])]]
                t_r = Table(resumen, colWidths=col_gg2)
                t_r.setStyle(TableStyle([
                    ('FONTNAME',      (0, 0), (-1, -1), 'Inter-Bold' if is_bold else 'Inter'),
                    ('FONTSIZE',      (0, 0), (-1, -1), 9),
                    ('GRID',          (0, 0), (-1, -1), 0.3, C_GREY),
                    ('LINEABOVE',     (0, 0), (-1, 0), 0.5, C_BLACK),
                    ('ALIGN',         (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN',         (1, 0), (1, -1), 'RIGHT'),
                    ('TOPPADDING',    (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]))
                story.append(t_r)
                story.append(Spacer(1, 0.1*cm))
                continue

            gg_items = conn.execute(
                "SELECT * FROM gastos_generales WHERE proyecto_id=? AND rubro=? ORDER BY orden",
                (proyecto_id, rub['codigo'])
            ).fetchall()

            gg_data = []
            gg_data.append([rub['nombre'].upper(), '', '', '', '', ''])
            if gg_items:
                gg_data.append(['Descripción', 'Unidad', '% Participación',
                                'Cantidad', 'Precio', 'Total'])
                grupo_rows = []
                for item in gg_items:
                    if item['tipo'] not in ('grupo', 'item'):
                        continue   # filas 'manual' no se listan en el detalle
                    if item['tipo'] == 'grupo':
                        grupo_rows.append(len(gg_data))
                        gg_data.append([item['descripcion'], '', '', '', '', ''])
                    else:
                        pctp = item['pct_participacion'] if item['pct_participacion'] is not None else 100
                        cant = item['cantidad'] or 0
                        pr_  = item['precio']   or 0
                        total = cant * (pctp / 100) * pr_
                        gg_data.append([
                            item['descripcion'] or '',
                            item['unidad'] or '',
                            _fmt(pctp),
                            _fmt(cant) if cant else '',
                            _fmt(pr_),
                            _fmt(total)
                        ])
            else:
                gg_data.append(['(sin detalle — valor calculado)', '', '', '', '', _fmt(rub['valor'])])
                grupo_rows = []

            subtot_gg_idx = len(gg_data)
            gg_data.append([rub['nombre'].upper(), '', '', '', '', _fmt(rub['valor'])])

            t_gg = Table(gg_data, colWidths=col_gg)
            ts_gg = TableStyle([
                ('FONTSIZE',      (0, 0), (-1, -1), 9),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING',    (0, 0), (-1, -1), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                ('ALIGN',         (2, 1), (-1, -1), 'RIGHT'),
                ('ALIGN',         (1, 1), (1, -1),  'CENTER'),
                ('LEFTPADDING',   (0, 0), (0, -1),  3),
                # Fila 0: encabezado rubro
                ('FONTNAME',      (0, 0), (-1, 0), 'Inter-Bold'),
                ('LINEABOVE',     (0, 0), (-1, 0), 1,   C_BLACK),
                ('LINEBELOW',     (0, 0), (-1, 0), 0.5, C_BLACK),
                ('SPAN',          (0, 0), (-1, 0)),
                # Fila 1: cabecera columnas
                ('FONTNAME',      (0, 1), (-1, 1), 'Inter-BoldItalic'),
                ('LINEBELOW',     (0, 1), (-1, 1), 1, C_BLACK),
                ('GRID',          (0, 1), (-1, -2), 0.3, C_GREY),
                # Última fila: subtotal
                ('FONTNAME',      (0, subtot_gg_idx), (-1, subtot_gg_idx), 'Inter-Bold'),
                ('LINEABOVE',     (0, subtot_gg_idx), (-1, subtot_gg_idx), 0.5, C_BLACK),
                ('LINEBELOW',     (0, subtot_gg_idx), (-1, subtot_gg_idx), 1,   C_BLACK),
                ('SPAN',          (0, subtot_gg_idx), (-2, subtot_gg_idx)),
                ('ALIGN',         (0, subtot_gg_idx), (-2, subtot_gg_idx), 'RIGHT'),
                ('GRID',          (0, subtot_gg_idx), (-1, subtot_gg_idx), 0.3, C_GREY),
                ('BOX',           (0, 0), (-1, -1), 0.5, C_BLACK),
            ])
            for gidx in (grupo_rows if gg_items else []):
                ts_gg.add('FONTNAME',  (0, gidx), (0, gidx), 'Inter-Bold')
                ts_gg.add('LINEABOVE', (0, gidx), (-1, gidx), 0.3, C_GREY)
                ts_gg.add('SPAN',      (0, gidx), (-1, gidx))
            t_gg.setStyle(ts_gg)
            story.append(t_gg)
            story.append(Spacer(1, 0.3*cm))
    else:
        gg_simple = [
            [f"GASTOS GENERALES ({proyecto['gf_pct']}% del CD)", '', '', '', '', _fmt(totales['gf'])],
            [f"UTILIDAD ({proyecto['utilidad_pct']}% del CD)",   '', '', '', '', _fmt(totales['utilidad'])],
        ]
        t_ggs = Table(gg_simple, colWidths=col_gg)
        t_ggs.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID',     (0, 0), (-1, -1), 0.3, C_GREY),
            ('SPAN',     (0, 0), (-2, 0)),
            ('SPAN',     (0, 1), (-2, 1)),
            ('ALIGN',    (-1, 0), (-1, -1),  'RIGHT'),
        ]))
        story.append(t_ggs)

    story.append(Spacer(1, 0.3*cm))
    if rubros_pie is not None:
        res_gg = [['TOTAL COSTO DIRECTO', _fmt(totales['cd'])]]
        for rub in rubros_pie:
            if rub['tipo'] == 'subtotal':
                res_gg.append([rub['nombre'].upper(), _fmt(rub['valor'])])
        res_gg.append(['MONTO TOTAL PRESUPUESTO', _fmt(monto_final)])
    else:
        res_gg = [
            ['TOTAL COSTO DIRECTO',       _fmt(totales['cd'])],
            ['TOTAL COSTOS INDIRECTOS',   _fmt(totales['total'] - totales['cd'])],
            ['MONTO TOTAL PRESUPUESTO',   _fmt(totales['total'])],
        ]
    t_res = Table(res_gg, colWidths=col_gg2)
    ts_res = TableStyle([
        ('FONTNAME',      (0, 0), (-1, -1), 'Inter-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('GRID',          (0, 0), (-1, -1), 0.3, C_GREY),
        ('LINEABOVE',     (0, 0), (-1, 0),  1, C_BLACK),
        ('LINEBELOW',     (0, -1), (-1, -1), 1, C_BLACK),
        ('ALIGN',         (0, 0), (0, -1),  'RIGHT'),
        ('ALIGN',         (1, 0), (1, -1),  'RIGHT'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])
    t_res.setStyle(ts_res)
    story.append(t_res)

    # ══════════════════════════════════════════════════════════════════════════
    # 6. ESPECIFICACIONES TÉCNICAS (si existen)
    # ══════════════════════════════════════════════════════════════════════════
    partidas_spec = conn.execute(
        """SELECT * FROM partidas WHERE proyecto_id=? AND es_titulo=0
           AND especificaciones IS NOT NULL AND especificaciones != ''
           ORDER BY item""", (proyecto_id,)
    ).fetchall()

    def _es_enc_spec(texto):
        t = texto.strip()
        if not t:
            return True
        if t.endswith(':'):
            return True
        if t == t.upper() and len(t) < 70:
            return True
        if t.startswith(('-', '•', '*', '–')):
            return True
        return False

    if partidas_spec:
        story.append(PageBreak())
        story += _encabezado('Especificaciones Técnicas')
        story.append(Spacer(1, 0.2*cm))
        for partida in partidas_spec:
            unidad_str = f" ({partida['unidad']})" if partida['unidad'] else ''
            titulo_spec = f"{partida['item']}  {partida['descripcion']}{unidad_str}"
            story.append(Paragraph(f"<b>{titulo_spec}</b>", st_spec_h))
            spec_text = partida['especificaciones'] or ''
            for linea in spec_text.split('\n'):
                linea = linea.strip()
                if linea:
                    estilo = st_spec_label if _es_enc_spec(linea) else st_spec
                    story.append(Paragraph(linea, estilo))
            story.append(Spacer(1, 0.3*cm))

    conn.close()

    PageNumCanvas = _make_page_canvas()
    doc.build(story, canvasmaker=PageNumCanvas)
    buf.seek(0)
    return buf


# ─── EXPORTAR PROYECTO COMO .DB (SQLite individual) ──────────────────────────

def exportar_proyecto_db(proyecto_id: int, dest_path: str) -> str:
    """Exporta un proyecto completo a un archivo SQLite ``.db`` independiente.

    Replica el esquema de la BD origen y copia únicamente las filas del
    proyecto y sus dependientes (partidas, acu_items, recursos referenciados,
    metrados_detalle, acero_detalle, cronograma_partidas, spec_imagenes,
    gastos_generales, pie_rubros, sub_presupuestos, formula_monomios,
    formula_periodos, pie_presupuesto). Simétrico con lo que reimporta
    ``ingepresupuestos_db_importer.importar_proyecto_db_directo``.
    Retorna ``dest_path``.

    Espejo de la ruta ``/proyecto/<pid>/exportar/proyecto-db`` del Flask.
    """
    import sqlite3
    conn_src = get_db()
    try:
        proy = conn_src.execute(
            "SELECT * FROM proyectos WHERE id=?", (proyecto_id,)
        ).fetchone()
        if not proy:
            raise ValueError(f"Proyecto {proyecto_id} no existe")

        # Eliminar archivo destino si existe (sqlite3.connect crea o abre)
        try:
            os.remove(dest_path)
        except FileNotFoundError:
            pass

        conn_dst = sqlite3.connect(dest_path)
        conn_dst.row_factory = sqlite3.Row
        try:
            # Replicar esquema completo desde origen
            for row in conn_src.execute(
                "SELECT sql FROM sqlite_master "
                "WHERE type='table' AND sql IS NOT NULL"
            ).fetchall():
                try:
                    conn_dst.execute(row['sql'])
                except Exception:
                    pass
            conn_dst.commit()

            def _insert(table_row):
                ph = ','.join(['?'] * len(table_row))
                return tuple(table_row), ph

            tup, ph = _insert(proy)
            conn_dst.execute(f"INSERT INTO proyectos VALUES ({ph})", tup)

            partidas = conn_src.execute(
                "SELECT * FROM partidas WHERE proyecto_id=?", (proyecto_id,)
            ).fetchall()
            part_ids = []
            for row in partidas:
                tup, ph = _insert(row)
                conn_dst.execute(f"INSERT INTO partidas VALUES ({ph})", tup)
                part_ids.append(row['id'])

            if part_ids:
                marcs = ','.join(['?'] * len(part_ids))

                acu_items = conn_src.execute(
                    f"SELECT * FROM acu_items WHERE partida_id IN ({marcs})",
                    part_ids
                ).fetchall()
                recurso_ids = list({r['recurso_id'] for r in acu_items})
                for row in acu_items:
                    tup, ph = _insert(row)
                    conn_dst.execute(f"INSERT INTO acu_items VALUES ({ph})", tup)

                if recurso_ids:
                    rmarcs = ','.join(['?'] * len(recurso_ids))
                    for row in conn_src.execute(
                        f"SELECT * FROM recursos WHERE id IN ({rmarcs})",
                        recurso_ids
                    ).fetchall():
                        tup, ph = _insert(row)
                        conn_dst.execute(
                            f"INSERT OR IGNORE INTO recursos VALUES ({ph})", tup
                        )

                # Tablas opcionales (saltar silenciosamente si no existen)
                for tabla in ("metrados_detalle", "acero_detalle",
                              "cronograma_partidas", "spec_imagenes"):
                    try:
                        for row in conn_src.execute(
                            f"SELECT * FROM {tabla} WHERE partida_id IN ({marcs})",
                            part_ids
                        ).fetchall():
                            tup, ph = _insert(row)
                            conn_dst.execute(
                                f"INSERT INTO {tabla} VALUES ({ph})", tup
                            )
                    except sqlite3.OperationalError:
                        pass

            for tabla in ("gastos_generales", "pie_rubros", "sub_presupuestos",
                          "formula_monomios", "formula_periodos",
                          "pie_presupuesto"):
                try:
                    for row in conn_src.execute(
                        f"SELECT * FROM {tabla} WHERE proyecto_id=?",
                        (proyecto_id,)
                    ).fetchall():
                        tup, ph = _insert(row)
                        conn_dst.execute(f"INSERT INTO {tabla} VALUES ({ph})", tup)
                except sqlite3.OperationalError:
                    pass

            conn_dst.commit()
        finally:
            conn_dst.close()
    finally:
        conn_src.close()
    return dest_path
