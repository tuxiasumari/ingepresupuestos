# SPDX-License-Identifier: GPL-3.0-or-later
# Copyright (C) 2026 Marco Sumari / Sumari SAC
# This file is part of IngePresupuestos — https://ingepresupuestos.com
# Licensed under the GNU GPL v3.0 or later. See the LICENSE file.
import openpyxl
import os
import re

from core.database import get_db, _siguiente_codigo_inei
from utils.formatting import pad_codigo as _pad_codigo


def _norm_desc(s: str) -> str:
    """Normaliza descripción para comparación tolerante."""
    return (s or '').strip().upper()


def _resolve_recurso(conn, codigo: str, descripcion: str, tipo: str,
                     unidad: str, precio: float, scope_map: dict = None) -> int:
    """Devuelve recurso_id reutilizando el catálogo (estilo PowerCost): un
    mismo insumo `(tipo, descripción, unidad)` se mantiene como UN solo
    recurso compartido por todos los proyectos. El precio NO se comparte —
    vive por línea en `acu_items.precio` (COALESCE(ai.precio, r.precio)), así
    que cada proyecto conserva su propio precio aunque comparta el recurso.

    Política (en orden):
    1. `scope_map` (caché de esta importación) → reúsa de inmediato.
    2. Existe un recurso con el mismo `(tipo, descripción, unidad)` en el
       catálogo → reúsalo (aunque el código difiera). Esto evita que cada
       import sume otra copia del insumo (PEON con otro código, etc.).
    3. Mismo código + misma descripción → reúsalo (compat).
    4. Código nuevo → crear con ese código.
    5. Código ya usado por OTRA descripción (colisión) → crear con código
       alternativo (siguiente libre en el mismo prefijo INEI).

    `scope_map` (dict opcional): caché por importación para no repetir el
    SELECT por cada línea del mismo insumo.
    """
    desc_n = _norm_desc(descripcion)
    und_n = (unidad or '').strip()
    skey = (tipo, desc_n, und_n)
    # 1. Caché de la importación en curso.
    if scope_map is not None and desc_n and skey in scope_map:
        return scope_map[skey]

    def _remember(rid):
        if scope_map is not None and desc_n:
            scope_map[skey] = rid
        return rid

    # 2. Reúso por insumo (tipo+descripción+unidad) en TODO el catálogo.
    if desc_n:
        match = conn.execute(
            "SELECT id FROM recursos WHERE tipo=? "
            "AND UPPER(TRIM(descripcion))=? AND TRIM(unidad)=? "
            "ORDER BY id LIMIT 1",
            (tipo, desc_n, und_n)
        ).fetchone()
        if match:
            return _remember(match['id'])

    ex = conn.execute(
        "SELECT id, descripcion FROM recursos WHERE codigo=?", (codigo,)
    ).fetchone()
    # 3. Mismo código + misma descripción.
    if ex and _norm_desc(ex['descripcion']) == desc_n:
        return _remember(ex['id'])
    # 4. Código libre → crear con el código pedido.
    if not ex:
        indice = codigo[:2] if len(codigo) >= 2 else ''
        c = conn.execute(
            "INSERT INTO recursos (codigo, descripcion, tipo, unidad, "
            "precio, indice_inei) VALUES (?,?,?,?,?,?)",
            (codigo, descripcion, tipo, unidad, precio, indice)
        )
        return _remember(c.lastrowid)
    # 5. Colisión: código ocupado por otra descripción → código alternativo.
    indice = codigo[:2] if len(codigo) >= 2 else ''
    nuevo_codigo = _siguiente_codigo_inei(conn, indice)
    c = conn.execute(
        "INSERT INTO recursos (codigo, descripcion, tipo, unidad, "
        "precio, indice_inei) VALUES (?,?,?,?,?,?)",
        (nuevo_codigo, descripcion, tipo, unidad, precio, indice)
    )
    return _remember(c.lastrowid)

# ─── UTILS ────────────────────────────────────────────────────────────────────

def safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(',', '.').strip())
    except:
        return 0.0

def safe_str(val):
    if val is None:
        return ''
    return str(val).strip()

def _find_header_cols(row, patterns):
    """Busca en una fila las columnas que coinciden con los patrones dados.
    patterns: dict {nombre: [keyword, ...]}
    Retorna dict {nombre: col_index}
    """
    result = {}
    for j, v in enumerate(row or []):
        vu = safe_str(v).upper()
        for name, keywords in patterns.items():
            if name not in result:
                if any(k in vu for k in keywords):
                    result[name] = j
    return result


# ─── DELPHIN EXPRESS ──────────────────────────────────────────────────────────
# Formato detectado en archivos de muestra:
#   Presupuesto.xlsx: metadatos filas 2-6 (col 0 = etiqueta, col 4 = valor)
#     Cabecera fila 7: col0=Item, col1=Desc, col8=Unid., col9=Cant., col10=Precio
#   Acu.xlsx: "Partida:" en col0, item en col1, desc en col3, rend en col12
#     Tipo en col0 (MANO DE OBRA/MATERIALES/EQUIPO), recurso: col0=código, col2=desc,
#     col9=unidad, col10=cuadrilla, col13=cantidad
#   Insumos.xlsx: tipo en col1 (cuando col0 es None), recurso: col0=código,
#     col1=desc, col7=unidad, col9=precio

def import_delphin_presupuesto(filepath):
    """Importa Presupuesto desde Excel de Delphin Express."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    info = {'nombre': '', 'cliente': '', 'ubicacion': '', 'sub_presupuesto': '', 'costo_al': ''}
    partidas = []
    header_row = None
    col_item = col_desc = col_und = col_met = col_pu = None

    for i, row in enumerate(rows):
        if not row:
            continue
        c0 = safe_str(row[0]).upper()
        # Delphin pone valor en col 4 (precedida por ":")
        c_val = safe_str(row[4]) if len(row) > 4 else ''
        c1 = safe_str(row[1]) if len(row) > 1 else ''

        if not info['nombre'] and ('PROYECTO' in c0 or 'OBRA' in c0):
            info['nombre'] = c_val or c1
        elif not info['cliente'] and ('PROPIETARIO' in c0 or 'CLIENTE' in c0):
            info['cliente'] = c_val or c1
        elif not info['ubicacion'] and ('UBICACI' in c0 or 'LUGAR' in c0):
            info['ubicacion'] = c_val or c1
        elif not info['costo_al'] and ('FECHA' in c0 or 'COSTO A' in c0):
            info['costo_al'] = c_val or c1

        row_text = ' '.join(safe_str(v).upper() for v in row if v)
        if ('ITEM' in row_text or 'ÍTEM' in row_text) and 'DESCRIPCI' in row_text:
            header_row = i
            cols = _find_header_cols(row, {
                'item':  ['ITEM', 'ÍTEM'],
                'desc':  ['DESCRIPCI'],
                'und':   ['UNID', 'UNIDAD', 'UND'],
                'met':   ['CANT', 'METRADO'],
                'pu':    ['PRECIO', 'C.U', 'P.U'],
            })
            col_item = cols.get('item', 0)
            col_desc = cols.get('desc', 1)
            col_und  = cols.get('und',  8)
            col_met  = cols.get('met',  9)
            col_pu   = cols.get('pu',   10)
            break

    if header_row is None:
        return info, partidas

    for row in rows[header_row + 1:]:
        if not row or col_item >= len(row) or row[col_item] is None:
            continue
        item = safe_str(row[col_item])
        if not item or not any(c.isdigit() for c in item):
            continue

        desc   = safe_str(row[col_desc]) if col_desc < len(row) else ''
        unidad = safe_str(row[col_und])  if col_und  < len(row) else ''
        metrado = safe_float(row[col_met]) if col_met < len(row) else 0
        precio  = safe_float(row[col_pu])  if col_pu  < len(row) else 0
        nivel   = len(item.split('.'))
        es_titulo = 1 if (not unidad and metrado == 0 and precio == 0) else 0

        partidas.append({
            'item': item, 'descripcion': desc, 'unidad': unidad,
            'metrado': metrado, 'precio_unitario': precio,
            'nivel': nivel, 'es_titulo': es_titulo,
        })

    return info, partidas


def import_delphin_acus(filepath):
    """Importa ACUs desde Excel de Delphin Express."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    acus = {}
    current_partida = None
    current_tipo = None
    TIPO_MAP = {
        'MANO DE OBRA': 'MO', 'M.O.': 'MO',
        'MATERIALES': 'MAT', 'MATERIAL': 'MAT',
        'EQUIPO': 'EQ', 'EQUIPOS': 'EQ', 'HERRAMIENTA': 'EQ',
    }

    for row in rows:
        if not row or all(v is None for v in row):
            continue
        c0 = safe_str(row[0])
        c0u = c0.upper().rstrip(':')

        # Cabecera de partida: col0 = "Partida:"
        if c0u == 'PARTIDA' or c0.upper().startswith('PARTIDA:'):
            item = safe_str(row[1]) if len(row) > 1 else ''
            rend = 1.0
            # Buscar texto de rendimiento en las últimas columnas ("Rendimiento:50 m²/Día")
            for idx in range(len(row) - 1, -1, -1):
                v = safe_str(row[idx]) if row[idx] else ''
                if v and ('RENDIMIENTO' in v.upper() or 'REND' in v.upper()):
                    m = re.search(r'(\d+(?:[.,]\d+)?)', v)
                    if m:
                        rend = safe_float(m.group(1))
                    break
            if item:
                current_partida = item.strip()
                acus[current_partida] = {'rendimiento': rend, 'items': []}
                current_tipo = None
            continue

        # Marcador de tipo de recurso (en col0)
        matched_tipo = None
        for k, v in TIPO_MAP.items():
            if c0u == k:
                matched_tipo = v
                break
        if matched_tipo:
            current_tipo = matched_tipo
            continue

        # Línea de recurso
        # Delphin: col0=código, col2=desc, col9=unidad, col10=cuadrilla, col13=cantidad
        if current_partida and current_tipo and c0 and len(row) > 9:
            # Skip filas de encabezado de columnas
            if c0u in ('CÓDIGO', 'CODIGO', 'ITEM', 'ÍTEM'):
                continue
            desc = safe_str(row[2]) if len(row) > 2 else ''
            if not desc or not any(c.isalpha() for c in desc):
                continue
            if desc.upper() in ('DESCRIPCIÓN', 'DESCRIPCION', 'RECURSO'):
                continue
            unidad = safe_str(row[9]) if len(row) > 9 else ''
            cuad_val = safe_str(row[10]) if len(row) > 10 else ''
            cuad = safe_float(cuad_val) if cuad_val not in ('-', '', 'Recursos', 'RECURSOS') else 0
            cant = safe_float(row[13]) if len(row) > 13 else 0
            precio = 0  # Delphin no incluye precio en ACU; se toma de insumos

            acus[current_partida]['items'].append({
                'codigo': c0, 'descripcion': desc, 'tipo': current_tipo,
                'unidad': unidad, 'cuadrilla': cuad, 'cantidad': cant, 'precio': precio,
            })

    return acus


def import_delphin_insumos(filepath):
    """Importa Insumos desde Excel de Delphin Express."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    recursos = []
    current_tipo = 'MAT'
    TIPO_MAP = {
        'MANO DE OBRA': 'MO', 'MATERIALES': 'MAT',
        'EQUIPO': 'EQ', 'EQUIPOS': 'EQ', 'HERRAMIENTA': 'EQ',
    }
    header_found = False

    for row in rows:
        if not row or all(v is None for v in row):
            continue
        c0 = safe_str(row[0])
        c1 = safe_str(row[1]) if len(row) > 1 else ''

        if not header_found:
            row_text = ' '.join(safe_str(v).upper() for v in row if v)
            if 'CÓDIGO' in row_text or 'CODIGO' in row_text:
                header_found = True
            continue

        # Tipo marker: col0 es None, col1 tiene el tipo
        if not c0 and c1:
            c1u = c1.upper()
            for k, v in TIPO_MAP.items():
                if k in c1u:
                    current_tipo = v
                    break
            continue

        # Recurso: col0=código, col1=desc, col7=unidad, col9=precio (Costo)
        if c0 and c1 and any(c.isdigit() for c in c0):
            if c1.upper() in TIPO_MAP or c1.upper() in ('DESCRIPCIÓN', 'DESCRIPCION'):
                continue
            unidad = safe_str(row[7]) if len(row) > 7 else ''
            # Recursos porcentuales (%mo, %mat) no tienen precio unitario real
            if '%' in unidad.lower() or unidad.lower() in ('%mo', '%mat', '%eq', '%pres'):
                precio = 0.0
            else:
                precio = safe_float(row[9]) if len(row) > 9 else 0
            recursos.append({
                'codigo': c0, 'descripcion': c1, 'tipo': current_tipo,
                'unidad': unidad, 'precio': precio,
            })

    return recursos


# ─── POWERCOST ────────────────────────────────────────────────────────────────
# Formato detectado en archivos de muestra:
#   Presupuesto de obra.xlsx: hoja "SP1"
#     Metadatos: Proyecto(rows 3-5 col2), Sub Presupuesto(row6 col2), Cliente(row7 col2)
#               Ubicación(row8 col2), Costo a (row8 col6-8)
#     Cabecera fila 11: col0=Item, col1=Desc, col3=Unidad, col4=Metrado, col5=Precio
#   Análisis de costos unitarios.xlsx: hoja "SP1"
#     Partida: col0="Partida", col1=item, col3=desc, col9=rendimiento
#     Tipo en col2 (Equipo/Mano de Obra/Materiales)
#     Recurso: col1=código, col2=desc, col4=unidad, col5=cuadrilla, col6=cant, col7=precio
#   Relación de insumos.xlsx:
#     Cabecera fila 10: IU Código(col0), Desc(col2), Unidad(col4), Cant(col5), Precio(col6)
#     Tipo en col1 (MATERIALES, MANO DE OBRA, etc.)

def import_powercost_presupuesto(filepath):
    """Importa Presupuesto desde Excel de PowerCost."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    # Buscar hoja correcta
    ws = None
    for name in wb.sheetnames:
        nu = name.upper()
        if 'PRESUPUESTO' in nu or 'SP' in nu or 'PPTO' in nu:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    info = {'nombre': '', 'cliente': '', 'ubicacion': '', 'sub_presupuesto': '', 'costo_al': ''}
    partidas = []
    header_row = None
    col_item = col_desc = col_und = col_met = col_pu = None
    col_subtot = col_total = None

    # PowerCost puede tener nombre de proyecto en múltiples filas (rows 3-5 col 2)
    nombre_parts = []

    for i, row in enumerate(rows):
        if not row:
            continue
        c0 = safe_str(row[0])
        c0u = c0.upper()
        c2 = safe_str(row[2]) if len(row) > 2 else ''

        # Detectar metadatos PowerCost
        if 'PROYECTO' in c0u or 'OBRA' in c0u:
            if c2:
                nombre_parts = [c2]
        elif not c0 and nombre_parts and c2 and not info['nombre']:
            # Continuación de nombre de proyecto en filas siguientes
            nombre_parts.append(c2)
            continue
        elif 'SUB PRESUPUESTO' in c0u or 'SUBPRESUPUESTO' in c0u:
            if nombre_parts:
                info['nombre'] = ' '.join(nombre_parts).strip('"').strip()
            info['sub_presupuesto'] = c2
        elif 'CLIENTE' in c0u or 'PROPIETARIO' in c0u:
            info['cliente'] = c2
        elif 'UBICACI' in c0u or 'LUGAR' in c0u:
            info['ubicacion'] = c2
            # "Costo a :" puede estar en col 6-8 de la misma fila
            for j in range(len(row) - 1, -1, -1):
                if row[j] is not None and safe_str(row[j]).strip():
                    cv = safe_str(row[j]).strip()
                    if cv and cv not in (c2, c0):
                        info['costo_al'] = cv
                    break

        # Detectar cabecera de tabla
        row_text = ' '.join(safe_str(v).upper() for v in row if v)
        if ('ITEM' in row_text or 'ÍTEM' in row_text) and 'DESCRIPCI' in row_text and 'UNIDAD' in row_text:
            if nombre_parts and not info['nombre']:
                info['nombre'] = ' '.join(nombre_parts).strip('"').strip()
            header_row = i
            cols = _find_header_cols(row, {
                'item': ['ITEM', 'ÍTEM'],
                'desc': ['DESCRIPCI'],
                'und':  ['UNIDAD', 'UNID', 'UND'],
                'met':  ['METRADO', 'CANTIDAD', 'CANT'],
                'pu':   ['PRECIO', 'P.U', 'C.U'],
            })
            col_item = cols.get('item', 0)
            col_desc = cols.get('desc', 1)
            col_und  = cols.get('und',  3)
            col_met  = cols.get('met',  4)
            col_pu   = cols.get('pu',   5)
            # Columnas «Subtotal»/«Total» — marcan la profundidad de los títulos.
            for j, v in enumerate(row):
                vu = safe_str(v).upper().strip()
                if col_subtot is None and ('SUBTOTAL' in vu or 'SUB TOTAL' in vu):
                    col_subtot = j
                elif col_total is None and vu == 'TOTAL':
                    col_total = j
            break

    if header_row is None:
        # Intentar formato simple (presupuesto sin encabezado explícito)
        if nombre_parts and not info['nombre']:
            info['nombre'] = ' '.join(nombre_parts).strip('"').strip()
        return info, partidas

    # ── Numeración jerárquica limpia ──────────────────────────────────────────
    # PowerCost usa códigos NO jerárquicos (101.A, 500. A, 501.A1) que el árbol
    # del proyecto no sabe anidar (anida por prefijo de ítem). Reasignamos códigos
    # limpios (01, 01.01, 01.01.01…) según la estructura de TÍTULOS, detectada por:
    #   · fila sin ítem + sin metrado/precio + sin Total/Subtotal → Componente (n1)
    #   · ítem-título con valor en col «Total»     → grupo (n2)
    #   · ítem-título con valor en col «Subtotal»  → subgrupo (n3)
    #   · ítem con metrado/precio/unidad           → partida hoja (cuelga del título)
    # El código PowerCost original se guarda en `item_origen` para casar el ACU/metrados.
    nums = [0, 0, 0, 0, 0, 0, 0, 0]

    def _code_for(depth):
        depth = max(1, min(depth, len(nums)))
        nums[depth - 1] += 1
        for k in range(depth, len(nums)):
            nums[k] = 0
        return '.'.join(f"{nums[i]:02d}" for i in range(depth))

    last_title_depth = 1
    for row in rows[header_row + 1:]:
        if not row or all(v is None for v in row):
            continue
        item    = safe_str(row[col_item]) if col_item < len(row) else ''
        desc    = safe_str(row[col_desc]) if col_desc < len(row) else ''
        unidad  = safe_str(row[col_und])  if col_und  < len(row) else ''
        metrado = safe_float(row[col_met]) if col_met < len(row) else 0
        precio  = safe_float(row[col_pu])  if col_pu  < len(row) else 0
        subtot  = row[col_subtot] if (col_subtot is not None and col_subtot < len(row)) else None
        total   = row[col_total]  if (col_total  is not None and col_total  < len(row)) else None
        has_sub = subtot not in (None, '') and safe_float(subtot) != 0
        has_tot = total  not in (None, '') and safe_float(total)  != 0

        es_hoja = bool(unidad) or metrado != 0 or precio != 0
        if not es_hoja:
            # Fila de título (o fila de resumen/total a descartar)
            if not desc or len(desc) < 2:
                continue
            if not item:
                if has_tot or has_sub:
                    continue            # fila de total/resumen sin ítem → no es título
                depth = 1               # Componente
            elif not any(c.isdigit() for c in item):
                continue                # etiqueta de pie ("Son : <monto en letras>",
                                        # firmas…): no es partida ni título
            elif has_sub:
                depth = 3               # Subgrupo (subtotal)
            else:
                depth = 2               # Grupo (total) o título sin marca
            last_title_depth = depth
            partidas.append({
                'item': _code_for(depth), 'item_origen': item,
                'descripcion': desc, 'unidad': '',
                'metrado': 0, 'precio_unitario': 0,
                'nivel': depth, 'es_titulo': 1,
            })
        else:
            if not item and not desc:
                continue
            depth = last_title_depth + 1
            partidas.append({
                'item': _code_for(depth), 'item_origen': item,
                'descripcion': desc, 'unidad': unidad,
                'metrado': metrado, 'precio_unitario': precio,
                'nivel': depth, 'es_titulo': 0,
            })

    return info, partidas


def import_powercost_acus(filepath):
    """Importa ACUs desde Excel de PowerCost."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = None
    for name in wb.sheetnames:
        nu = name.upper()
        if 'ACU' in nu or 'ANALISIS' in nu or 'ANÁLISIS' in nu or 'COSTOS' in nu or 'SP' in nu:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    acus = {}
    current_partida = None
    current_tipo = None
    TIPO_MAP = {
        'MANO DE OBRA': 'MO', 'M.O.': 'MO',
        'MATERIALES': 'MAT', 'MATERIAL': 'MAT',
        'EQUIPO': 'EQ', 'EQUIPOS': 'EQ', 'HERRAMIENTA': 'EQ',
        'SUBCONTRATO': 'SC', 'SUBCONTRATOS': 'SC',
        'SUB PARTIDA': 'SC', 'SUBPARTIDA': 'SC',
    }

    for row in rows:
        if not row or all(v is None for v in row):
            continue
        c0 = safe_str(row[0])
        c0u = c0.upper()
        c1 = safe_str(row[1]) if len(row) > 1 else ''

        # Cabecera de partida REAL: col0="Partida", col1=item, col3=desc, col9=rend.
        # OJO: PowerCost lista, tras cada partida, sus «Sub Partida» y «Detalle Ins.»
        # (sub-análisis) que NO deben fusionarse en la partida. Esas cabeceras van en
        # col1 (col0 vacío) y se ignoran; el corte real lo da la fila «Costo Unitario».
        if c0u == 'PARTIDA':
            item = safe_str(row[1]) if len(row) > 1 else ''
            rend = 1.0
            # Rendimiento en col 9: "1.0000  GLB/DIA"
            for idx in [9, 10, 8]:
                if len(row) > idx and row[idx] is not None:
                    v = safe_str(row[idx])
                    m = re.search(r'(\d+(?:[.,]\d+)?)', v)
                    if m:
                        rend = safe_float(m.group(1))
                        break
            if item:
                current_partida = item.strip()
                acus[current_partida] = {'rendimiento': rend, 'items': []}
                current_tipo = None
            continue

        # Fin de la composición directa de la partida → dejar de capturar. Lo que
        # sigue (sub-partidas / detalles de insumo) son análisis aparte.
        if current_partida and any('COSTO UNITARIO' in safe_str(v).upper() for v in row):
            current_partida = None
            current_tipo = None
            continue

        # Tipo de recurso (cabecera de sección): col1 vacío + col2 = nombre de sección.
        # El guard `not c1` evita confundir un INSUMO cuya descripción contenga
        # "EQUIPO"/"MATERIAL" (p.ej. "EQUIPO DE SOLDAR") con una cabecera de sección.
        c2 = safe_str(row[2]) if len(row) > 2 else ''
        c2u = c2.upper()
        if not c1:
            matched = None
            for k, v in TIPO_MAP.items():
                if k in c2u:
                    matched = v
                    break
            if matched:
                current_tipo = matched
                continue

        # Recurso: col1=código, col2=desc, col4=unidad, col5=cuadrilla, col6=cant, col7=precio
        if current_partida and current_tipo and c1 and c2 and len(row) >= 7:
            if c2u in ('DESCRIPCIÓN INSUMO', 'DESCRIPCION INSUMO', 'DESCRIPCIÓN', 'DESCRIPCION'):
                continue
            unidad = safe_str(row[4]) if len(row) > 4 else ''
            cuad_val = safe_str(row[5]) if len(row) > 5 else ''
            cuad = safe_float(cuad_val) if cuad_val.strip() not in ('-', '') else 0
            cant  = safe_float(row[6]) if len(row) > 6 else 0
            precio = safe_float(row[7]) if len(row) > 7 else 0
            if c2 and any(c.isalpha() for c in c2):
                acus[current_partida]['items'].append({
                    'codigo': c1, 'descripcion': c2, 'tipo': current_tipo,
                    'unidad': unidad, 'cuadrilla': cuad, 'cantidad': cant, 'precio': precio,
                })

    return acus


def import_powercost_insumos(filepath):
    """Importa Insumos desde Excel de PowerCost (Relación de Insumos)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    recursos = []
    current_tipo = 'MAT'
    TIPO_MAP = {
        'MANO DE OBRA': 'MO', 'MATERIALES': 'MAT',
        'EQUIPO': 'EQ', 'EQUIPOS': 'EQ', 'HERRAMIENTA': 'EQ',
        'SUBCONTRATO': 'EQ',
    }
    header_found = False

    for row in rows:
        if not row or all(v is None for v in row):
            continue
        c0 = safe_str(row[0])
        c1 = safe_str(row[1]) if len(row) > 1 else ''

        if not header_found:
            row_text = ' '.join(safe_str(v).upper() for v in row if v)
            if 'CÓDIGO' in row_text or 'CODIGO' in row_text or 'IU' in row_text:
                header_found = True
            continue

        # Tipo en col1 cuando col0 es None
        if not c0 and c1:
            c1u = c1.upper()
            for k, v in TIPO_MAP.items():
                if k in c1u:
                    current_tipo = v
                    break
            continue

        # Recurso: col0=código(con espacios), col2=desc, col4=unidad, col5=cant, col6=precio
        if c0:
            codigo = c0.strip()
            desc   = safe_str(row[2]) if len(row) > 2 else ''
            if not desc or desc.upper() in ('DESCRIPCIÓN', 'DESCRIPCION'):
                continue
            unidad = safe_str(row[4]) if len(row) > 4 else ''
            precio = safe_float(row[6]) if len(row) > 6 else 0
            recursos.append({
                'codigo': codigo, 'descripcion': desc, 'tipo': current_tipo,
                'unidad': unidad, 'precio': precio,
            })

    return recursos


# ─── S10 COSTOS Y PRESUPUESTOS ────────────────────────────────────────────────
# Formato detectado en archivos .xls (xlrd necesario):
#   Presupuesto.xls: hoja Sheet1
#     Col1="Presupuesto", col3=código, col4=nombre (fila 7)
#     Col1="Subpresupuesto", col3=código, col4=sub_presupuesto (fila 11)
#     Col1="Cliente", col3=cliente, col10="Costo al", col14=fecha (fila 15)
#     Col1="Lugar", col3=ubicacion (fila 17)
#     Cabecera fila 19: col1=Item, col3=Desc, col7=Und., col8=Metrado, col10=Precio
#   Insumos.xls: hoja Hoja1
#     Cabecera fila 4: col2=Código, col5=Recurso, col6=Unidad, col7=cantidad, col8=precio
#     Tipo en col2 (MANO DE OBRA, MATERIALES, EQUIPOS)

def import_s10_presupuesto(filepath):
    """Importa Presupuesto desde Excel .xls exportado por S10."""
    try:
        import xlrd
        return _import_s10_presupuesto_xlrd(filepath)
    except ImportError:
        pass
    # Fallback: intentar con openpyxl (si es xlsx)
    return _import_s10_presupuesto_openpyxl(filepath)


def _import_s10_presupuesto_xlrd(filepath):
    import xlrd
    wb = xlrd.open_workbook(filepath)
    # Buscar la hoja correcta
    ws = None
    for name in wb.sheet_names():
        nu = name.upper()
        if 'PRESUPUESTO' in nu or 'SHEET' in nu or 'PPTO' in nu:
            ws = wb.sheet_by_name(name)
            break
    if ws is None:
        ws = wb.sheet_by_index(0)

    def cell(row, col):
        try:
            v = ws.cell_value(row, col)
            return str(v).strip() if v != '' else ''
        except:
            return ''

    def num(row, col):
        try:
            return float(ws.cell_value(row, col))
        except:
            return 0.0

    info = {'nombre': '', 'cliente': '', 'ubicacion': '', 'sub_presupuesto': '', 'costo_al': ''}
    partidas = []
    header_row = None
    col_item = col_desc = col_und = col_met = col_pu = None

    for i in range(min(30, ws.nrows)):
        c1 = cell(i, 1).upper() if ws.ncols > 1 else ''
        c3 = cell(i, 3) if ws.ncols > 3 else ''
        c4 = cell(i, 4) if ws.ncols > 4 else ''

        if not info['nombre'] and 'PRESUPUESTO' in c1 and c4:
            info['nombre'] = c4
        if not info['sub_presupuesto'] and 'SUBPRESUPUESTO' in c1 and c4:
            info['sub_presupuesto'] = c4
        if not info['cliente'] and 'CLIENTE' in c1 and c3:
            info['cliente'] = c3
            if ws.ncols > 14:
                info['costo_al'] = cell(i, 14)
        if not info['ubicacion'] and ('LUGAR' in c1 or 'UBICACI' in c1) and c3:
            info['ubicacion'] = c3

        # Detectar cabecera: col1=Item, col3=Descripción
        row_cells = [cell(i, j).upper() for j in range(min(16, ws.ncols))]
        row_text = ' '.join(row_cells)
        if ('ITEM' in row_text or 'ÍTEM' in row_text) and 'DESCRIPCI' in row_text:
            header_row = i
            for j, vu in enumerate(row_cells):
                if vu in ('ITEM', 'ÍTEM', 'CÓDIGO', 'CODIGO'):
                    col_item = j
                elif 'DESCRIPCI' in vu:
                    col_desc = j
                elif vu in ('UND', 'UND.', 'UNIDAD', 'UNID.'):
                    col_und = j
                elif 'METRADO' in vu or 'CANTIDAD' in vu:
                    col_met = j
                elif 'PRECIO' in vu or 'P.U' in vu:
                    col_pu = j
            # Defaults S10
            if col_item is None: col_item = 1
            if col_desc is None: col_desc = 3
            if col_und  is None: col_und  = 7
            if col_met  is None: col_met  = 8
            if col_pu   is None: col_pu   = 10
            break

    if header_row is None:
        return info, partidas

    for i in range(header_row + 1, ws.nrows):
        item = cell(i, col_item)
        if not item or not any(c.isdigit() for c in item):
            continue
        desc   = cell(i, col_desc)
        unidad = cell(i, col_und)
        metrado = num(i, col_met)
        precio  = num(i, col_pu)
        nivel   = len(item.split('.'))
        es_titulo = 1 if (not unidad and metrado == 0 and precio == 0) else 0

        partidas.append({
            'item': item, 'descripcion': desc.strip(), 'unidad': unidad,
            'metrado': metrado, 'precio_unitario': precio,
            'nivel': nivel, 'es_titulo': es_titulo,
        })

    return info, partidas


def _import_s10_presupuesto_openpyxl(filepath):
    """Fallback openpyxl para S10 xlsx."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    info = {'nombre': '', 'cliente': '', 'ubicacion': '', 'sub_presupuesto': '', 'costo_al': ''}
    partidas = []
    header_row = None
    col_item = 1; col_desc = 3; col_und = 7; col_met = 8; col_pu = 10

    for i, row in enumerate(rows[:30]):
        if not row:
            continue
        c1 = safe_str(row[1]).upper() if len(row) > 1 else ''
        c3 = safe_str(row[3]) if len(row) > 3 else ''
        c4 = safe_str(row[4]) if len(row) > 4 else ''

        if not info['nombre'] and 'PRESUPUESTO' in c1 and c4:
            info['nombre'] = c4
        if not info['sub_presupuesto'] and 'SUBPRESUPUESTO' in c1 and c4:
            info['sub_presupuesto'] = c4
        if not info['cliente'] and 'CLIENTE' in c1 and c3:
            info['cliente'] = c3
            info['costo_al'] = safe_str(row[14]) if len(row) > 14 else ''
        if not info['ubicacion'] and ('LUGAR' in c1 or 'UBICACI' in c1) and c3:
            info['ubicacion'] = c3

        row_text = ' '.join(safe_str(v).upper() for v in row if v)
        if ('ITEM' in row_text or 'ÍTEM' in row_text) and 'DESCRIPCI' in row_text:
            header_row = i
            for j, v in enumerate(row):
                vu = safe_str(v).upper()
                if vu in ('ITEM', 'ÍTEM', 'CÓDIGO', 'CODIGO'): col_item = j
                elif 'DESCRIPCI' in vu: col_desc = j
                elif vu in ('UND', 'UND.', 'UNIDAD'): col_und = j
                elif 'METRADO' in vu: col_met = j
                elif 'PRECIO' in vu or 'P.U' in vu: col_pu = j
            break

    if header_row is None:
        return info, partidas

    for row in rows[header_row + 1:]:
        if not row or col_item >= len(row) or row[col_item] is None:
            continue
        item = safe_str(row[col_item])
        if not item or not any(c.isdigit() for c in item):
            continue
        desc   = safe_str(row[col_desc]) if col_desc < len(row) else ''
        unidad = safe_str(row[col_und])  if col_und  < len(row) else ''
        metrado = safe_float(row[col_met]) if col_met < len(row) else 0
        precio  = safe_float(row[col_pu])  if col_pu  < len(row) else 0
        nivel   = len(item.split('.'))
        es_titulo = 1 if (not unidad and metrado == 0 and precio == 0) else 0
        partidas.append({
            'item': item, 'descripcion': desc.strip(), 'unidad': unidad,
            'metrado': metrado, 'precio_unitario': precio,
            'nivel': nivel, 'es_titulo': es_titulo,
        })

    return info, partidas


def import_s10_acus(filepath):
    """Importa ACUs desde Excel exportado por S10."""
    # S10 generalmente no exporta ACUs en formato separado, retornar vacío
    # Si en el futuro hay soporte, se implementa aquí
    try:
        import xlrd
        wb = xlrd.open_workbook(filepath)
    except:
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except:
            return {}
    return {}


def import_s10_insumos(filepath):
    """Importa Insumos desde Excel .xls exportado por S10 (hoja Hoja1)."""
    try:
        import xlrd
        return _import_s10_insumos_xlrd(filepath)
    except ImportError:
        return []


def _import_s10_insumos_xlrd(filepath):
    import xlrd
    wb = xlrd.open_workbook(filepath)

    # Buscar hoja con datos (Hoja1 es la más completa)
    ws = None
    for name in ['Hoja1', 'HOJA1', 'hoja1']:
        if name in wb.sheet_names():
            ws = wb.sheet_by_name(name)
            break
    if ws is None:
        ws = wb.sheet_by_index(0)

    def cell(row, col):
        try:
            v = ws.cell_value(row, col)
            return str(v).strip() if v != '' else ''
        except:
            return ''

    def num(row, col):
        try:
            return float(ws.cell_value(row, col))
        except:
            return 0.0

    recursos = []
    current_tipo = 'MAT'
    TIPO_MAP = {
        'MANO DE OBRA': 'MO', 'MATERIALES': 'MAT',
        'EQUIPO': 'EQ', 'EQUIPOS': 'EQ', 'HERRAMIENTA': 'EQ',
        'SUBCONTRATO': 'EQ',
    }
    header_found = False

    for i in range(ws.nrows):
        row_vals = [cell(i, j) for j in range(min(10, ws.ncols))]
        row_text = ' '.join(v.upper() for v in row_vals if v)

        if not header_found:
            if 'CÓDIGO' in row_text or 'CODIGO' in row_text or 'RECURSO' in row_text:
                header_found = True
            continue

        c2 = cell(i, 2)  # Código
        c5 = cell(i, 5)  # Recurso/Descripción

        # Tipo marker
        c2u = c2.upper()
        if not c5 and c2:
            for k, v in TIPO_MAP.items():
                if k in c2u:
                    current_tipo = v
                    break
            continue

        # Recurso: col2=código, col5=desc, col6=unidad, col8=precio
        if c2 and c5 and any(ch.isdigit() for ch in c2):
            if c2u in TIPO_MAP:
                continue
            unidad = cell(i, 6)
            precio = num(i, 8)
            recursos.append({
                'codigo': c2, 'descripcion': c5, 'tipo': current_tipo,
                'unidad': unidad, 'precio': precio,
            })

    return recursos


# ─── GENÉRICO (compatibilidad hacia atrás) ────────────────────────────────────

def import_presupuesto(filepath):
    """Importa Presupuesto.xlsx genérico (formato PowerCost o similar).
    Detecta automáticamente el formato."""
    # Intentar PowerCost primero (el más común)
    try:
        return import_powercost_presupuesto(filepath)
    except Exception:
        pass
    # Fallback genérico
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    info = {'nombre': '', 'cliente': '', 'ubicacion': '', 'sub_presupuesto': '', 'costo_al': ''}
    partidas = []
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    col_item = 0; col_desc = 1; col_und = 3; col_met = 4; col_pu = 5

    for i, row in enumerate(rows[:20]):
        if not row:
            continue
        c0 = safe_str(row[0])
        c0u = c0.upper()
        cv = safe_str(row[2]) if len(row) > 2 else safe_str(row[1]) if len(row) > 1 else ''
        if 'PROYECTO' in c0u or 'OBRA' in c0u: info['nombre'] = cv
        elif 'SUB PRESUPUESTO' in c0u: info['sub_presupuesto'] = cv
        elif 'CLIENTE' in c0u: info['cliente'] = cv
        elif 'UBICACI' in c0u: info['ubicacion'] = cv

        row_text = ' '.join(safe_str(v).upper() for v in row if v)
        if ('ITEM' in row_text or 'ÍTEM' in row_text) and 'DESCRIPCI' in row_text:
            header_row = i
            break

    if header_row is None:
        return info, partidas

    for row in rows[header_row + 1:]:
        if not row or row[col_item] is None:
            continue
        item = safe_str(row[col_item])
        if not item or not any(c.isdigit() for c in item):
            continue
        desc   = safe_str(row[col_desc]) if col_desc < len(row) else ''
        unidad = safe_str(row[col_und])  if col_und  < len(row) else ''
        metrado = safe_float(row[col_met]) if col_met < len(row) else 0
        precio  = safe_float(row[col_pu])  if col_pu  < len(row) else 0
        nivel   = len(item.split('.'))
        es_titulo = 1 if (not unidad and metrado == 0 and precio == 0) else 0
        partidas.append({
            'item': item, 'descripcion': desc, 'unidad': unidad,
            'metrado': metrado, 'precio_unitario': precio,
            'nivel': nivel, 'es_titulo': es_titulo,
        })

    return info, partidas


def import_acus(filepath):
    """Importa ACUs genérico (intenta PowerCost primero)."""
    try:
        return import_powercost_acus(filepath)
    except Exception:
        return {}


def import_insumos(filepath):
    """Importa insumos genérico - detecta formato automáticamente."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        return _import_s10_insumos_xlrd(filepath)

    # Intentar detectar PowerCost o Delphin por contenido
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # PowerCost: header "IU  Código" en col0
        # Delphin: header "Código" en col0, descripción en col1
        for row in rows[:15]:
            if not row:
                continue
            c0 = safe_str(row[0]).upper()
            c1 = safe_str(row[1]).upper() if len(row) > 1 else ''
            if 'IU' in c0 or 'RELACIÓN' in c0:
                return import_powercost_insumos(filepath)
            if 'LISTA DE INSUMOS' in c0 or ('CÓDIGO' in c0 and 'DESCRIPCI' in c1):
                return import_delphin_insumos(filepath)

        # Fallback genérico
        return import_powercost_insumos(filepath)
    except Exception:
        return []


# ─── POWERCOST METRADOS ───────────────────────────────────────────────────────

def import_powercost_metrados(filepath):
    """
    Importa Planilla de Sustento de Metrados desde Excel de PowerCost.
    Retorna dict {item: [{'descripcion','n_estructuras','n_elementos','area',
                          'largo','ancho','alto','parcial'}, ...]}
    """
    def _num(val):
        """Convierte a float o None; espacios y cadenas vacías → None."""
        if val is None:
            return None
        s = str(val).strip()
        if s == '' or s == ' ':
            return None
        try:
            return float(s.replace(',', '.'))
        except Exception:
            return None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    # Buscar hoja correcta
    ws = None
    for name in wb.sheetnames:
        nu = name.upper()
        if 'METRADO' in nu or 'SP' in nu or 'SUSTENTO' in nu:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    result = {}        # {item: [filas]}
    current_item = None

    for row in rows:
        if not row or all(v is None for v in row):
            continue

        c0 = safe_str(row[0])
        c1 = safe_str(row[1]) if len(row) > 1 else ''
        c1u = c1.upper()

        # Fila de cabecera de partida: col[0]=='Partida'
        if c0.upper() == 'PARTIDA':
            item = safe_str(row[1])
            if item:
                current_item = item
                result[current_item] = []
            continue

        # Fila de encabezado de columnas (contiene 'DESCRIPCI' y 'N ESTR')
        if 'DESCRIPCI' in c1u or ('N ESTR' in c1u and 'N ELEM' in safe_str(row[6] if len(row)>6 else '').upper()):
            continue

        # Fila de datos: tiene descripción en col[1] y al menos una dimensión
        if current_item is not None and c1 and c1 != ' ':
            # Saltar si es una fila residual de partida
            if c0.upper() in ('PARTIDA', 'PRESUPUESTO', 'PROYECTO', 'SUB PRESUPUESTO',
                              'CLIENTE', 'UBICACI'):
                continue
            n_estr  = _num(row[5])  if len(row) > 5  else None
            n_elem  = _num(row[6])  if len(row) > 6  else None
            area    = _num(row[7])  if len(row) > 7  else None
            largo   = _num(row[8])  if len(row) > 8  else None
            ancho   = _num(row[9])  if len(row) > 9  else None
            alto    = _num(row[10]) if len(row) > 10 else None
            parcial = _num(row[11]) if len(row) > 11 else None

            # Calcular parcial si falta
            if parcial is None:
                p = 1.0
                has_val = False
                for v in [n_estr, n_elem, area, largo, ancho, alto]:
                    if v is not None:
                        p *= v; has_val = True
                parcial = p if has_val else 0.0

            result[current_item].append({
                'descripcion':   c1,
                'n_estructuras': n_estr,
                'n_elementos':   n_elem,
                'area':          area,
                'largo':         largo,
                'ancho':         ancho,
                'alto':          alto,
                'parcial':       parcial,
            })

    return result


# ─── GUARDAR IMPORTACIÓN ──────────────────────────────────────────────────────

def guardar_importacion(info: dict, partidas_data: list, acus_data: dict = None,
                        recursos_data: list = None, metrados_data: dict = None) -> int:
    """Persiste el resultado de cualquier importador (PowerCost, Delphin, S10,
    IFC, PDF…) y retorna el ``id`` del proyecto creado.

    Espejo de ``_guardar_importacion`` del Flask original. La normalización de
    precios al final asegura que un mismo recurso tenga el mismo precio en
    todo el proyecto.
    """
    acus_data = acus_data or {}
    conn = get_db()
    cur = conn.execute(
        """INSERT INTO proyectos (nombre, cliente, ubicacion, sub_presupuesto, costo_al)
           VALUES (?,?,?,?,?)""",
        (info.get('nombre') or 'Proyecto importado', info.get('cliente', ''),
         info.get('ubicacion', ''), info.get('sub_presupuesto', ''),
         info.get('costo_al', ''))
    )
    pid = cur.lastrowid

    # Aplicar porcentajes del pie si vienen explícitos en el archivo origen.
    # Usamos `is not None` (no `or`) para respetar 0% explícito.
    pcts_explicitos = any(info.get(k) is not None for k in
                          ('gf_pct', 'utilidad_pct', 'igv_pct'))
    for campo in ('gf_pct', 'utilidad_pct', 'igv_pct'):
        if info.get(campo) is not None:
            conn.execute(f"UPDATE proyectos SET {campo}=? WHERE id=?",
                         (float(info[campo]), pid))

    # Si el archivo origen trajo porcentajes, sembrar pie_rubros directamente.
    # El flag `pie_activo_default` (en info) controla si los rubros se crean
    # activos o inactivos:
    #   - PowerCost  → activo (sus reportes incluyen CD+GG+Util+IGV)
    #   - Delphin    → inactivo por default (adicionales suelen ser solo CD)
    # Los porcentajes reales se preservan en cualquier caso.
    if pcts_explicitos:
        gf   = float(info.get('gf_pct')       or 0)
        util = float(info.get('utilidad_pct') or 0)
        igv  = float(info.get('igv_pct')      or 0)
        # Si pie_activo_default=True → rubros con valor > 0 se activan
        activar = bool(info.get('pie_activo_default'))
        _on = lambda pct: (1 if activar and pct > 0 else 0)
        rubros_default = [
            ('GG',   'Gastos Generales',   gf,   _on(gf),   0, 'rubro',    1),
            ('UTIL', 'Utilidad',            util, _on(util), 1, 'pct_cd',   1),
            ('SUB',  'Sub Total',           0,    1 if activar else 0,
                                                   2, 'subtotal', 1),
            ('SUP',  'Supervisión',         5.0,  0, 3, 'rubro',    0),
            ('ET',   'Expediente Técnico',  3.0,  0, 4, 'rubro',    0),
            ('LQ',   'Liquidación de Obra', 2.0,  0, 5, 'rubro',    0),
            ('IGV',  f'IGV ({int(igv) if igv else 18}%)',
                                            igv,  _on(igv), 6, 'pct_sub', 1),
        ]
        for codigo, nombre, pct, activo, orden, tipo, mostrar_pct in rubros_default:
            conn.execute(
                "INSERT INTO pie_rubros"
                " (proyecto_id, codigo, nombre, pct, activo, orden, tipo, mostrar_pct)"
                " VALUES (?,?,?,?,?,?,?,?)",
                (pid, codigo, nombre, pct, activo, orden, tipo, mostrar_pct)
            )
    partida_map = {}
    for p in partidas_data:
        c = conn.execute(
            """INSERT INTO partidas (proyecto_id, item, descripcion, unidad, metrado,
               precio_unitario, nivel, es_titulo) VALUES (?,?,?,?,?,?,?,?)""",
            (pid, p['item'], p['descripcion'], p['unidad'],
             p['metrado'], p['precio_unitario'], p['nivel'], p['es_titulo'])
        )
        partida_map[p['item']] = c.lastrowid

    # Código de origen → ítem limpio. PowerCost reasigna ítems jerárquicos
    # (01, 01.01…) pero el ACU y los metrados vienen con el código original del
    # archivo (101.A, 501.A1…). Este mapa los reconcilia. Otros importadores
    # (Delphin/S10) no traen `item_origen` → el mapa queda vacío y se usa el
    # código tal cual (item == código original).
    raw_to_clean = {p['item_origen']: p['item']
                    for p in partidas_data if p.get('item_origen')}

    def _pid_de(code):
        return partida_map.get(raw_to_clean.get(code, code))

    # Mapa de recursos acotado a ESTA importación: colapsa mismo insumo
    # (tipo+desc+unidad) bajo distintos códigos en un solo recurso → evita
    # duplicados intra-proyecto (PEON con 2 códigos, etc.). Compartido por
    # los dos bucles para que listado de insumos y ACU resuelvan igual.
    scope_recursos: dict = {}

    if recursos_data:
        for rec in recursos_data:
            rec['codigo'] = _pad_codigo(rec['codigo'])
            _resolve_recurso(
                conn, rec['codigo'], rec['descripcion'], rec['tipo'],
                rec['unidad'], rec['precio'], scope_map=scope_recursos
            )

    for item_code, acu in acus_data.items():
        part_id = _pid_de(item_code)
        if not part_id:
            continue
        conn.execute("UPDATE partidas SET rendimiento=? WHERE id=?",
                     (acu['rendimiento'], part_id))
        for it in acu['items']:
            it['codigo'] = _pad_codigo(it['codigo'])
            rec_id = _resolve_recurso(
                conn, it['codigo'], it['descripcion'], it['tipo'],
                it['unidad'], it['precio'], scope_map=scope_recursos
            )
            conn.execute(
                "INSERT INTO acu_items (partida_id, recurso_id, cuadrilla, "
                "cantidad, precio) VALUES (?,?,?,?,?)",
                (part_id, rec_id, it['cuadrilla'], it['cantidad'],
                 it.get('precio', 0) or 0)
            )

    # Rellenar precios faltantes desde otros acu_items del mismo recurso
    # SOLO cuando precio actual = 0 o NULL (evita machacar precios reales
    # diferentes que vienen del archivo origen, p.ej. Delphin permite mismo
    # recurso con precios distintos en partidas diferentes del mismo proyecto).
    conn.execute(
        """UPDATE acu_items SET precio = (
               SELECT MAX(ai2.precio) FROM acu_items ai2
               JOIN partidas p2 ON p2.id = ai2.partida_id
               WHERE ai2.recurso_id = acu_items.recurso_id
                 AND p2.proyecto_id = ?
                 AND ai2.precio > 0
           )
           WHERE partida_id IN (SELECT id FROM partidas WHERE proyecto_id=?)
             AND (precio IS NULL OR precio = 0)
             AND EXISTS (
               SELECT 1 FROM acu_items ai2
               JOIN partidas p2 ON p2.id = ai2.partida_id
               WHERE ai2.recurso_id = acu_items.recurso_id
                 AND p2.proyecto_id = ? AND ai2.precio > 0
             )""",
        (pid, pid, pid)
    )

    if metrados_data:
        for item, filas in metrados_data.items():
            part_id = _pid_de(item)
            if not part_id or not filas:
                continue
            for orden, f in enumerate(filas):
                conn.execute(
                    """INSERT INTO metrados_detalle
                       (partida_id, orden, descripcion, n_estructuras, n_elementos,
                        area, largo, ancho, alto, parcial)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (part_id, orden,
                     f.get('descripcion', ''),
                     f.get('n_estructuras'), f.get('n_elementos'),
                     f.get('area'), f.get('largo'), f.get('ancho'),
                     f.get('alto'), f.get('parcial', 0))
                )
            total = sum(f.get('parcial') or 0 for f in filas)
            if total:
                conn.execute("UPDATE partidas SET metrado=? WHERE id=?",
                             (total, part_id))

    # ── Enriquecer la biblioteca CU con los ACUs únicos del proyecto ──────
    # Por cada partida-hoja con ACU detallado (>=1 item) la guardamos en
    # biblioteca_cu. Dedup por (descripcion, unidad, grupo='') tras strip().
    # Si el CU ya existe pero NO tiene items, RELLENA sus items con esta
    # importación — así nunca quedan CU huérfanos en la biblioteca.
    for item_code, acu in acus_data.items():
        # Si el ACU no tiene items, no aporta nada a la biblioteca
        items_acu = acu.get('items') or []
        if not items_acu:
            continue
        part_id = _pid_de(item_code)
        if not part_id:
            continue
        p_row = conn.execute(
            "SELECT descripcion, unidad, precio_unitario FROM partidas WHERE id=?",
            (part_id,)
        ).fetchone()
        if not p_row:
            continue
        desc = (p_row['descripcion'] or '').strip()
        und  = (p_row['unidad'] or '').strip()
        if not desc:
            continue
        # ¿Ya existe en la biblioteca?
        existente = conn.execute(
            "SELECT id FROM biblioteca_cu "
            "WHERE descripcion=? AND unidad=? AND grupo=''",
            (desc, und)
        ).fetchone()
        if existente:
            cu_id = existente['id']
            # ¿Tiene items? Si no, los rellenamos con esta importación
            n_items = conn.execute(
                "SELECT COUNT(*) FROM biblioteca_acu_items WHERE cu_id=?",
                (cu_id,)
            ).fetchone()[0]
            if n_items > 0:
                continue   # ya está completo, respetar
        else:
            cur = conn.execute(
                "INSERT INTO biblioteca_cu "
                "(descripcion, unidad, rendimiento, costo_unitario, grupo, "
                " especificaciones, usos) VALUES (?,?,?,?,'','',0)",
                (desc, und, float(acu.get('rendimiento') or 1.0),
                 float(p_row['precio_unitario'] or 0))
            )
            cu_id = cur.lastrowid
        # Insertar items del ACU
        for it in items_acu:
            cod = _pad_codigo(it.get('codigo') or '')
            if not cod:
                continue
            rec = conn.execute(
                "SELECT id, precio FROM recursos WHERE codigo=?", (cod,)
            ).fetchone()
            if not rec:
                continue
            precio = it.get('precio')
            if precio in (None, '', 0):
                precio = rec['precio']
            conn.execute(
                "INSERT INTO biblioteca_acu_items "
                "(cu_id, recurso_id, cuadrilla, cantidad, precio) VALUES (?,?,?,?,?)",
                (cu_id, rec['id'],
                 float(it.get('cuadrilla') or 0),
                 float(it.get('cantidad') or 0),
                 float(precio) if precio is not None else None)
            )

    conn.commit()
    conn.close()
    return pid
