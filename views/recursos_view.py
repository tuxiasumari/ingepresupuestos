# SPDX-License-Identifier: GPL-3.0-or-later
# Copyright (C) 2026 Marco Sumari / Sumari SAC
# This file is part of IngePresupuestos — https://ingepresupuestos.com
# Licensed under the GNU GPL v3.0 or later. See the LICENSE file.
"""recursos_view — Catálogo de Insumos (≈ recursos.html de Flask).

Lista todos los recursos del catálogo (MO/MAT/EQ) con filtros, KPIs, edición
inline del precio, CRUD completo (nuevo/editar/duplicar/eliminar) y
import/export Excel. Reusa el pool de Índices INEI del recurso_selector.
"""
from __future__ import annotations

from PySide6.QtCore import Qt, QModelIndex, QRect, QSize, QTimer, Signal
from PySide6.QtGui import (
    QAction, QFont, QKeySequence, QShortcut, QColor, QPalette, QPainter,
)
from PySide6.QtWidgets import (
    QWidget, QFrame, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QComboBox,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QMessageBox, QFileDialog, QMenu, QDialog, QDialogButtonBox, QFormLayout,
    QSizePolicy, QStyledItemDelegate, QStyle, QStyleOptionViewItem, QApplication,
)

from core.database import get_db, _siguiente_codigo_inei
from core.config import TIPOS_RECURSO, INEI_DEFAULT
from utils.formatting import fmt, parse_num
from utils.icons import icon
from utils.theme import C


# ── Paleta — aliases de los tokens centralizados (utils/theme.py) ───────────
BLUE_500   = C.brand            # naranja marca (legacy name)
BLUE_700   = C.brand_hover
SLATE_700  = C.text
SLATE_500  = C.text_secondary
SLATE_300  = C.text_muted
SILVER_100 = C.bg
SILVER_200 = C.surface_subtle
SILVER_300 = C.border
ORANGE_500 = C.brand
GREEN_500  = "#68B723"
RED_500    = C.error

# Tonos para los badges MO/MAT/EQ (canónicos en el tema)
TIPO_FONDO = {'MO': C.mo_bg, 'MAT': C.mat_bg, 'EQ': C.eq_bg, 'SC': C.sc_bg}
TIPO_TEXTO = {'MO': C.mo_fg, 'MAT': C.mat_fg, 'EQ': C.eq_fg, 'SC': C.sc_fg}
TIPO_LARGO = {
    'MO': 'Mano de Obra',
    'MAT': 'Material',
    'EQ': 'Equipo / Herram.',
    'SC': 'Sub-contrato / Servicio',
}


# ── Catálogo de Índices Unificados INEI (80 entradas) ────────────────────────
INEI_CATALOG = [
    ("01", "Aceite"), ("02", "Acero de construcción liso"),
    ("03", "Acero de construcción corrugado"), ("04", "Agregado fino"),
    ("05", "Agregado grueso"), ("06", "Alambre y cable de cobre desnudo"),
    ("07", "Alambre y cable tipo TW y THW"), ("08", "Alambre y cable tipo WP"),
    ("09", "Alcantarilla metálica"), ("10", "Aparato sanitario con grifería"),
    ("11", "Artefacto de alumbrado exterior"), ("12", "Artefacto de alumbrado interior"),
    ("13", "Asfalto"), ("14", "Baldosa acústica"), ("15", "Baldosa asfáltica"),
    ("16", "Baldosa vinílica"), ("17", "Bloque y ladrillo"), ("18", "Cable telefónico"),
    ("19", "Cable NYY-N2XY"), ("20", "Cemento asfáltico"),
    ("21", "Cemento Portland tipo I"), ("22", "Cemento Portland tipo II"),
    ("23", "Cemento Portland tipo V"), ("24", "Cerámica esmaltada y sin esmaltar"),
    ("26", "Cerrajería nacional"), ("27", "Detonante"), ("28", "Dinamita"),
    ("29", "Dólar"), ("30", "Dólar más inflación USA / General ponderado"),
    ("31", "Ducto de concreto"), ("32", "Flete terrestre"), ("33", "Flete aéreo"),
    ("34", "Gasolina"), ("37", "Herramienta manual"), ("38", "Hormigón"),
    ("39", "Índice general de precios al consumidor (IPC)"), ("40", "Loseta"),
    ("41", "Madera en tiras para piso"),
    ("42", "Madera importada para encofrado y carpintería"),
    ("43", "Madera nacional para encofrado y carpintería"),
    ("44", "Madera terciada para encofrado y carpintería"),
    ("45", "Madera terciada para encofrado"), ("46", "Malla de acero"),
    ("47", "Mano de obra (incluido leyes sociales)"),
    ("48", "Maquinaria y equipo nacional"), ("49", "Maquinaria y equipo importado"),
    ("50", "Marco y tapa de hierro fundido"), ("51", "Perfil de acero liviano"),
    ("52", "Perfil de aluminio"), ("53", "Petróleo diesel"),
    ("54", "Pintura látex"), ("55", "Pintura temple"),
    ("56", "Plancha de Aero LAC"), ("57", "Plancha de Aero LAF"),
    ("59", "Plancha de fibro-cemento"), ("60", "Plancha de poliuretano"),
    ("61", "Plancha galvanizada"), ("62", "Poste de concreto"),
    ("64", "Terrazo"), ("65", "Tubería de acero negro y/o galvanizado"),
    ("66", "Tubería de PVC para agua potable y alcantarillado"),
    ("68", "Tubería de cobre"), ("69", "Tubería de concreto simple"),
    ("70", "Tubería de concreto reforzado"), ("71", "Tubería de fierro fundido"),
    ("72", "Tubería de PVC para agua"), ("73", "Ducto telefónico de PVC"),
    ("74", "Tubería de PVC para electricidad (SAP)"),
    ("77", "Válvula de bronce nacional"),
    ("78", "Válvula de fierro fundido nacional"),
    ("79", "Vidrio incoloro nacional"), ("80", "Concreto premezclado"),
]
INEI_NOMBRE = dict(INEI_CATALOG)


# ── Delegate: badge de tipo (píldora coloreada) ──────────────────────────────
class _TipoBadgeDelegate(QStyledItemDelegate):
    """Pinta una píldora coloreada con MO/MAT/EQ — sin fondo de fila alterno."""

    def paint(self, painter, option, index):
        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        opt.text = ""
        QApplication.style().drawControl(QStyle.CE_ItemViewItem, opt, painter, opt.widget)

        tipo = (index.data(Qt.DisplayRole) or "").strip()
        if not tipo:
            return

        bg = QColor(TIPO_FONDO.get(tipo, '#EEEEEE'))
        fg = QColor(TIPO_TEXTO.get(tipo, '#444444'))

        rect = option.rect
        cx = rect.center().x()
        cy = rect.center().y()
        w, h = 38, 18
        pill = QRect(cx - w // 2, cy - h // 2, w, h)

        painter.save()
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setBrush(bg)
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(pill, 9, 9)
        painter.setPen(fg)
        f = painter.font()
        f.setBold(True)
        f.setPointSize(9)
        painter.setFont(f)
        painter.drawText(pill, Qt.AlignCenter, tipo)
        painter.restore()


# ── Diálogo: formulario de recurso (nuevo / editar) ──────────────────────────
class RecursoFormDialog(QDialog):
    """Form para crear o editar un recurso del catálogo."""

    def __init__(self, parent=None, recurso: dict | None = None):
        super().__init__(parent)
        self.recurso = recurso or {}
        self.es_edicion = bool(recurso and recurso.get('id'))
        self.setWindowTitle("Editar insumo" if self.es_edicion else "Nuevo insumo")
        self.setMinimumWidth(480)
        self._build()
        if self.es_edicion:
            self._cargar()
        else:
            self._on_tipo_change()  # preselecciona índice por defecto

    def _build(self):
        v = QVBoxLayout(self)
        v.setContentsMargins(20, 18, 20, 16)
        v.setSpacing(10)

        # Cabecera color marca
        head = QFrame()
        head.setStyleSheet(
            f"QFrame {{ background:{SLATE_500}; border-radius:6px; }}"
            f"QLabel {{ color:white; padding:8px 12px; font-weight:600; font-size:13px; }}"
        )
        hl = QHBoxLayout(head)
        hl.setContentsMargins(0, 0, 0, 0)
        ico_h = QLabel()
        ico_h.setPixmap(icon("editar" if self.es_edicion else "add").pixmap(20, 20))
        hl.addWidget(ico_h)
        hl.addWidget(QLabel("Editar insumo" if self.es_edicion else "Nuevo insumo"))
        v.addWidget(head)

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignRight)
        form.setContentsMargins(0, 4, 0, 4)
        form.setSpacing(8)

        # Tipo
        self.cmb_tipo = QComboBox()
        for t in TIPOS_RECURSO:
            self.cmb_tipo.addItem(f"{t} — {TIPO_LARGO[t]}", t)
        self.cmb_tipo.currentIndexChanged.connect(self._on_tipo_change)
        form.addRow("Tipo *:", self.cmb_tipo)

        # Índice INEI
        self.cmb_inei = QComboBox()
        self.cmb_inei.setEditable(True)
        self.cmb_inei.addItem("— Seleccione índice INEI —", "")
        for cod, desc in INEI_CATALOG:
            self.cmb_inei.addItem(f"{cod} — {desc}", cod)
        self.cmb_inei.addItem("Otro (ingresar manualmente)", "OTRO")
        self.cmb_inei.currentIndexChanged.connect(self._on_inei_change)
        form.addRow("Índice INEI *:", self.cmb_inei)

        self.inp_inei_libre = QLineEdit()
        self.inp_inei_libre.setPlaceholderText("Código de 2 dígitos, ej: 15")
        self.inp_inei_libre.setMaxLength(2)
        self.inp_inei_libre.setVisible(False)
        self.inp_inei_libre.textChanged.connect(self._auto_codigo)
        form.addRow("", self.inp_inei_libre)

        # Código + bloqueo manual
        cod_row = QHBoxLayout()
        cod_row.setContentsMargins(0, 0, 0, 0)
        cod_row.setSpacing(6)
        self.inp_codigo = QLineEdit()
        self.inp_codigo.setMaxLength(7)
        self.inp_codigo.setPlaceholderText("Se genera al elegir el índice")
        self.inp_codigo.setReadOnly(True)
        self._codigo_lock_style(True)
        self.btn_codigo_manual = QPushButton("Editar")
        self.btn_codigo_manual.setFixedWidth(56)
        self.btn_codigo_manual.setToolTip("Editar manualmente")
        self.btn_codigo_manual.clicked.connect(self._toggle_codigo_manual)
        cod_row.addWidget(self.inp_codigo, 1)
        cod_row.addWidget(self.btn_codigo_manual)
        wcod = QWidget()
        wcod.setLayout(cod_row)
        form.addRow("Código:", wcod)
        self._codigo_manual = False

        # Descripción
        self.inp_desc = QLineEdit()
        self.inp_desc.setPlaceholderText("Descripción del insumo")
        form.addRow("Descripción *:", self.inp_desc)

        # Unidad
        self.inp_unidad = QLineEdit()
        self.inp_unidad.setPlaceholderText("hh · m³ · kg · glb …")
        self.inp_unidad.textEdited.connect(self._auto_superindice_unidad)
        form.addRow("Unidad:", self.inp_unidad)

        # Precio
        self.inp_precio = QLineEdit("0.00")
        self.inp_precio.setMaximumWidth(140)
        form.addRow("Precio:", self.inp_precio)

        v.addLayout(form)

        # Mensaje de error
        self.lbl_err = QLabel("")
        self.lbl_err.setStyleSheet(f"color:{RED_500}; font-size:12px; padding:2px 0;")
        v.addWidget(self.lbl_err)

        # Botones
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText(
            "Guardar cambios" if self.es_edicion else "Crear insumo"
        )
        btns.accepted.connect(self._aceptar)
        btns.rejected.connect(self.reject)
        v.addWidget(btns)

    def _codigo_lock_style(self, locked: bool):
        if locked:
            self.inp_codigo.setStyleSheet(
                f"QLineEdit {{ background:{SILVER_100}; color:{SLATE_300}; "
                "border:1px solid #D4D4D4; border-radius:4px; padding:4px 8px; "
                "font-family: monospace; }"
            )
        else:
            self.inp_codigo.setStyleSheet(
                "QLineEdit { background:white; border:1px solid #C0621A; "
                "border-radius:4px; padding:4px 8px; font-family: monospace; }"
            )

    def _toggle_codigo_manual(self):
        self._codigo_manual = not self._codigo_manual
        self.inp_codigo.setReadOnly(not self._codigo_manual)
        self._codigo_lock_style(not self._codigo_manual)
        self.btn_codigo_manual.setText("Auto" if self._codigo_manual else "Editar")
        self.btn_codigo_manual.setToolTip(
            "Volver a auto-generado" if self._codigo_manual else "Editar manualmente"
        )

    def _on_tipo_change(self):
        if self.es_edicion:
            return  # no auto-modificar al editar
        tipo = self.cmb_tipo.currentData()
        inei_def = INEI_DEFAULT.get(tipo, '39')
        for i in range(self.cmb_inei.count()):
            if self.cmb_inei.itemData(i) == inei_def:
                self.cmb_inei.setCurrentIndex(i)
                return

    def _on_inei_change(self):
        es_otro = self.cmb_inei.currentData() == 'OTRO'
        self.inp_inei_libre.setVisible(es_otro)
        if es_otro:
            self.inp_inei_libre.setFocus()
            return
        self._auto_codigo()

    def _inei_actual(self) -> str:
        v = self.cmb_inei.currentData() or ''
        if v == 'OTRO':
            return self.inp_inei_libre.text().strip()
        return v

    def _auto_codigo(self):
        if self._codigo_manual:
            return
        inei = self._inei_actual()
        if len(inei) == 2 and inei.isdigit():
            conn = get_db()
            self.inp_codigo.setText(_siguiente_codigo_inei(conn, inei))
            conn.close()
        else:
            if not self.es_edicion:
                self.inp_codigo.clear()

    def _auto_superindice_unidad(self, txt: str):
        import re
        m = re.match(r'^([a-zA-Z/]+)([23])$', txt)
        if not m:
            return
        sup = '²' if m.group(2) == '2' else '³'
        self.inp_unidad.blockSignals(True)
        self.inp_unidad.setText(m.group(1) + sup)
        self.inp_unidad.blockSignals(False)

    def _cargar(self):
        r = self.recurso
        for i in range(self.cmb_tipo.count()):
            if self.cmb_tipo.itemData(i) == r.get('tipo'):
                self.cmb_tipo.setCurrentIndex(i)
                break
        ind = (r.get('indice_inei') or
               (r.get('codigo', '')[:2] if r.get('codigo') and len(r.get('codigo')) >= 2 else ''))
        encontrado = False
        for i in range(self.cmb_inei.count()):
            if self.cmb_inei.itemData(i) == ind:
                self.cmb_inei.setCurrentIndex(i)
                encontrado = True
                break
        if not encontrado and ind:
            for i in range(self.cmb_inei.count()):
                if self.cmb_inei.itemData(i) == 'OTRO':
                    self.cmb_inei.setCurrentIndex(i)
                    self.inp_inei_libre.setText(ind)
                    break
        self.inp_codigo.setText(r.get('codigo', '') or '')
        self.inp_desc.setText(r.get('descripcion', '') or '')
        self.inp_unidad.setText(r.get('unidad', '') or '')
        precio = r.get('precio', 0) or 0
        self.inp_precio.setText(f"{float(precio):.2f}")

    def _aceptar(self):
        desc = self.inp_desc.text().strip()
        if not desc:
            self.lbl_err.setText("La descripción es obligatoria.")
            self.inp_desc.setFocus()
            return

        inei = self._inei_actual()
        if not (len(inei) == 2 and inei.isdigit()):
            self.lbl_err.setText("El Índice INEI es obligatorio (2 dígitos).")
            self.cmb_inei.setFocus()
            return

        tipo = self.cmb_tipo.currentData()
        unidad = self.inp_unidad.text().strip()
        precio = parse_num(self.inp_precio.text())
        codigo = self.inp_codigo.text().strip()

        conn = get_db()
        if not codigo:
            codigo = _siguiente_codigo_inei(conn, inei)

        # Para overhead (% MO / % MAT / % EQ) el precio debe ser 0
        if unidad.startswith('%'):
            precio = 0.0

        try:
            if self.es_edicion:
                conn.execute(
                    "UPDATE recursos SET codigo=?, descripcion=?, tipo=?, unidad=?, "
                    "precio=?, indice_inei=? WHERE id=?",
                    (codigo, desc, tipo, unidad, precio, inei, self.recurso['id'])
                )
                self.recurso_id = self.recurso['id']
            else:
                cur = conn.execute(
                    "INSERT INTO recursos "
                    "(codigo, descripcion, tipo, unidad, precio, indice_inei) "
                    "VALUES (?,?,?,?,?,?)",
                    (codigo, desc, tipo, unidad, precio, inei)
                )
                self.recurso_id = cur.lastrowid
            conn.commit()
        finally:
            conn.close()
        self.accept()


# ── Vista principal del catálogo ─────────────────────────────────────────────
class RecursosView(QWidget):
    """Catálogo de Insumos — lista, filtros, KPIs y CRUD completo."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setProperty("vista_nombre", "recursos")
        self._row_ids: list[int] = []
        self._build()
        self.cargar()

    # -- construcción UI -------------------------------------------------------
    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 18, 20, 16)
        layout.setSpacing(12)

        # ── Topbar ──
        top = QHBoxLayout()
        top.setSpacing(10)
        ico_title = QLabel()
        ico_title.setPixmap(icon("paquete").pixmap(28, 28))
        ico_title.setFixedSize(28, 28)
        top.addWidget(ico_title)

        title = QLabel("Catálogo de Insumos")
        f = QFont()
        f.setPointSize(15)
        f.setWeight(QFont.DemiBold)
        title.setFont(f)
        title.setStyleSheet(f"color:{SLATE_700};")
        top.addWidget(title)

        self.lbl_subt = QLabel("0 insumos")
        self.lbl_subt.setStyleSheet(f"color:{SLATE_300}; padding-left:6px;")
        top.addWidget(self.lbl_subt)
        top.addStretch(1)

        self.btn_nuevo = self._mk_btn("Nuevo insumo", primary=True, icon_name="add")
        self.btn_nuevo.clicked.connect(self._nuevo)
        top.addWidget(self.btn_nuevo)

        # Menú Importar (Excel / JSON)
        from PySide6.QtWidgets import QMenu as _QMenu
        self.btn_import = self._mk_btn("Importar ▾", icon_name="importar")
        menu_imp = _QMenu(self.btn_import)
        a_xlsx_i = menu_imp.addAction(icon("xlsx"), "Desde Excel (.xlsx)")
        a_xlsx_i.triggered.connect(self._importar_excel)
        a_json_i = menu_imp.addAction(icon("rep-presupuesto"), "Desde JSON (.json)")
        a_json_i.triggered.connect(self._importar_json)
        self.btn_import.setMenu(menu_imp)
        top.addWidget(self.btn_import)

        # Menú Exportar (Excel / JSON)
        self.btn_export = self._mk_btn("Exportar ▾", icon_name="exportar")
        menu_exp = _QMenu(self.btn_export)
        a_xlsx_e = menu_exp.addAction(icon("xlsx"), "A Excel (.xlsx) — vista filtrada")
        a_xlsx_e.triggered.connect(self._exportar_excel)
        a_json_e = menu_exp.addAction(icon("rep-presupuesto"), "A JSON (.json) — catálogo completo")
        a_json_e.triggered.connect(self._exportar_json)
        self.btn_export.setMenu(menu_exp)
        top.addWidget(self.btn_export)

        layout.addLayout(top)

        # ── Tarjetas KPI ──
        kpis = QHBoxLayout()
        kpis.setSpacing(10)
        self.kpi_total = self._mk_kpi("Insumos", "0", SLATE_500)
        self.kpi_mo = self._mk_kpi("Mano de Obra", "0", TIPO_TEXTO['MO'])
        self.kpi_mat = self._mk_kpi("Materiales", "0", TIPO_TEXTO['MAT'])
        self.kpi_eq = self._mk_kpi("Equipos", "0", TIPO_TEXTO['EQ'])
        self.kpi_sc = self._mk_kpi("Sub-contratos", "0", TIPO_TEXTO['SC'])
        from utils.theme import accent_color as _acc
        self.kpi_valor = self._mk_kpi("Valor catálogo", fmt(0), _acc())
        for k in (self.kpi_total, self.kpi_mo, self.kpi_mat,
                  self.kpi_eq, self.kpi_sc, self.kpi_valor):
            kpis.addWidget(k, 1)
        layout.addLayout(kpis)

        # ── Filtros ──
        filt_card = QFrame()
        filt_card.setStyleSheet(
            f"QFrame {{ background:white; border:1px solid {SILVER_300}; border-radius:6px; }}"
        )
        fl = QHBoxLayout(filt_card)
        fl.setContentsMargins(10, 8, 10, 8)
        fl.setSpacing(8)

        ico_search = QLabel()
        ico_search.setPixmap(icon("buscar").pixmap(18, 18))
        ico_search.setFixedSize(20, 20)
        fl.addWidget(ico_search)
        self.inp_q = QLineEdit()
        self.inp_q.setPlaceholderText("Buscar por descripción o código…")
        self.inp_q.setClearButtonEnabled(True)
        self._timer_q = QTimer(self)
        self._timer_q.setSingleShot(True)
        self._timer_q.timeout.connect(self.cargar)
        self.inp_q.textChanged.connect(lambda _: self._timer_q.start(250))
        fl.addWidget(self.inp_q, 2)

        self.cmb_tipo = QComboBox()
        self.cmb_tipo.addItem("Todos los tipos", "")
        for t in TIPOS_RECURSO:
            self.cmb_tipo.addItem(TIPO_LARGO[t], t)
        self.cmb_tipo.currentIndexChanged.connect(self.cargar)
        fl.addWidget(self.cmb_tipo)

        self.cmb_inei = QComboBox()
        self.cmb_inei.addItem("Todos los índices INEI", "")
        for cod, desc in INEI_CATALOG:
            self.cmb_inei.addItem(f"{cod} — {desc}", cod)
        self.cmb_inei.currentIndexChanged.connect(self.cargar)
        fl.addWidget(self.cmb_inei)

        self.cmb_uso = QComboBox()
        self.cmb_uso.addItem("Todos", "")
        self.cmb_uso.addItem("En uso", "usados")
        self.cmb_uso.addItem("Sin uso", "no_usados")
        self.cmb_uso.currentIndexChanged.connect(self.cargar)
        fl.addWidget(self.cmb_uso)

        self.btn_limpiar = self._mk_btn("Limpiar", icon_name="limpiar")
        self.btn_limpiar.clicked.connect(self._limpiar_filtros)
        fl.addWidget(self.btn_limpiar)

        layout.addWidget(filt_card)

        # ── Tabla ──
        self.tbl = QTableWidget(0, 7)
        self.tbl.setHorizontalHeaderLabels(
            ["Código", "Tipo", "Descripción", "Unidad", "Precio", "INEI", "Usos"]
        )
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)  # solo Precio editable (controlado)
        self.tbl.setSortingEnabled(True)
        self.tbl.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tbl.customContextMenuRequested.connect(self._menu_contextual)
        self.tbl.cellDoubleClicked.connect(self._on_double_click)
        self.tbl.itemChanged.connect(self._on_item_changed)
        self.tbl.setItemDelegateForColumn(1, _TipoBadgeDelegate(self.tbl))
        # Hint visual: doble clic edita, clic derecho menú
        self.tbl.setToolTip("Doble clic: editar  ·  Clic derecho: menú  ·  Del: eliminar")

        h = self.tbl.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.Fixed)
        h.resizeSection(0, 100)
        h.setSectionResizeMode(1, QHeaderView.Fixed)
        h.resizeSection(1, 70)
        h.setSectionResizeMode(2, QHeaderView.Stretch)
        h.setSectionResizeMode(3, QHeaderView.Fixed)
        h.resizeSection(3, 70)
        h.setSectionResizeMode(4, QHeaderView.Fixed)
        h.resizeSection(4, 110)
        h.setSectionResizeMode(5, QHeaderView.Fixed)
        h.resizeSection(5, 60)
        h.setSectionResizeMode(6, QHeaderView.Fixed)

        layout.addWidget(self.tbl, 1)

        # ── Atajos ──
        QShortcut(QKeySequence("Ctrl+N"), self, activated=self._nuevo)
        QShortcut(QKeySequence("Delete"), self, activated=self._eliminar_seleccion)
        QShortcut(QKeySequence("F5"), self, activated=self.cargar)
        QShortcut(QKeySequence("Ctrl+F"), self, activated=lambda: self.inp_q.setFocus())

    def _mk_btn(self, text: str, primary: bool = False,
                icon_name: str | None = None) -> QPushButton:
        b = QPushButton(text)
        b.setCursor(Qt.PointingHandCursor)
        b.setMinimumHeight(32)
        if icon_name:
            b.setIcon(icon(icon_name))
            b.setIconSize(QSize(18, 18))
        if primary:
            b.setStyleSheet(
                f"QPushButton {{ background:{ORANGE_500}; color:white; border:none; "
                f"border-radius:6px; padding:6px 14px; font-weight:600; }}"
                f"QPushButton:hover {{ background:{BLUE_700}; }}"
            )
        return b

    def _mk_kpi(self, etiqueta: str, valor: str, color: str) -> QFrame:
        from utils.theme import apply_shadow
        card = QFrame()
        card.setObjectName("kpiCard")
        card.setAttribute(Qt.WA_StyledBackground, True)
        card.setStyleSheet(
            f"QFrame#kpiCard {{ background:white; border:1px solid {SILVER_300}; "
            f"border-radius:8px; }}"
        )
        apply_shadow(card, 'sm')
        v = QVBoxLayout(card)
        v.setContentsMargins(14, 10, 14, 10)
        v.setSpacing(2)
        l_e = QLabel(etiqueta)
        l_e.setStyleSheet(
            f"color:{SLATE_300}; font-size:11px; letter-spacing:0.4px; "
            f"background:transparent; border:none;"
        )
        l_v = QLabel(valor)
        f = QFont()
        f.setPointSize(14)
        f.setWeight(QFont.DemiBold)
        l_v.setFont(f)
        l_v.setStyleSheet(f"color:{color}; background:transparent; border:none;")
        v.addWidget(l_e)
        v.addWidget(l_v)
        # accesores
        card.lbl_etiqueta = l_e
        card.lbl_valor = l_v
        return card

    # -- carga / consulta ------------------------------------------------------
    def cargar(self):
        from utils.formatting import norm_busqueda
        q = norm_busqueda(self.inp_q.text().strip()) if hasattr(self, 'inp_q') else ''
        tipo = self.cmb_tipo.currentData() if hasattr(self, 'cmb_tipo') else ''
        inei = self.cmb_inei.currentData() if hasattr(self, 'cmb_inei') else ''
        uso = self.cmb_uso.currentData() if hasattr(self, 'cmb_uso') else ''

        sql = (
            "SELECT r.id, r.codigo, r.descripcion, r.tipo, r.unidad, r.precio, "
            "       r.indice_inei, "
            "       (SELECT COUNT(*) FROM acu_items ai WHERE ai.recurso_id=r.id) AS usos "
            "FROM recursos r WHERE 1=1"
        )
        params: list = []
        if q:
            sql += " AND (_norm(r.descripcion) LIKE ? OR _norm(r.codigo) LIKE ?)"
            like = f"%{q}%"
            params += [like, like]
        if tipo:
            sql += " AND r.tipo = ?"
            params.append(tipo)
        if inei:
            sql += " AND (r.indice_inei = ? OR (COALESCE(r.indice_inei,'')='' AND SUBSTR(r.codigo,1,2)=?))"
            params += [inei, inei]
        if uso == 'usados':
            sql += " AND EXISTS (SELECT 1 FROM acu_items ai WHERE ai.recurso_id=r.id)"
        elif uso == 'no_usados':
            sql += " AND NOT EXISTS (SELECT 1 FROM acu_items ai WHERE ai.recurso_id=r.id)"
        sql += " ORDER BY r.tipo, r.codigo, r.descripcion"

        conn = get_db()
        conn.create_function("_norm", 1, norm_busqueda)
        rows = conn.execute(sql, params).fetchall()
        # KPIs globales (no afectados por filtros) — un único query
        tot = conn.execute(
            "SELECT tipo, COUNT(*) AS n, COALESCE(SUM(precio),0) AS valor "
            "FROM recursos GROUP BY tipo"
        ).fetchall()
        conn.close()

        # ── Llenar tabla (optimizado para 7K+ filas) ──
        # Cachear colores/fuentes una sola vez
        col_codigo = QColor(SLATE_500)
        col_muted = QColor(SLATE_300)
        font_mono = QFont("monospace")
        font_pre = QFont()
        font_pre.setWeight(QFont.DemiBold)
        flag_no_edit = Qt.ItemIsSelectable | Qt.ItemIsEnabled
        flag_edit = flag_no_edit | Qt.ItemIsEditable
        align_center = Qt.AlignCenter
        align_right = Qt.AlignRight | Qt.AlignVCenter

        self.tbl.setUpdatesEnabled(False)
        self.tbl.blockSignals(True)
        try:
            self.tbl.setSortingEnabled(False)
            self.tbl.clearContents()
            self.tbl.setRowCount(len(rows))
            self._row_ids = []
            for row, r in enumerate(rows):
                self._row_ids.append(r['id'])

                it_cod = QTableWidgetItem(r['codigo'] or '—')
                it_cod.setFont(font_mono)
                it_cod.setForeground(col_codigo)
                it_cod.setFlags(flag_no_edit)
                it_cod.setData(Qt.UserRole, r['id'])
                self.tbl.setItem(row, 0, it_cod)

                it_tipo = QTableWidgetItem(r['tipo'] or '')
                it_tipo.setTextAlignment(align_center)
                it_tipo.setFlags(flag_no_edit)
                self.tbl.setItem(row, 1, it_tipo)

                it_desc = QTableWidgetItem(r['descripcion'] or '')
                it_desc.setFlags(flag_no_edit)
                self.tbl.setItem(row, 2, it_desc)

                it_und = QTableWidgetItem(r['unidad'] or '')
                it_und.setTextAlignment(align_center)
                it_und.setFlags(flag_no_edit)
                self.tbl.setItem(row, 3, it_und)

                precio = float(r['precio'] or 0)
                it_pre = QTableWidgetItem(fmt(precio))
                it_pre.setData(Qt.UserRole, precio)
                it_pre.setTextAlignment(align_right)
                es_overhead = (r['unidad'] or '').startswith('%')
                if es_overhead:
                    it_pre.setFlags(flag_no_edit)
                    it_pre.setForeground(col_muted)
                else:
                    it_pre.setFlags(flag_edit)
                    it_pre.setFont(font_pre)
                self.tbl.setItem(row, 4, it_pre)

                cod_inei = r['indice_inei']
                it_inei = QTableWidgetItem(cod_inei or '')
                it_inei.setTextAlignment(align_center)
                it_inei.setFlags(flag_no_edit)
                self.tbl.setItem(row, 5, it_inei)

                usos = r['usos'] or 0
                it_usos = QTableWidgetItem(str(usos) if usos else '—')
                it_usos.setTextAlignment(align_center)
                it_usos.setFlags(flag_no_edit)
                if usos == 0:
                    it_usos.setForeground(col_muted)
                self.tbl.setItem(row, 6, it_usos)

            self.tbl.setSortingEnabled(True)
        finally:
            self.tbl.blockSignals(False)
            self.tbl.setUpdatesEnabled(True)

        # ── KPIs ──
        n_total = sum(r['n'] for r in tot) or 0
        n_mo = next((r['n'] for r in tot if r['tipo'] == 'MO'), 0)
        n_mat = next((r['n'] for r in tot if r['tipo'] == 'MAT'), 0)
        n_eq = next((r['n'] for r in tot if r['tipo'] == 'EQ'), 0)
        n_sc = next((r['n'] for r in tot if r['tipo'] == 'SC'), 0)
        valor = sum(r['valor'] for r in tot) or 0.0
        self.kpi_total.lbl_valor.setText(str(n_total))
        self.kpi_mo.lbl_valor.setText(str(n_mo))
        self.kpi_mat.lbl_valor.setText(str(n_mat))
        self.kpi_eq.lbl_valor.setText(str(n_eq))
        self.kpi_sc.lbl_valor.setText(str(n_sc))
        self.kpi_valor.lbl_valor.setText(fmt(valor))

        n_filt = len(rows)
        if n_filt == n_total:
            self.lbl_subt.setText(f"{n_total} insumo{'s' if n_total != 1 else ''}")
        else:
            self.lbl_subt.setText(f"{n_filt} de {n_total} insumos")

    # -- helper: obtener rid robusto a sorting --------------------------------
    def _rid_at(self, row: int) -> int | None:
        if row < 0 or row >= self.tbl.rowCount():
            return None
        it = self.tbl.item(row, 0)
        if not it:
            return None
        v = it.data(Qt.UserRole)
        return int(v) if v is not None else None

    # -- edición inline del precio --------------------------------------------
    def _on_item_changed(self, item: QTableWidgetItem):
        if item.column() != 4:
            return
        rid = self._rid_at(item.row())
        if rid is None:
            return
        nuevo = parse_num(item.text())
        anterior = float(item.data(Qt.UserRole) or 0)
        if abs(nuevo - anterior) < 0.0001:
            self.tbl.blockSignals(True)
            item.setText(fmt(anterior))
            self.tbl.blockSignals(False)
            return
        conn = get_db()
        conn.execute("UPDATE recursos SET precio=? WHERE id=?", (nuevo, rid))
        conn.commit()
        conn.close()
        self.tbl.blockSignals(True)
        item.setData(Qt.UserRole, nuevo)
        item.setText(fmt(nuevo))
        self.tbl.blockSignals(False)

    def _on_double_click(self, row: int, col: int):
        if col == 4:
            it = self.tbl.item(row, col)
            if it and (it.flags() & Qt.ItemIsEditable):
                self.tbl.editItem(it)
            return
        rid = self._rid_at(row)
        if rid is not None:
            self._editar_id(rid)

    # -- CRUD ------------------------------------------------------------------
    def _nuevo(self):
        dlg = RecursoFormDialog(self)
        if dlg.exec() == QDialog.Accepted:
            self.cargar()

    def _editar_id(self, rid: int):
        conn = get_db()
        r = conn.execute(
            "SELECT id, codigo, descripcion, tipo, unidad, precio, indice_inei "
            "FROM recursos WHERE id=?", (rid,)
        ).fetchone()
        conn.close()
        if not r:
            return
        dlg = RecursoFormDialog(self, recurso=dict(r))
        if dlg.exec() == QDialog.Accepted:
            self.cargar()

    def _duplicar_id(self, rid: int):
        conn = get_db()
        r = conn.execute(
            "SELECT codigo, descripcion, tipo, unidad, precio, indice_inei "
            "FROM recursos WHERE id=?", (rid,)
        ).fetchone()
        if not r:
            conn.close()
            return
        inei = r['indice_inei'] or (r['codigo'][:2] if r['codigo'] else '39')
        nuevo_cod = _siguiente_codigo_inei(conn, inei)
        conn.execute(
            "INSERT INTO recursos (codigo, descripcion, tipo, unidad, precio, indice_inei) "
            "VALUES (?,?,?,?,?,?)",
            (nuevo_cod, r['descripcion'] + " (copia)", r['tipo'],
             r['unidad'], r['precio'], inei)
        )
        conn.commit()
        conn.close()
        self.cargar()

    def _eliminar_id(self, rid: int):
        self._eliminar_ids([rid])

    def _eliminar_seleccion(self):
        rows = sorted({i.row() for i in self.tbl.selectedIndexes()})
        ids = [self._rid_at(r) for r in rows]
        ids = [i for i in ids if i is not None]
        if ids:
            self._eliminar_ids(ids)

    def _eliminar_ids(self, ids: list[int]):
        if not ids:
            return
        conn = get_db()
        try:
            placeholders = ','.join('?' * len(ids))
            usados = conn.execute(
                f"SELECT r.id, r.descripcion, COUNT(ai.id) AS n "
                f"FROM recursos r LEFT JOIN acu_items ai ON ai.recurso_id = r.id "
                f"WHERE r.id IN ({placeholders}) GROUP BY r.id HAVING n > 0",
                ids
            ).fetchall()
            if usados:
                detalle = "\n".join(f"  • {u['descripcion']}  ({u['n']} usos)" for u in usados[:5])
                extra = f"\n  …y {len(usados) - 5} más" if len(usados) > 5 else ""
                QMessageBox.warning(
                    self, "No se puede eliminar",
                    f"{len(usados)} recurso(s) están en uso por análisis de costos:\n\n"
                    f"{detalle}{extra}\n\nElimine primero los ítems del ACU."
                )
                return
            n = len(ids)
            msg = (f"¿Eliminar {n} insumos del catálogo?" if n > 1
                   else "¿Eliminar este insumo del catálogo?")
            res = QMessageBox.question(self, "Confirmar eliminación", msg,
                                       QMessageBox.Yes | QMessageBox.No)
            if res != QMessageBox.Yes:
                return
            conn.execute(f"DELETE FROM recursos WHERE id IN ({placeholders})", ids)
            conn.commit()
        finally:
            conn.close()
        self.cargar()

    # -- menú contextual -------------------------------------------------------
    def _menu_contextual(self, pos):
        idx = self.tbl.indexAt(pos)
        if not idx.isValid():
            return
        rid = self._rid_at(idx.row())
        if rid is None:
            return
        seleccion = sorted({i.row() for i in self.tbl.selectedIndexes()})
        ids_sel = [self._rid_at(r) for r in seleccion]
        ids_sel = [i for i in ids_sel if i is not None]

        m = QMenu(self)
        a_edit = QAction(icon("editar"), "Editar", self)
        a_edit.triggered.connect(lambda: self._editar_id(rid))
        m.addAction(a_edit)
        a_dup = QAction(icon("duplicar"), "Duplicar", self)
        a_dup.triggered.connect(lambda: self._duplicar_id(rid))
        m.addAction(a_dup)
        m.addSeparator()
        if len(ids_sel) > 1:
            a_del = QAction(icon("eliminar"), f"Eliminar {len(ids_sel)} seleccionados", self)
            a_del.triggered.connect(lambda: self._eliminar_ids(ids_sel))
        else:
            a_del = QAction(icon("eliminar"), "Eliminar", self)
            a_del.triggered.connect(lambda: self._eliminar_id(rid))
        m.addAction(a_del)
        m.exec(self.tbl.viewport().mapToGlobal(pos))

    # -- filtros ---------------------------------------------------------------
    def _limpiar_filtros(self):
        self.inp_q.blockSignals(True)
        self.inp_q.clear()
        self.inp_q.blockSignals(False)
        self.cmb_tipo.setCurrentIndex(0)
        self.cmb_inei.setCurrentIndex(0)
        self.cmb_uso.setCurrentIndex(0)
        self.cargar()

    # -- import / export Excel -------------------------------------------------
    def _exportar_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.warning(self, "openpyxl no instalado",
                                "Instale openpyxl para exportar:\n\npip install openpyxl")
            return

        ruta, _ = QFileDialog.getSaveFileName(
            self, "Exportar catálogo de insumos",
            "catalogo_insumos.xlsx", "Excel (*.xlsx)"
        )
        if not ruta:
            return
        if not ruta.lower().endswith(".xlsx"):
            ruta += ".xlsx"

        # Exporta lo que esté visible (respeta filtros actuales)
        filas = []
        for row in range(self.tbl.rowCount()):
            rid = self._rid_at(row)
            if rid:
                filas.append(rid)

        if not filas:
            QMessageBox.information(self, "Sin datos", "No hay insumos para exportar.")
            return

        conn = get_db()
        placeholders = ','.join('?' * len(filas))
        rows = conn.execute(
            f"SELECT codigo, descripcion, tipo, unidad, precio, indice_inei "
            f"FROM recursos WHERE id IN ({placeholders}) "
            f"ORDER BY tipo, codigo, descripcion",
            filas
        ).fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Insumos"
        head_fill = PatternFill("solid", fgColor="485A6C")
        head_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="D4D4D4")
        bd = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        right = Alignment(horizontal="right", vertical="center")

        headers = ["Código", "Tipo", "Descripción", "Unidad", "Precio", "Índice INEI"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = center
            cell.border = bd

        for i, r in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=r['codigo'] or '').border = bd
            tcell = ws.cell(row=i, column=2, value=r['tipo'] or '')
            tcell.alignment = center
            tcell.border = bd
            ws.cell(row=i, column=3, value=r['descripcion'] or '').border = bd
            ucell = ws.cell(row=i, column=4, value=r['unidad'] or '')
            ucell.alignment = center
            ucell.border = bd
            pcell = ws.cell(row=i, column=5, value=float(r['precio'] or 0))
            pcell.number_format = '#,##0.00'
            pcell.alignment = right
            pcell.border = bd
            icell = ws.cell(row=i, column=6, value=r['indice_inei'] or '')
            icell.alignment = center
            icell.border = bd

        widths = [12, 8, 60, 10, 14, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"

        try:
            wb.save(ruta)
        except PermissionError:
            QMessageBox.critical(self, "Error",
                                 "No se pudo escribir el archivo. ¿Está abierto en Excel?")
            return
        QMessageBox.information(self, "Exportado",
                                f"Catálogo exportado:\n{ruta}\n\n{len(rows)} insumos.")

    def _importar_excel(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            QMessageBox.warning(self, "openpyxl no instalado",
                                "Instale openpyxl para importar:\n\npip install openpyxl")
            return

        ruta, _ = QFileDialog.getOpenFileName(
            self, "Importar catálogo de insumos", "", "Excel (*.xlsx *.xls)"
        )
        if not ruta:
            return
        try:
            wb = load_workbook(ruta, read_only=True, data_only=True)
        except Exception as e:
            QMessageBox.critical(self, "Error",
                                 f"No se pudo abrir el archivo:\n{e}")
            return
        ws = wb.active
        # Buscar fila de cabecera (primera no vacía)
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            QMessageBox.warning(self, "Vacío", "El archivo no tiene datos.")
            return

        # Mapear por nombre — tolerante a mayúsculas/espacios
        def norm(s: str) -> str:
            return (s or '').strip().lower().replace('í', 'i').replace('ó', 'o')

        idx = {norm(h): i for i, h in enumerate(header)}
        col_cod = idx.get('codigo', idx.get('código', -1))
        col_desc = idx.get('descripcion', idx.get('descripción', -1))
        col_tipo = idx.get('tipo', -1)
        col_und = idx.get('unidad', -1)
        col_pre = idx.get('precio', -1)
        col_inei = idx.get('indice inei', idx.get('índice inei', idx.get('inei', -1)))

        if col_desc < 0 or col_tipo < 0:
            QMessageBox.warning(self, "Cabecera incorrecta",
                                "El Excel debe tener al menos las columnas:\n"
                                "Código · Tipo · Descripción · Unidad · Precio · Índice INEI")
            return

        creados = 0
        actualizados = 0
        errores = 0
        conn = get_db()
        try:
            for fila in rows_iter:
                if not fila or all(c is None or str(c).strip() == '' for c in fila):
                    continue
                desc = (str(fila[col_desc]).strip() if col_desc >= 0 and fila[col_desc] is not None else '')
                tipo = (str(fila[col_tipo]).strip().upper() if col_tipo >= 0 and fila[col_tipo] is not None else '')
                if not desc or tipo not in TIPOS_RECURSO:
                    errores += 1
                    continue
                cod = (str(fila[col_cod]).strip() if col_cod >= 0 and fila[col_cod] is not None else '')
                und = (str(fila[col_und]).strip() if col_und >= 0 and fila[col_und] is not None else '')
                pre_raw = fila[col_pre] if col_pre >= 0 else 0
                pre = parse_num(str(pre_raw)) if pre_raw is not None else 0.0
                inei = (str(fila[col_inei]).strip().zfill(2)[:2]
                        if col_inei >= 0 and fila[col_inei] is not None
                        else (cod[:2] if len(cod) >= 2 else INEI_DEFAULT.get(tipo, '39')))

                if not cod:
                    cod = _siguiente_codigo_inei(conn, inei)

                # Buscar duplicado por código exacto
                ex = conn.execute("SELECT id FROM recursos WHERE codigo=?", (cod,)).fetchone()
                if ex:
                    conn.execute(
                        "UPDATE recursos SET descripcion=?, tipo=?, unidad=?, precio=?, "
                        "indice_inei=? WHERE id=?",
                        (desc, tipo, und, pre, inei, ex['id'])
                    )
                    actualizados += 1
                else:
                    conn.execute(
                        "INSERT INTO recursos (codigo, descripcion, tipo, unidad, precio, indice_inei) "
                        "VALUES (?,?,?,?,?,?)",
                        (cod, desc, tipo, und, pre, inei)
                    )
                    creados += 1
            conn.commit()
        finally:
            conn.close()
        self.cargar()
        QMessageBox.information(
            self, "Importación completa",
            f"{creados} nuevos\n{actualizados} actualizados"
            + (f"\n{errores} ignorados (datos inválidos)" if errores else "")
        )

    # ── Import / Export JSON ───────────────────────────────────────────────
    def _exportar_json(self):
        from datetime import datetime as _dt
        from core.catalogos_json import exportar_recursos_json
        fecha = _dt.now().strftime("%Y%m%d_%H%M")
        sugerido = f"catalogo_insumos_{fecha}.json"
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportar catálogo a JSON", sugerido, "JSON (*.json)"
        )
        if not path:
            return
        if not path.lower().endswith(".json"):
            path += ".json"
        try:
            n = exportar_recursos_json(path)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo exportar:\n{e}")
            return
        QMessageBox.information(
            self, "Exportado",
            f"{n} recursos exportados a:\n{path}"
        )

    def _importar_json(self):
        from core.catalogos_json import importar_recursos_json
        path, _ = QFileDialog.getOpenFileName(
            self, "Importar catálogo desde JSON", "", "JSON (*.json)"
        )
        if not path:
            return

        # Preguntar modo
        from PySide6.QtWidgets import QMessageBox as _QMB
        res = _QMB.question(
            self, "Modo de importación",
            "¿Cómo manejar los recursos que ya existen (por código)?\n\n"
            "• Sí (Merge): actualiza los existentes con los datos del JSON\n"
            "• No (Solo nuevos): ignora los existentes, solo agrega los nuevos",
            _QMB.Yes | _QMB.No | _QMB.Cancel
        )
        if res == _QMB.Cancel:
            return
        modo = 'merge' if res == _QMB.Yes else 'solo_nuevos'

        try:
            r = importar_recursos_json(path, modo=modo)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo importar:\n{e}")
            return
        if not r['ok']:
            QMessageBox.warning(self, "Importar", r.get('msg'))
            return
        self.cargar()
        QMessageBox.information(
            self, "Importación completa",
            f"{r['msg']}\nTotal en archivo: {r['n_total']}"
        )
